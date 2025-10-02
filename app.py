# app.py
# Este é o arquivo principal da sua aplicação Flask.

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    send_from_directory,
    jsonify,
    send_file,
    has_request_context,
    g,
)
import psycopg2
from psycopg2 import extras, sql
from psycopg2.errors import DuplicateObject
import os
import posixpath
from datetime import datetime, timedelta, date
import calendar
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import subprocess
import json
import re
from collections import OrderedDict
from werkzeug.utils import secure_filename  # Para lidar com nomes de arquivos de upload
from decimal import Decimal, InvalidOperation
import io
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Importa a configuração do banco de dados e outras variáveis
from config import DATABASE_URL, SECRET_KEY, UPLOAD_FOLDER, ALLOWED_EXTENSIONS
from caixa_banco import init_app as init_caixa_banco, db
from contas_receber import init_app as init_contas_receber
from cobranca import init_app as init_cobrancas
from cobranca.models import Cobranca
from contas_receber.models import ContaReceber, Pessoa
from caixa_banco.models import (
    ContaCaixa,
    ContaBanco,
    Conciliacao,
    MovimentoFinanceiro,
    PosicaoDiaria,
)
from caixa_banco.services import (
    criar_movimento,
    importar_cnab,
    atualizar_movimento,
    deletar_movimento,
    recalcular_posicoes,
)
from sqlalchemy import func
from sqlalchemy.orm import load_only

app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Cria as pastas de uploads se elas não existirem
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Pastas de uploads utilizadas pelos módulos
os.makedirs(os.path.join(UPLOAD_FOLDER, "backups"), exist_ok=True)
os.makedirs(
    os.path.join(UPLOAD_FOLDER, "imoveis_anexos"), exist_ok=True
)  # Pasta para anexos de imóveis
os.makedirs(
    os.path.join(UPLOAD_FOLDER, "contratos_anexos"), exist_ok=True
)  # Pasta para anexos de contratos

# ------------------ Utilidades para Modelos/Placeholders ------------------
PLACEHOLDER_PATTERN = re.compile(r"\[([^\[\]]+)\]")


def _normalize_key(key: str) -> str:
    return re.sub(r"\s+", "", key.strip().lower())


def login_required(f):
    """Decorador simples para garantir que o usuário esteja autenticado."""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Você precisa estar logado para acessar esta página.", "info")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def permission_required(module, action):
    """Decorador para validar se o usuário autenticado possui a permissão."""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "user_id" not in session:
                flash("Você precisa estar logado para acessar esta página.", "info")
                return redirect(url_for("login"))

            user_id = session["user_id"]
            conn = get_db_connection()
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            try:
                # Usuários do tipo "Master" possuem todas as permissões
                cur.execute("SELECT tipo_usuario FROM usuarios WHERE id = %s", (user_id,))
                user = cur.fetchone()
                is_master = bool(user and user["tipo_usuario"] == "Master")

                has_permission = True
                if not is_master:
                    # Verifica se há permissão específica para o módulo/ação
                    cur.execute(
                        """
                        SELECT COUNT(*) FROM permissoes
                        WHERE usuario_id = %s AND modulo = %s AND acao = %s
                        """,
                        (user_id, module, action),
                    )
                    has_permission = cur.fetchone()[0] > 0
            finally:
                cur.close()
                conn.close()

            if not has_permission:
                flash(
                    f"Você não tem permissão para realizar esta ação no módulo {module}.",
                    "danger",
                )
                return redirect(url_for("dashboard"))

            error = None
            try:
                response = f(*args, **kwargs)
                return response
            except Exception as exc:
                error = exc
                raise
            finally:
                if (
                    error is None
                    and has_request_context()
                    and session.get("user_id")
                    and not getattr(g, "_user_action_logged", False)
                    and request.method not in ("GET", "HEAD", "OPTIONS")
                    and action.lower() not in {"consultar", "visualizar"}
                ):
                    payload = _collect_audit_payload()
                    descricao_auto = _build_auto_log_description(
                        module,
                        action,
                        request.view_args or {},
                        payload,
                    )
                    log_user_action(action, module, descricao_auto)

        return decorated_function

    return decorator


def format_currency(value):
    """Formata valores monetários no padrão brasileiro: R$ 1.234,56.
    Aceita Decimal, float, int ou string numérica.
    """
    try:
        if isinstance(value, Decimal):
            d = value
        else:
            d = Decimal(str(value or 0))
    except Exception:
        d = Decimal("0")
    s = f"{d:,.2f}"
    # Converte 1,234.56 -> 1.234,56
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def build_contrato_context(cur, contrato_id: int) -> dict:
    """Monta um dicionário com dados do contrato, inquilino, imóvel e empresa.
    As chaves são normalizadas para facilitar matching de placeholders.
    """
    ctx = {}

    # Contrato + Imóvel
    cur.execute(
        """
        SELECT c.*, i.endereco AS imovel_endereco, i.bairro AS imovel_bairro,
               i.cidade AS imovel_cidade, i.estado AS imovel_estado,
               i.cep AS imovel_cep, i.matricula, i.inscricao_iptu,
               i.tipo_imovel AS imovel_tipo_imovel
        FROM contratos_aluguel c
        JOIN imoveis i ON c.imovel_id = i.id
        WHERE c.id = %s
        """,
        (contrato_id,),
    )
    contrato = cur.fetchone()
    if not contrato:
        return ctx

    # Empresa licenciada (primeira ativa)
    cur.execute(
        "SELECT documento, razao_social_nome, nome_fantasia, endereco, bairro, cidade, estado, cep, telefone FROM empresa_licenciada WHERE status = 'Ativo' ORDER BY id LIMIT 1"
    )
    empresa = cur.fetchone()

    # Pessoa (cliente) para Documento (CPF/CNPJ)
    cur.execute("SELECT * FROM pessoas WHERE id = %s", (contrato["cliente_id"],))
    pessoa = cur.fetchone()

    def put(k: str, v):
        if v is None:
            v = ""
        ctx[_normalize_key(k)] = str(v)

    # Contrato
    put("ContratoId", contrato["id"])
    put("Finalidade", contrato["finalidade"])
    put("DataInicio", contrato["data_inicio"].strftime("%d/%m/%Y") if contrato["data_inicio"] else "")
    put("DataFim", contrato["data_fim"].strftime("%d/%m/%Y") if contrato["data_fim"] else "")
    put("QuantidadeParcelas", contrato.get("quantidade_parcelas"))
    put("ValorParcela", contrato.get("valor_parcela"))
    put("QuantidadeCalcao", contrato.get("quantidade_calcao"))
    put("ValorCalcao", contrato.get("valor_calcao"))
    put("StatusContrato", contrato.get("status_contrato"))
    put("Observacao", contrato.get("observacao"))

    # Inquilino (snapshot no contrato)
    put("Cliente", contrato.get("nome_inquilino"))
    put("NomeInquilino", contrato.get("nome_inquilino"))
    put("EnderecoCliente", contrato.get("endereco_inquilino"))
    put("BairroCliente", contrato.get("bairro_inquilino"))
    put("CidadeCliente", contrato.get("cidade_inquilino"))
    put("EstadoCliente", contrato.get("estado_inquilino"))
    put("CEPCliente", contrato.get("cep_inquilino"))
    put("TelefoneCliente", contrato.get("telefone_inquilino"))

    # Documento do cliente (CPF/CNPJ)
    if pessoa:
        put("CPF", pessoa.get("documento"))
        put("CPF_CNPJ", pessoa.get("documento"))
        put("DocumentoCliente", pessoa.get("documento"))
        put("NacionalidadeCliente", pessoa.get("nacionalidade"))
        put("EstadoCivilCliente", pessoa.get("estado_civil"))
        put("ProfissaoCliente", pessoa.get("profissao"))
        put("RGCliente", pessoa.get("rg"))
        responsavel_cpf = pessoa.get("responsavel_cpf")
        if responsavel_cpf:
            responsavel_cpf = "".join(ch for ch in responsavel_cpf if ch.isdigit())
        responsavel_uf = pessoa.get("responsavel_uf")
        if responsavel_uf:
            responsavel_uf = responsavel_uf.upper()
        put("ResponsavelNome", pessoa.get("responsavel_nome"))
        put("ResponsavelCPF", responsavel_cpf)
        put("ResponsavelEndereco", pessoa.get("responsavel_endereco"))
        put("ResponsavelBairro", pessoa.get("responsavel_bairro"))
        put("ResponsavelCidade", pessoa.get("responsavel_cidade"))
        put("ResponsavelEstado", pessoa.get("responsavel_estado"))
        put("ResponsavelUF", responsavel_uf)
        put("ResponsavelEstadoCivil", pessoa.get("responsavel_estado_civil"))

    # Imóvel
    put("EnderecoImovel", contrato.get("imovel_endereco"))
    put("BairroImovel", contrato.get("imovel_bairro"))
    put("CidadeImovel", contrato.get("imovel_cidade"))
    put("EstadoImovel", contrato.get("imovel_estado"))
    put("CEPImovel", contrato.get("imovel_cep"))
    put("MatriculaImovel", contrato.get("matricula"))
    put("InscricaoIPTU", contrato.get("inscricao_iptu"))
    put("TipoImovel", contrato.get("imovel_tipo_imovel"))

    # Campos derivados solicitados para modelos de contrato
    # 1) Mês e ano do início do contrato (ex: "Janeiro de 2025")
    data_inicio_dt = contrato.get("data_inicio")
    data_fim_dt = contrato.get("data_fim")
    if data_inicio_dt and data_fim_dt and data_fim_dt >= data_inicio_dt:
        months_diff = (data_fim_dt.year - data_inicio_dt.year) * 12 + (data_fim_dt.month - data_inicio_dt.month)
        if data_fim_dt.day > data_inicio_dt.day:
            months_diff += 1
        months_diff = max(months_diff, 0)
        put("DuracaoContratoMeses", months_diff)
    else:
        put("DuracaoContratoMeses", "")
    if data_inicio_dt:
        meses_pt = [
            "",
            "Janeiro",
            "Fevereiro",
            "Março",
            "Abril",
            "Maio",
            "Junho",
            "Julho",
            "Agosto",
            "Setembro",
            "Outubro",
            "Novembro",
            "Dezembro",
        ]
        put("MesEAnoInicioContrato", f"{meses_pt[data_inicio_dt.month]} de {data_inicio_dt.year}")
        # Data de início por extenso, com mês em minúsculas (ex: "10 de janeiro de 2025")
        meses_pt_lower = [
            "",
            "janeiro",
            "fevereiro",
            "março",
            "abril",
            "maio",
            "junho",
            "julho",
            "agosto",
            "setembro",
            "outubro",
            "novembro",
            "dezembro",
        ]
        put(
            "DataInicioExtenso",
            f"{data_inicio_dt.day} de {meses_pt_lower[data_inicio_dt.month]} de {data_inicio_dt.year}",
        )
    else:
        put("MesEAnoInicioContrato", "")
        put("DataInicioExtenso", "")

    # 2) Dia do vencimento e 3) Data do Primeiro Vencimento
    # Considera apenas parcelas de aluguel (ignora títulos de calção que começam com 'C')
    cur.execute(
        """
        SELECT MIN(data_vencimento) AS primeiro
          FROM contas_a_receber
         WHERE contrato_id = %s
           AND (titulo IS NULL OR titulo NOT LIKE 'C%%')
        """,
        (contrato_id,),
    )
    row_first = cur.fetchone()
    primeiro_venc = row_first["primeiro"] if row_first and row_first.get("primeiro") else None
    if primeiro_venc:
        put("DiaVencimento", str(primeiro_venc.day))
        put("DataPrimeiroVencimento", primeiro_venc.strftime("%d/%m/%Y"))
    else:
        put("DiaVencimento", "")
        put("DataPrimeiroVencimento", "")

    # Empresa licenciada
    if empresa:
        put("EmpresaNome", empresa.get("razao_social_nome"))
        put("EmpresaFantasia", empresa.get("nome_fantasia"))
        put("EmpresaDocumento", empresa.get("documento"))
        put("EmpresaEndereco", empresa.get("endereco"))
        put("EmpresaBairro", empresa.get("bairro"))
        put("EmpresaCidade", empresa.get("cidade"))
        put("EmpresaEstado", empresa.get("estado"))
        put("EmpresaCEP", empresa.get("cep"))
        put("EmpresaTelefone", empresa.get("telefone"))

    return ctx


def render_placeholders(html: str, ctx: dict) -> str:
    """Substitui [Placeholders] pelo valor do contexto (case-insensitive)."""
    if not html:
        return html

    def repl(match):
        raw = match.group(1)
        key_norm = _normalize_key(raw)
        return ctx.get(key_norm, f"[{raw}]")

    return PLACEHOLDER_PATTERN.sub(repl, html)

# Inicializa o módulo de Caixa e Banco (SQLAlchemy e rotas REST)
init_caixa_banco(app)
init_contas_receber(app)
init_cobrancas(app)

# Variáveis globais para o sistema (exemplo)
SYSTEM_VERSION = "1.0"

# Módulos disponíveis no sistema para configuração de permissões
MODULES = [
    "Cadastro Fornecedores/Clientes",
    "Cadastro Imoveis",
    "Cadastro Despesas",
    "Cadastro Origens",
    "Cadastro Receitas",
    "Gestao Contratos",
    "Financeiro",
    "Administracao Sistema",
]

# Ações possíveis
ACTIONS = ["Incluir", "Editar", "Consultar", "Excluir", "Bloquear"]

PRESTACAO_RECEITA_DESCR = "PRESTACAO ENCERRAMENTO CONTRATO"
PRESTACAO_DESPESA_DESCR = "PRESTACAO ENCERRAMENTO CONTRATO"


# Função para conectar ao banco de dados
def get_db_connection():
    conn = psycopg2.connect(DATABASE_URL)
    return conn


def ensure_status_contrato_enum_values():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        conn.autocommit = True
        cur = conn.cursor()
        for value in ("Finalizado", "Renovar"):
            cur.execute(f"ALTER TYPE status_contrato_enum ADD VALUE IF NOT EXISTS '{value}'")
    except Exception:
        pass
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


ensure_status_contrato_enum_values()


def ensure_auditoria_table():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS auditoria_logs (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                username VARCHAR(150),
                modulo VARCHAR(150),
                acao VARCHAR(50),
                descricao TEXT,
                ip VARCHAR(45),
                criado_em TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        conn.commit()
        # Remove coluna "dados" de instalações antigas
        cur.execute("ALTER TABLE auditoria_logs DROP COLUMN IF EXISTS dados")
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        app.logger.exception("Erro ao garantir tabela de auditoria: %s", e)
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


ensure_auditoria_table()


def _get_request_ip():
    if not has_request_context():
        return None
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr


def _sanitize_audit_value(value, *, max_length=200):
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        if len(cleaned) > max_length:
            return cleaned[: max_length - 3] + "..."
        return cleaned
    if isinstance(value, (int, float, Decimal)):
        return value
    if isinstance(value, (list, tuple, set)):
        sanitized = [
            _sanitize_audit_value(v, max_length=max_length)
            for v in value
        ]
        sanitized = [v for v in sanitized if v is not None]
        if not sanitized:
            return None
        return sanitized
    if value is None:
        return None
    return _sanitize_audit_value(str(value), max_length=max_length)


def _collect_audit_payload():
    payload = OrderedDict()
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return payload

    sensitive_keys = {"senha", "password", "password_confirm", "confirm_password"}

    json_payload = request.get_json(silent=True) if request.is_json else None
    if isinstance(json_payload, dict):
        for key, value in json_payload.items():
            if not isinstance(key, str):
                continue
            sanitized = _sanitize_audit_value(value)
            if sanitized is not None:
                payload[key] = sanitized
    elif json_payload is not None:
        sanitized = _sanitize_audit_value(json_payload)
        if sanitized is not None:
            payload["_json"] = sanitized

    if request.form:
        for key in request.form:
            key_lower = key.lower()
            if key_lower in sensitive_keys or "senha" in key_lower:
                continue
            if key_lower == "csrf_token":
                continue
            values = request.form.getlist(key)
            sanitized_values = [
                _sanitize_audit_value(v)
                for v in values
            ]
            sanitized_values = [v for v in sanitized_values if v is not None]
            if not sanitized_values:
                continue
            payload[key] = (
                sanitized_values[0]
                if len(sanitized_values) == 1
                else sanitized_values
            )

    if request.files:
        for key in request.files:
            arquivos = request.files.getlist(key)
            nomes = [
                getattr(arquivo, "filename", "")
                for arquivo in arquivos
            ]
            nomes = [nome for nome in nomes if nome]
            if not nomes:
                continue
            payload[f"arquivo:{key}"] = nomes[0] if len(nomes) == 1 else nomes

    return payload


def _stringify_audit_value(value, max_length=200):
    if isinstance(value, list):
        parts = [
            _stringify_audit_value(item, max_length=max_length)
            for item in value
            if item is not None
        ]
        parts = [part for part in parts if part]
        return ", ".join(parts)
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    text = str(value).strip()
    if len(text) > max_length:
        text = text[: max_length - 3] + "..."
    return text


def _format_audit_label(key: str) -> str:
    base = key
    lower = base.lower()
    for prefix in ("old_", "anterior_", "novo_", "new_"):
        if lower.startswith(prefix):
            base = base[len(prefix) :]
            lower = base.lower()
            break
    for suffix in ("_old", "_anterior", "_novo", "_new"):
        if lower.endswith(suffix):
            base = base[: -len(suffix)]
            lower = base.lower()
            break
    label = re.sub(r"[_\-]+", " ", base).strip()
    if not label:
        label = key
    lowered = label.lower()
    if lowered in {"cpf", "cnpj"}:
        return label.upper()
    return label.title()


def _split_audit_key_role(key: str):
    key_lower = key.lower()
    for prefix in ("old_", "anterior_"):
        if key_lower.startswith(prefix):
            return key[len(prefix) :], "old"
    for suffix in ("_old", "_anterior"):
        if key_lower.endswith(suffix):
            return key[: -len(suffix)], "old"
    for prefix in ("novo_", "new_"):
        if key_lower.startswith(prefix):
            return key[len(prefix) :], "new"
    for suffix in ("_novo", "_new"):
        if key_lower.endswith(suffix):
            return key[: -len(suffix)], "new"
    return key, "value"


def _summarize_audit_payload(payload: OrderedDict):
    if not payload:
        return ""

    groups = {}
    for key, value in payload.items():
        base, role = _split_audit_key_role(key)
        groups.setdefault(base, {})[role] = {"key": key, "value": value}

    handled = set()
    summaries = []

    for base, info in groups.items():
        if "old" in info and "new" in info:
            old_value = _stringify_audit_value(info["old"]["value"])
            new_value = _stringify_audit_value(info["new"]["value"])
            if old_value or new_value:
                label = _format_audit_label(base)
                old_repr = json.dumps(old_value, ensure_ascii=False)
                new_repr = json.dumps(new_value, ensure_ascii=False)
                summaries.append(f"{label}: {old_repr} -> {new_repr}")
            handled.add(info["old"]["key"])
            handled.add(info["new"]["key"])

    for base, info in groups.items():
        for entry in info.values():
            key = entry["key"]
            if key in handled:
                continue
            key_lower = key.lower()
            if key_lower in {"csrf_token", "submit"}:
                continue
            if "senha" in key_lower or "password" in key_lower:
                continue
            value_text = _stringify_audit_value(entry["value"])
            if not value_text:
                continue
            label = _format_audit_label(key)
            value_repr = json.dumps(value_text, ensure_ascii=False)
            summaries.append(f"{label} = {value_repr}")
            handled.add(key)

    return "; ".join(summaries)


def _find_matching_value(source: OrderedDict, keywords):
    if not source:
        return None
    for key, value in source.items():
        key_lower = key.lower()
        if any(keyword in key_lower for keyword in keywords):
            if isinstance(value, list):
                for item in value:
                    if item is not None:
                        text = _stringify_audit_value(item)
                        if text:
                            return text
            else:
                text = _stringify_audit_value(value)
                if text:
                    return text
    return None


def _extract_audit_identifier(view_args, payload):
    candidates = []
    if view_args:
        candidates.append(view_args.items())
    if payload:
        candidates.append(payload.items())
    for items in candidates:
        for key, value in items:
            if not isinstance(key, str):
                continue
            key_lower = key.lower()
            if "id" in key_lower or "codigo" in key_lower or "código" in key_lower or key_lower.endswith("numero") or key_lower.endswith("número"):
                if isinstance(value, list):
                    for item in value:
                        text = _stringify_audit_value(item)
                        if text:
                            return text
                else:
                    text = _stringify_audit_value(value)
                    if text:
                        return text
    return None


ACTION_VERB_MAP = {
    "incluir": "adicionado",
    "inclusao": "adicionado",
    "adicionar": "adicionado",
    "cadastrar": "cadastrado",
    "criar": "criado",
    "editar": "editado",
    "alterar": "alterado",
    "atualizar": "atualizado",
    "ajustar": "ajustado",
    "excluir": "excluído",
    "remover": "removido",
    "deletar": "deletado",
    "apagar": "apagado",
    "bloquear": "bloqueado",
    "desbloquear": "desbloqueado",
    "aprovar": "aprovado",
    "reprovar": "reprovado",
    "gerar": "gerado",
}


def _build_auto_log_description(module, action, view_args, payload):
    action_lower = (action or "").strip().lower()
    verbo = ACTION_VERB_MAP.get(action_lower, action or "Ação executada")
    modulo_nome = module or "Recurso"

    identificador = _extract_audit_identifier(view_args, payload)
    nome_principal = _find_matching_value(
        payload,
        [
            "nome",
            "razao",
            "descr",
            "titulo",
            "pessoa",
            "cliente",
            "fornecedor",
            "contrat",
        ],
    )
    documento = _find_matching_value(payload, ["cpf", "cnpj", "documento"])

    partes = [f"{modulo_nome} {verbo}".strip()]
    if identificador:
        partes[-1] += f" ID {identificador}"
    if nome_principal:
        if documento:
            partes[-1] += f" para {nome_principal} ({documento})"
        else:
            partes[-1] += f" para {nome_principal}"
    elif documento:
        partes[-1] += f" ({documento})"

    resumo_campos = _summarize_audit_payload(payload)
    detalhes_tecnicos = []

    if resumo_campos:
        partes.append(f"Campos: {resumo_campos}")

    if request.endpoint:
        detalhes_tecnicos.append(f"endpoint={request.endpoint}")
    if request.path:
        detalhes_tecnicos.append(f"path={request.path}")
    if request.view_args:
        args_resumo = {
            chave: _stringify_audit_value(valor)
            for chave, valor in request.view_args.items()
        }
        detalhes_tecnicos.append(
            "parâmetros="
            + json.dumps(args_resumo, ensure_ascii=False)
        )
    if request.args:
        query_resumo = {
            chave: request.args.getlist(chave)
            if len(request.args.getlist(chave)) > 1
            else request.args.get(chave)
            for chave in request.args
        }
        detalhes_tecnicos.append(
            "query=" + json.dumps(query_resumo, ensure_ascii=False)
        )

    descricao = ". ".join(partes)
    if detalhes_tecnicos:
        descricao += ". Contexto: " + "; ".join(detalhes_tecnicos)
    return descricao


def log_user_action(acao, modulo, descricao=None, dados=None):
    if not has_request_context():
        return
    if "user_id" not in session:
        return
    setattr(g, "_user_action_logged", True)
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        descricao_final = descricao or ""
        if dados is not None:
            try:
                dados_serializados = json.dumps(dados, ensure_ascii=False)
            except (TypeError, ValueError):
                dados_serializados = str(dados)
            if descricao_final:
                descricao_final = f"{descricao_final}\nDados: {dados_serializados}"
            else:
                descricao_final = f"Dados: {dados_serializados}"
        if not descricao_final:
            descricao_final = None
        cur.execute(
            """
            INSERT INTO auditoria_logs (user_id, username, modulo, acao, descricao, ip)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                session.get("user_id"),
                session.get("username"),
                modulo,
                acao,
                descricao_final,
                _get_request_ip(),
            ),
        )
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        app.logger.exception("Erro ao registrar log de auditoria: %s", e)
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


# --- Gerencial ---


@app.route("/gerencial/logs")
@login_required
@permission_required("Administracao Sistema", "Visualizar")
def gerencial_logs():
    page = request.args.get("page", 1, type=int)
    if page < 1:
        page = 1
    per_page = 50

    username = request.args.get("username", "").strip()
    modulo = request.args.get("modulo", "").strip()
    acao = request.args.get("acao", "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()

    filtros_sql = []
    params = []

    if username:
        filtros_sql.append("LOWER(username) LIKE %s")
        params.append(f"%{username.lower()}%")
    if modulo:
        filtros_sql.append("modulo = %s")
        params.append(modulo)
    if acao:
        filtros_sql.append("acao = %s")
        params.append(acao)

    if data_inicio:
        try:
            filtros_sql.append("criado_em >= %s")
            params.append(datetime.strptime(data_inicio, "%Y-%m-%d"))
        except ValueError:
            flash("Data inicial inválida. Utilize o formato AAAA-MM-DD.", "warning")

    if data_fim:
        try:
            filtros_sql.append("criado_em < %s")
            params.append(datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            flash("Data final inválida. Utilize o formato AAAA-MM-DD.", "warning")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    where_clause = ""
    if filtros_sql:
        where_clause = " WHERE " + " AND ".join(filtros_sql)

    count_sql = "SELECT COUNT(*) FROM auditoria_logs" + where_clause
    cur.execute(count_sql, params)
    total = cur.fetchone()[0] or 0

    total_pages = (total + per_page - 1) // per_page if total else 1
    if page > total_pages:
        page = total_pages if total_pages > 0 else 1

    offset = (page - 1) * per_page if total else 0

    select_sql = (
        "SELECT id, user_id, username, modulo, acao, descricao, ip, criado_em "
        "FROM auditoria_logs"
        + where_clause
        + " ORDER BY criado_em DESC LIMIT %s OFFSET %s"
    )
    query_params = list(params)
    query_params.extend([per_page, offset])
    cur.execute(select_sql, query_params)
    rows = cur.fetchall()

    logs = []
    for row in rows:
        logs.append(
            {
                "id": row["id"],
                "user_id": row["user_id"],
                "username": row["username"],
                "modulo": row["modulo"],
                "acao": row["acao"],
                "descricao": row["descricao"],
                "ip": row["ip"],
                "criado_em": row["criado_em"],
            }
        )

    cur.execute(
        "SELECT DISTINCT modulo FROM auditoria_logs WHERE modulo IS NOT NULL AND modulo <> '' ORDER BY modulo"
    )
    modulos = [row[0] for row in cur.fetchall()]

    cur.execute(
        "SELECT DISTINCT acao FROM auditoria_logs WHERE acao IS NOT NULL AND acao <> '' ORDER BY acao"
    )
    acoes = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    query_args = {
        "username": username or None,
        "modulo": modulo or None,
        "acao": acao or None,
        "data_inicio": data_inicio or None,
        "data_fim": data_fim or None,
    }
    query_args = {k: v for k, v in query_args.items() if v}

    pagination = {
        "page": page,
        "pages": total_pages,
        "per_page": per_page,
        "total": total,
        "has_prev": page > 1,
        "has_next": page < total_pages and total > 0,
        "prev_url": url_for("gerencial_logs", page=page - 1, **query_args) if page > 1 else None,
        "next_url": url_for("gerencial_logs", page=page + 1, **query_args)
        if page < total_pages and total > 0
        else None,
    }

    filter_values = {
        "username": username,
        "modulo": modulo,
        "acao": acao,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
    }

    return render_template(
        "gerencial/logs/index.html",
        logs=logs,
        modulos=modulos,
        acoes=acoes,
        filter_values=filter_values,
        pagination=pagination,
    )


# Função auxiliar para verificar extensões de arquivo permitidas
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# Converte valores decimais enviados pelo formulário, aceitando vírgulas ou
# pontos como separadores de milhares e decimais. Retorna None se o valor
# estiver vazio ou não puder ser convertido em número.
def parse_decimal(value):
    if value is None:
        return None
    value = str(value).strip()
    if value == "":
        return None

    # Remove quaisquer símbolos de moeda e caracteres não numéricos,
    # preservando apenas dígitos, vírgula, ponto e sinal de menos
    value = re.sub(r"[^0-9,\.\-]", "", value)

    # Se contiver tanto "," quanto ".", assume que o separador decimal é o
    # último caracter dentre eles e remove os demais como separadores de
    # milhares.
    if "," in value and "." in value:
        if value.rfind(",") > value.rfind("."):
            value = value.replace(".", "").replace(",", ".")
        else:
            value = value.replace(",", "")
    elif "," in value:
        value = value.replace(".", "").replace(",", ".")
    else:
        value = value.replace(",", "")

    try:
        return Decimal(value)
    except (ValueError, InvalidOperation):
        return None


def parse_int(value):
    """Converte valores para inteiro ou retorna None se vazio."""
    if value is None:
        return None
    value = str(value).strip()
    if value == "" or value.lower() == "none":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_date(value):
    """Converte uma string no formato ISO (YYYY-MM-DD) para date."""
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def decimal_to_string(value) -> str:
    if value is None:
        value = Decimal("0")
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return format(value.quantize(Decimal("0.01")), ".2f")


def get_or_create_categoria(cur, tabela: str, descricao: str) -> int:
    cur.execute(f"SELECT id FROM {tabela} WHERE descricao = %s", (descricao,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        f"INSERT INTO {tabela} (descricao) VALUES (%s) RETURNING id",
        (descricao,),
    )
    return cur.fetchone()[0]


def fetch_contrato_info(cur, contrato_id: int):
    cur.execute(
        """
        SELECT c.*, p.razao_social_nome AS cliente_nome, p.documento AS cliente_documento,
               p.endereco AS cliente_endereco, p.bairro AS cliente_bairro,
               p.cidade AS cliente_cidade, p.estado AS cliente_estado,
               p.cep AS cliente_cep,
               i.endereco AS imovel_endereco, i.bairro AS imovel_bairro,
               i.cidade AS imovel_cidade, i.estado AS imovel_estado
        FROM contratos_aluguel c
        JOIN pessoas p ON c.cliente_id = p.id
        JOIN imoveis i ON c.imovel_id = i.id
        WHERE c.id = %s
        """,
        (contrato_id,),
    )
    return cur.fetchone()


def listar_creditos_calcao(cur, contrato_id: int):
    cur.execute(
        """
        SELECT id, titulo, data_pagamento, valor_pago, valor_previsto
        FROM contas_a_receber
        WHERE contrato_id = %s
          AND titulo LIKE 'C%%'
          AND status_conta = 'Paga'
        ORDER BY data_pagamento NULLS LAST, id
        """,
        (contrato_id,),
    )
    creditos = []
    for row in cur.fetchall():
        valor_pago = row["valor_pago"]
        if valor_pago is None:
            continue
        valor_pago = Decimal(str(valor_pago))
        if valor_pago <= 0:
            continue
        creditos.append(
            {
                "id": row["id"],
                "titulo": row["titulo"],
                "data_pagamento": row["data_pagamento"],
                "valor": valor_pago,
                "valor_previsto": Decimal(str(row["valor_previsto"]))
                if row["valor_previsto"] is not None
                else valor_pago,
            }
        )
    return creditos


def listar_debitos_para_prestacao(cur, contrato_id: int, data_encerramento: date):
    cur.execute(
        """
        SELECT id, titulo, data_vencimento, data_pagamento, valor_previsto, valor_pago, valor_pendente, status_conta
        FROM contas_a_receber
        WHERE contrato_id = %s
          AND status_conta IN ('Aberta', 'Vencida', 'Parcial')
          AND (data_vencimento IS NULL OR data_vencimento >= %s)
        ORDER BY data_vencimento NULLS LAST, id
        """,
        (contrato_id, data_encerramento),
    )
    debitos = []
    for row in cur.fetchall():
        valor_previsto = Decimal(str(row["valor_previsto"]))
        valor_pago = (
            Decimal(str(row["valor_pago"]))
            if row["valor_pago"] is not None
            else Decimal("0")
        )
        valor_pendente = (
            Decimal(str(row["valor_pendente"]))
            if row["valor_pendente"] is not None
            else valor_previsto - valor_pago
        )
        if valor_pendente <= 0:
            continue
        debitos.append(
            {
                "id": row["id"],
                "titulo": row["titulo"],
                "data_vencimento": row["data_vencimento"],
                "valor_previsto": valor_previsto,
                "valor_pago": valor_pago,
                "valor_pendente": valor_pendente,
                "status": row["status_conta"],
                "data_pagamento": row["data_pagamento"],
            }
        )
    return debitos


def calcular_totais_prestacao(creditos, debitos, despesas, creditos_extras=None):
    creditos_extras = creditos_extras or []
    total_creditos_base = sum(item["valor"] for item in creditos)
    total_creditos_extras = sum(item["valor"] for item in creditos_extras)
    total_creditos = total_creditos_base + total_creditos_extras
    total_debitos = sum(item["valor_pendente"] for item in debitos)
    total_despesas = sum(item["valor"] for item in despesas)
    saldo_final = total_creditos - total_debitos - total_despesas
    return {
        "total_creditos": total_creditos,
        "total_debitos": total_debitos,
        "total_despesas": total_despesas,
        "saldo_final": saldo_final,
    }


def preparar_creditos_extras(payload):
    creditos = []
    if not payload:
        return creditos
    for item in payload:
        if not isinstance(item, dict):
            continue
        descricao = str(item.get("descricao", "")).strip()
        valor = parse_decimal(item.get("valor"))
        if not descricao or valor is None:
            continue
        try:
            valor = valor.quantize(Decimal("0.01"))
        except InvalidOperation:
            continue
        if valor <= 0:
            continue
        creditos.append({"descricao": descricao, "valor": valor})
    return creditos


def preparar_despesas(payload):
    despesas = []
    if not payload:
        return despesas
    for item in payload:
        if not isinstance(item, dict):
            continue
        descricao = str(item.get("descricao", "")).strip()
        valor = parse_decimal(item.get("valor"))
        if not descricao or valor is None:
            continue
        try:
            valor = valor.quantize(Decimal("0.01"))
        except InvalidOperation:
            continue
        if valor <= 0:
            continue
        despesas.append({"descricao": descricao, "valor": valor})
    return despesas


def reverter_prestacao(cur, prestacao_id: int):
    cur.execute(
        "SELECT conta_pagar_id, conta_receber_id FROM prestacoes_contas WHERE id = %s",
        (prestacao_id,),
    )
    prestacao = cur.fetchone()
    if prestacao is None:
        raise ValueError("Prestacao de contas não encontrada.")

    cur.execute(
        """
        SELECT conta_receber_id, status_original, valor_pago_original, valor_pendente_original, data_pagamento_original
        FROM prestacoes_contas_itens
        WHERE prestacao_id = %s AND tipo = 'debito'
        """,
        (prestacao_id,),
    )
    for row in cur.fetchall():
        conta_id = row["conta_receber_id"]
        if conta_id is None:
            continue
        status_original = row["status_original"] or "Aberta"
        cur.execute(
            """
            UPDATE contas_a_receber
            SET status_conta = %s,
                valor_pago = %s,
                valor_pendente = %s,
                data_pagamento = %s
            WHERE id = %s
            """,
            (
                status_original,
                row["valor_pago_original"],
                row["valor_pendente_original"],
                row["data_pagamento_original"],
                conta_id,
            ),
        )

    conta_pagar_id = prestacao["conta_pagar_id"]
    conta_receber_id = prestacao["conta_receber_id"]
    if conta_pagar_id:
        cur.execute("DELETE FROM contas_a_pagar WHERE id = %s", (conta_pagar_id,))
    if conta_receber_id:
        cur.execute("DELETE FROM contas_a_receber WHERE id = %s", (conta_receber_id,))


def processar_prestacao(
    cur,
    contrato,
    data_encerramento,
    payload,
    prestacao_id=None,
):
    contrato_id = contrato["id"]
    creditos = listar_creditos_calcao(cur, contrato_id)
    debitos_disponiveis = listar_debitos_para_prestacao(cur, contrato_id, data_encerramento)
    debitos_map = {item["id"]: item for item in debitos_disponiveis}

    debitos_ids = set()
    for conta_id in payload.get("debitos_incluir", []):
        try:
            debitos_ids.add(int(conta_id))
        except (TypeError, ValueError):
            raise ValueError("Identificador de parcela inválido.")

    debitos = []
    for conta_id in debitos_ids:
        debito = debitos_map.get(conta_id)
        if not debito:
            raise ValueError("Parcela selecionada não é válida para a prestação de contas.")
        debitos.append(debito)

    despesas = preparar_despesas(payload.get("despesas"))
    creditos_extras = preparar_creditos_extras(payload.get("creditos_extras"))
    observacoes = (payload.get("observacoes") or "").strip()

    totais = calcular_totais_prestacao(creditos, debitos, despesas, creditos_extras)
    total_creditos = totais["total_creditos"].quantize(Decimal("0.01"))
    total_debitos = totais["total_debitos"].quantize(Decimal("0.01"))
    total_despesas = totais["total_despesas"].quantize(Decimal("0.01"))
    saldo_final = totais["saldo_final"].quantize(Decimal("0.01"))

    if prestacao_id is None:
        cur.execute(
            """
            INSERT INTO prestacoes_contas (
                contrato_id, data_encerramento,
                total_creditos, total_debitos, total_despesas, saldo_final, observacoes
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                contrato_id,
                data_encerramento,
                total_creditos,
                total_debitos,
                total_despesas,
                saldo_final,
                observacoes if observacoes else None,
            ),
        )
        prestacao_id = cur.fetchone()[0]
    else:
        cur.execute(
            """
            UPDATE prestacoes_contas
            SET data_encerramento = %s,
                total_creditos = %s,
                total_debitos = %s,
                total_despesas = %s,
                saldo_final = %s,
                observacoes = %s,
                conta_pagar_id = NULL,
                conta_receber_id = NULL,
                atualizado_em = NOW()
            WHERE id = %s
            """,
            (
                data_encerramento,
                total_creditos,
                total_debitos,
                total_despesas,
                saldo_final,
                observacoes if observacoes else None,
                prestacao_id,
            ),
        )
        if cur.rowcount == 0:
            raise ValueError("Prestação de contas não encontrada para atualização.")
        cur.execute(
            "DELETE FROM prestacoes_contas_itens WHERE prestacao_id = %s",
            (prestacao_id,),
        )

    for credito in creditos:
        cur.execute(
            """
            INSERT INTO prestacoes_contas_itens (
                prestacao_id, conta_receber_id, tipo, descricao, valor
            ) VALUES (%s, %s, 'credito', %s, %s)
            """,
            (
                prestacao_id,
                credito["id"],
                credito["titulo"],
                credito["valor"].quantize(Decimal("0.01")),
            ),
        )

    for credito_extra in creditos_extras:
        cur.execute(
            """
            INSERT INTO prestacoes_contas_itens (
                prestacao_id, conta_receber_id, tipo, descricao, valor
            ) VALUES (%s, %s, 'credito_extra', %s, %s)
            """,
            (
                prestacao_id,
                None,
                credito_extra["descricao"],
                credito_extra["valor"].quantize(Decimal("0.01")),
            ),
        )

    for debito in debitos:
        cur.execute(
            """
            INSERT INTO prestacoes_contas_itens (
                prestacao_id, conta_receber_id, tipo, descricao, valor,
                status_original, valor_pago_original, valor_pendente_original, data_pagamento_original
            ) VALUES (%s, %s, 'debito', %s, %s, %s, %s, %s, %s)
            """,
            (
                prestacao_id,
                debito["id"],
                debito["titulo"],
                debito["valor_pendente"].quantize(Decimal("0.01")),
                debito["status"],
                debito["valor_pago"],
                debito["valor_pendente"],
                debito["data_pagamento"],
            ),
        )
        cur.execute(
            """
            UPDATE contas_a_receber
            SET status_conta = 'Paga',
                valor_pago = %s,
                valor_pendente = 0,
                data_pagamento = %s
            WHERE id = %s
            """,
            (
                debito["valor_previsto"],
                data_encerramento,
                debito["id"],
            ),
        )

    for despesa in despesas:
        cur.execute(
            """
            INSERT INTO prestacoes_contas_itens (prestacao_id, tipo, descricao, valor)
            VALUES (%s, 'despesa', %s, %s)
            """,
            (
                prestacao_id,
                despesa["descricao"],
                despesa["valor"].quantize(Decimal("0.01")),
            ),
        )

    conta_pagar_id = None
    conta_receber_id = None
    titulo = f"Prestação de contas contrato #{contrato_id}"
    if saldo_final > 0:
        despesa_id = get_or_create_categoria(cur, "despesas_cadastro", PRESTACAO_DESPESA_DESCR)
        cur.execute(
            """
            INSERT INTO contas_a_pagar (
                despesa_id, fornecedor_id, titulo, data_vencimento, competencia,
                valor_previsto, valor_pendente, observacao, status_conta
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Aberta')
            RETURNING id
            """,
            (
                despesa_id,
                contrato["cliente_id"],
                titulo,
                data_encerramento,
                data_encerramento,
                saldo_final,
                saldo_final,
                observacoes if observacoes else None,
            ),
        )
        conta_pagar_id = cur.fetchone()[0]
    elif saldo_final < 0:
        receita_id = get_or_create_categoria(cur, "receitas_cadastro", PRESTACAO_RECEITA_DESCR)
        valor_receber = abs(saldo_final)
        cur.execute(
            """
            INSERT INTO contas_a_receber (
                contrato_id, receita_id, cliente_id, titulo, data_vencimento,
                valor_previsto, valor_pendente, observacao, status_conta
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Aberta')
            RETURNING id
            """,
            (
                contrato_id,
                receita_id,
                contrato["cliente_id"],
                titulo,
                data_encerramento,
                valor_receber,
                valor_receber,
                observacoes if observacoes else None,
            ),
        )
        conta_receber_id = cur.fetchone()[0]

    cur.execute(
        """
        UPDATE prestacoes_contas
        SET conta_pagar_id = %s,
            conta_receber_id = %s,
            atualizado_em = NOW()
        WHERE id = %s
        """,
        (conta_pagar_id, conta_receber_id, prestacao_id),
    )

    return {
        "prestacao_id": prestacao_id,
        "totais": {
            "total_creditos": total_creditos,
            "total_debitos": total_debitos,
            "total_despesas": total_despesas,
            "saldo_final": saldo_final,
        },
        "conta_pagar_id": conta_pagar_id,
        "conta_receber_id": conta_receber_id,
    }


def format_currency(value):
    try:
        return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "R$ 0,00"


app.jinja_env.filters["currency"] = format_currency


def add_months(date_obj, months):
    """Retorna a data acrescida do número de meses informado."""
    month = date_obj.month - 1 + months
    year = date_obj.year + month // 12
    month = month % 12 + 1
    day = min(date_obj.day, calendar.monthrange(year, month)[1])
    return date_obj.replace(year=year, month=month, day=day)


def calcular_status_conta(data_vencimento, data_pagamento, contrato_id, cur):
    if data_pagamento:
        return "Paga"
    data_venc = datetime.strptime(data_vencimento, "%Y-%m-%d").date()
    status = "Vencida" if data_venc < datetime.today().date() else "Aberta"
    if contrato_id:
        cur.execute(
            "SELECT status_contrato FROM contratos_aluguel WHERE id = %s",
            (contrato_id,),
        )
        contrato = cur.fetchone()
        if (
            contrato
            and contrato.get("status_contrato") in {"Encerrado", "Finalizado", "Renovar"}
            and status == "Aberta"
        ):
            status = "Cancelada"
    return status


def atualizar_status_contas_a_receber(cur):
    cur.execute(
        """
        UPDATE contas_a_receber cr
           SET status_conta = CASE
                WHEN cr.status_conta = 'Cancelada'::status_conta_enum THEN 'Cancelada'::status_conta_enum
                WHEN cr.valor_pago >= cr.valor_previsto AND cr.valor_pago IS NOT NULL THEN 'Paga'::status_conta_enum
                WHEN cr.valor_pago > 0 THEN 'Parcial'::status_conta_enum
                WHEN cr.data_vencimento < CURRENT_DATE THEN 'Vencida'::status_conta_enum
                ELSE 'Aberta'::status_conta_enum
           END,
               valor_pendente = CASE
                   WHEN cr.status_conta = 'Cancelada'::status_conta_enum THEN 0
                   ELSE cr.valor_previsto - COALESCE(cr.valor_pago,0)
               END
        """
    )


def atualizar_status_contas_a_pagar(cur):
    cur.execute(
        """
        UPDATE contas_a_pagar cp
           SET status_conta = CASE
                WHEN cp.data_pagamento IS NOT NULL THEN 'Paga'::status_conta_enum
                WHEN cp.data_vencimento < CURRENT_DATE THEN 'Vencida'::status_conta_enum
                ELSE 'Aberta'::status_conta_enum
           END
        """
    )

def atualizar_status_contratos(cur):
    cur.execute(
        """
        UPDATE contratos_aluguel
           SET status_contrato = 'Renovar'::status_contrato_enum
         WHERE data_fim < CURRENT_DATE
           AND status_contrato IN ('Ativo', 'Pendente')
        """
    )

@app.before_request
def atualizar_contratos_expirados():
    endpoint = request.endpoint or ''
    if endpoint == 'static':
        return
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        atualizar_status_contratos(cur)
        conn.commit()
    except Exception:
        if conn:
            conn.rollback()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

# Garante que a coluna max_contratos exista na tabela imoveis
def ensure_max_contratos_column():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_name='imoveis' AND column_name='max_contratos'
        """
    )
    exists = cur.fetchone() is not None
    if not exists:
        cur.execute(
            "ALTER TABLE imoveis ADD COLUMN max_contratos INTEGER DEFAULT 1"
        )
        conn.commit()
    cur.close()
    conn.close()


def ensure_finalidade_column():
    """Garante que o tipo e a coluna de finalidade existam."""
    conn = get_db_connection()
    cur = conn.cursor()

    # Garante que o tipo enum exista
    cur.execute("SELECT 1 FROM pg_type WHERE typname = 'finalidade_contrato_enum'")
    if cur.fetchone() is None:
        cur.execute(
            "CREATE TYPE finalidade_contrato_enum AS ENUM ('Comercial', 'Residencial', 'Comodato')"
        )

    # Garante que a coluna exista na tabela
    cur.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_name='contratos_aluguel' AND column_name='finalidade'
        """
    )
    exists = cur.fetchone() is not None
    if not exists:
        cur.execute(
            """
            ALTER TABLE contratos_aluguel
            ADD COLUMN finalidade finalidade_contrato_enum NOT NULL DEFAULT 'Residencial'
            """
        )
        cur.execute(
            "ALTER TABLE contratos_aluguel ALTER COLUMN finalidade DROP DEFAULT"
        )

    conn.commit()
    cur.close()
    conn.close()


def ensure_calcao_columns():
    """Garante que as colunas de calção existam em contratos_aluguel."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT column_name FROM information_schema.columns
        WHERE table_name='contratos_aluguel'
          AND column_name IN ('quantidade_calcao', 'valor_calcao')
        """
    )
    existing = {row[0] for row in cur.fetchall()}
    if "quantidade_calcao" not in existing:
        cur.execute(
            "ALTER TABLE contratos_aluguel ADD COLUMN quantidade_calcao INTEGER"
        )
    if "valor_calcao" not in existing:
        cur.execute(
            "ALTER TABLE contratos_aluguel ADD COLUMN valor_calcao NUMERIC(10,2)"
        )
    conn.commit()
    cur.close()
    conn.close()


def ensure_contrato_renovacoes_table():
    """Cria a tabela de histórico de renovações de contrato, se necessário."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS contrato_renovacoes (
            id SERIAL PRIMARY KEY,
            contrato_id INTEGER NOT NULL REFERENCES contratos_aluguel(id) ON DELETE CASCADE,
            data_inicio_anterior DATE,
            data_fim_anterior DATE,
            valor_parcela_anterior NUMERIC(10,2),
            data_inicio_novo DATE NOT NULL,
            data_fim_novo DATE NOT NULL,
            valor_parcela_novo NUMERIC(10,2) NOT NULL,
            observacao TEXT,
            usuario_id INTEGER REFERENCES usuarios(id),
            data_renovacao TIMESTAMP NOT NULL DEFAULT NOW()
        )
        """
    )
    conn.commit()
    cur.close()
    conn.close()


def ensure_tipo_pessoa_enum():
    """Garante que o enum tipo_pessoa_enum contenha 'Cliente/Fornecedor'."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM pg_type WHERE typname = 'tipo_pessoa_enum'")
    if cur.fetchone():
        cur.execute(
            """
            SELECT 1 FROM pg_enum
            WHERE enumtypid = 'tipo_pessoa_enum'::regtype AND enumlabel = 'Cliente/Fornecedor'
            """
        )
        if cur.fetchone() is None:
            cur.execute("ALTER TYPE tipo_pessoa_enum ADD VALUE 'Cliente/Fornecedor'")
            conn.commit()
    cur.close()
    conn.close()


# Assegura colunas necessárias no banco de dados
ensure_max_contratos_column()
ensure_calcao_columns()
ensure_contrato_renovacoes_table()
ensure_tipo_pessoa_enum()
 


def ensure_ordens_pagamento_tables():
    """Cria as tabelas de Ordens de Pagamento (cabecalho e itens) caso não existam."""
    conn = get_db_connection()
    cur = conn.cursor()
    # Tabela principal de ordens de pagamento
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ordens_pagamento (
            id SERIAL PRIMARY KEY,
            data_emissao TIMESTAMP NOT NULL DEFAULT NOW(),
            imovel_id INTEGER REFERENCES imoveis(id),
            fornecedor_id INTEGER NOT NULL REFERENCES pessoas(id),
            forma_pagamento VARCHAR(30),
            data_vencimento DATE,
            descricao_servico TEXT,
            valor_servico NUMERIC(12,2) NOT NULL DEFAULT 0,
            desconto_manual NUMERIC(12,2) NOT NULL DEFAULT 0,
            subtotal_servicos NUMERIC(12,2) NOT NULL DEFAULT 0,
            subtotal_produtos NUMERIC(12,2) NOT NULL DEFAULT 0,
            desconto_produtos NUMERIC(12,2) NOT NULL DEFAULT 0,
            total NUMERIC(12,2) NOT NULL DEFAULT 0,
            observacoes TEXT
        )
        """
    )
    # Garante colunas auxiliares para integração com contas a pagar
    cur.execute(
        """
        SELECT 1 FROM information_schema.columns
         WHERE table_name='ordens_pagamento' AND column_name='despesa_id'
        """
    )
    if cur.fetchone() is None:
        cur.execute(
            "ALTER TABLE ordens_pagamento ADD COLUMN despesa_id INTEGER REFERENCES despesas_cadastro(id)"
        )

    cur.execute(
        """
        SELECT 1 FROM information_schema.columns
         WHERE table_name='ordens_pagamento' AND column_name='origem_id'
        """
    )
    if cur.fetchone() is None:
        cur.execute(
            "ALTER TABLE ordens_pagamento ADD COLUMN origem_id INTEGER REFERENCES origens_cadastro(id)"
        )
    # Itens de produtos da OP
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ordem_pagamento_itens (
            id SERIAL PRIMARY KEY,
            ordem_id INTEGER NOT NULL REFERENCES ordens_pagamento(id) ON DELETE CASCADE,
            codigo_produto VARCHAR(50),
            descricao_produto TEXT,
            quantidade NUMERIC(12,2) NOT NULL DEFAULT 0,
            valor_unitario NUMERIC(12,2) NOT NULL DEFAULT 0,
            valor_desconto NUMERIC(12,2) NOT NULL DEFAULT 0,
            valor_total NUMERIC(12,2) NOT NULL DEFAULT 0
        )
        """
    )
    # Parcelas (títulos) da OP
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ordem_pagamento_parcelas (
            id SERIAL PRIMARY KEY,
            ordem_id INTEGER NOT NULL REFERENCES ordens_pagamento(id) ON DELETE CASCADE,
            numero INTEGER,
            data_vencimento DATE NOT NULL,
            valor NUMERIC(12,2) NOT NULL DEFAULT 0
        )
        """
    )
    # Relacionamento com contas a pagar para evitar duplicidades
    cur.execute(
        "SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = 'contas_a_pagar')"
    )
    if cur.fetchone()[0]:
        cur.execute(
            """
            SELECT 1 FROM information_schema.columns
             WHERE table_name='contas_a_pagar' AND column_name='ordem_pagamento_id'
            """
        )
        if cur.fetchone() is None:
            cur.execute(
                "ALTER TABLE contas_a_pagar ADD COLUMN ordem_pagamento_id INTEGER REFERENCES ordens_pagamento(id)"
            )
    conn.commit()
    


def ensure_dre_tables():
    """Cria as tabelas de DRE caso não existam."""
    conn = get_db_connection()
    cur = conn.cursor()
    # Mascaras
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dre_mascaras (
            id SERIAL PRIMARY KEY,
            nome VARCHAR(255) NOT NULL,
            descricao TEXT,
            ordem INTEGER NOT NULL DEFAULT 0,
            eh_formula BOOLEAN NOT NULL DEFAULT FALSE,
            formula TEXT,
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            data_cadastro TIMESTAMP NOT NULL DEFAULT NOW()
        )
        """
    )
    # Commit imediato para garantir existência da tabela antes de tentar alterar
    conn.commit()
    # Garante coluna 'ordem' para bases já existentes (sem usar IF NOT EXISTS para compatibilidade)
    try:
        cur.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name='dre_mascaras' AND column_name='ordem'"
        )
        tem_ordem = cur.fetchone() is not None
        if not tem_ordem:
            cur.execute("ALTER TABLE dre_mascaras ADD COLUMN ordem INTEGER NOT NULL DEFAULT 0")
    except Exception:
        # Evita transação abortada caso haja erro em ambiente legado
        conn.rollback()

    # Garante colunas de fórmula
    try:
        cur.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name='dre_mascaras' AND column_name='eh_formula'"
        )
        tem_eh_formula = cur.fetchone() is not None
        if not tem_eh_formula:
            cur.execute("ALTER TABLE dre_mascaras ADD COLUMN eh_formula BOOLEAN NOT NULL DEFAULT FALSE")
        cur.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name='dre_mascaras' AND column_name='formula'"
        )
        tem_formula = cur.fetchone() is not None
        if not tem_formula:
            cur.execute("ALTER TABLE dre_mascaras ADD COLUMN formula TEXT")
    except Exception:
        conn.rollback()
    
    # Nós da máscara (hierarquia)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dre_nos (
            id SERIAL PRIMARY KEY,
            mascara_id INTEGER NOT NULL REFERENCES dre_mascaras(id) ON DELETE CASCADE,
            parent_id INTEGER REFERENCES dre_nos(id) ON DELETE CASCADE,
            titulo VARCHAR(255) NOT NULL,
            tipo VARCHAR(20) NOT NULL DEFAULT 'grupo',
            ordem INTEGER NOT NULL DEFAULT 0,
            CONSTRAINT dre_nos_tipo_chk CHECK (tipo IN ('grupo','receita','despesa'))
        )
        """
    )
    # Mapeamentos de categorias de receitas/despesas para nós
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dre_no_receitas (
            no_id INTEGER NOT NULL REFERENCES dre_nos(id) ON DELETE CASCADE,
            receita_id INTEGER NOT NULL REFERENCES receitas_cadastro(id) ON DELETE CASCADE,
            PRIMARY KEY (no_id, receita_id)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dre_no_despesas (
            no_id INTEGER NOT NULL REFERENCES dre_nos(id) ON DELETE CASCADE,
            despesa_id INTEGER NOT NULL REFERENCES despesas_cadastro(id) ON DELETE CASCADE,
            PRIMARY KEY (no_id, despesa_id)
        )
        """
    )
    conn.commit()
    cur.close()
    conn.close()


ensure_dre_tables()
ensure_ordens_pagamento_tables()


def ensure_column_exists(cur, table_name, column_name, definition_sql):
    """Create a column if it is missing.

    Older versions of PostgreSQL used in some environments do not support
    ``ADD COLUMN IF NOT EXISTS``. To keep compatibility we check the
    ``information_schema`` before attempting to add a column. The
    ``definition_sql`` parameter should contain only the column definition
    (type, constraints, defaults, etc.).
    """

    cur.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_name = %s AND column_name = %s
        """,
        (table_name, column_name),
    )
    if cur.fetchone():
        return

    cur.execute(
        sql.SQL("ALTER TABLE {} ADD COLUMN {} {}").format(
            sql.Identifier(table_name),
            sql.Identifier(column_name),
            sql.SQL(definition_sql),
        )
    )



def ensure_prestacao_enum_credito_extra(conn):
    """Ensure the enum prestacao_item_tipo includes the credito_extra value."""
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM pg_type WHERE typname = 'prestacao_item_tipo'")
        type_exists = cur.fetchone() is not None
    if not type_exists:
        return

    with conn.cursor() as cur:
        cur.execute("""
            SELECT 1
            FROM pg_enum e
            JOIN pg_type t ON e.enumtypid = t.oid
            WHERE t.typname = 'prestacao_item_tipo'
              AND e.enumlabel = 'credito_extra'
        """)
        if cur.fetchone():
            return

    conn.commit()
    previous_autocommit = conn.autocommit
    try:
        conn.autocommit = True
        with conn.cursor() as cur:
            try:
                cur.execute("ALTER TYPE prestacao_item_tipo ADD VALUE 'credito_extra'")
            except DuplicateObject:
                pass
    finally:
        conn.autocommit = previous_autocommit



def ensure_prestacao_contas_tables():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_type WHERE typname = 'prestacao_item_tipo'
                ) THEN
                    CREATE TYPE prestacao_item_tipo AS ENUM ('credito', 'debito', 'despesa');
                END IF;
            END$$;
            """
        )
        cur.close()
        # Ensure new enum value exists outside transactional context
        ensure_prestacao_enum_credito_extra(conn)
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS prestacoes_contas (
                id SERIAL PRIMARY KEY,
                contrato_id INTEGER NOT NULL REFERENCES contratos_aluguel(id) ON DELETE CASCADE,
                data_encerramento DATE NOT NULL,
                total_creditos NUMERIC(10,2) NOT NULL DEFAULT 0,
                total_debitos NUMERIC(10,2) NOT NULL DEFAULT 0,
                total_despesas NUMERIC(10,2) NOT NULL DEFAULT 0,
                saldo_final NUMERIC(10,2) NOT NULL DEFAULT 0,
                observacoes TEXT,
                conta_pagar_id INTEGER REFERENCES contas_a_pagar(id),
                conta_receber_id INTEGER REFERENCES contas_a_receber(id),
                criado_em TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                atualizado_em TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS prestacoes_contas_itens (
                id SERIAL PRIMARY KEY,
                prestacao_id INTEGER NOT NULL REFERENCES prestacoes_contas(id) ON DELETE CASCADE,
                conta_receber_id INTEGER REFERENCES contas_a_receber(id),
                tipo prestacao_item_tipo NOT NULL,
                descricao TEXT NOT NULL,
                valor NUMERIC(10,2) NOT NULL DEFAULT 0,
                status_original status_conta_enum,
                valor_pago_original NUMERIC(10,2),
                valor_pendente_original NUMERIC(10,2),
                data_pagamento_original DATE
            )
            """
        )

        for table_name, column_name, definition in [
            (
                "prestacoes_contas",
                "conta_pagar_id",
                "INTEGER REFERENCES contas_a_pagar(id)",
            ),
            (
                "prestacoes_contas",
                "conta_receber_id",
                "INTEGER REFERENCES contas_a_receber(id)",
            ),
            (
                "prestacoes_contas",
                "atualizado_em",
                "TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()",
            ),
            (
                "prestacoes_contas_itens",
                "status_original",
                "status_conta_enum",
            ),
            (
                "prestacoes_contas_itens",
                "valor_pago_original",
                "NUMERIC(10,2)",
            ),
            (
                "prestacoes_contas_itens",
                "valor_pendente_original",
                "NUMERIC(10,2)",
            ),
            (
                "prestacoes_contas_itens",
                "data_pagamento_original",
                "DATE",
            ),
        ]:
            ensure_column_exists(cur, table_name, column_name, definition)
        conn.commit()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


ensure_prestacao_contas_tables()


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    """Serve arquivos enviados pelo usuário."""
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/imoveis/fotos/<int:imovel_id>")
@login_required
def imoveis_fotos(imovel_id):
    """Retorna as URLs das fotos de um imóvel em formato JSON."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        "SELECT nome_arquivo FROM imovel_anexos WHERE imovel_id = %s AND tipo_anexo = 'foto'",
        (imovel_id,),
    )
    fotos = [
        url_for(
            "uploaded_file",
            filename=posixpath.join("imoveis_anexos", row["nome_arquivo"]),
        )
        for row in cur.fetchall()
    ]
    cur.close()
    conn.close()
    return jsonify(fotos)


# Context processor para injetar variáveis em todos os templates
@app.context_processor
def inject_global_vars():
    return {
        "system_version": SYSTEM_VERSION,
        "usuario_logado": session.get("username", "Convidado"),
    }


# --- Rotas de Autenticação ---
@app.route("/login", methods=["GET", "POST"])
def login():
    # Se o usuário já estiver logado, redireciona para o dashboard
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
    
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(
            "SELECT id, nome_usuario, senha_hash, status FROM usuarios WHERE nome_usuario = %s",
            (username,),
        )
        user = cur.fetchone()
        cur.close()
        conn.close()

        if (
            user
            and user["status"] == "Ativo"
            and check_password_hash(user["senha_hash"], password)
        ):
            session["user_id"] = user["id"]
            session["username"] = user["nome_usuario"]
            log_user_action(
                "Acesso",
                "Autenticação",
                "Login realizado com sucesso.",
                {"user_id": user["id"]},
            )
            flash("Login realizado com sucesso!", "success")
            return redirect(
                url_for("dashboard")
            )  # Redireciona para o dashboard após o login
        else:
            flash("Usuário ou senha inválidos, ou usuário inativo.", "danger")
    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    # Se o usuário já estiver logado, redireciona para o dashboard
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:
            flash("As senhas não coincidem.", "danger")
            return render_template("register.html")

        # Hash da senha antes de armazenar
        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            # Insere o novo usuário com tipo "Operador" por padrão
            cur.execute(
                "INSERT INTO usuarios (nome_usuario, senha_hash, tipo_usuario, status) VALUES (%s, %s, %s, %s) RETURNING id",
                (username, hashed_password, "Operador", "Ativo")
            )
            new_user_id = cur.fetchone()[0]
            conn.commit()

            # Opcional: Atribuir permissões básicas para o novo usuário "Operador"
            # Isso pode ser feito de forma mais sofisticada, mas aqui é um exemplo básico.
            # Por exemplo, dar permissão de consulta para alguns módulos.
            modules_to_grant_access = [
                "Cadastro Fornecedores/Clientes",
                "Cadastro Imoveis",
                "Gestao Contratos",
                "Financeiro",
                "Cadastro Despesas",  # Nova permissão
                "Cadastro Origens",  # Nova permissão
                "Cadastro Receitas",  # Nova permissão
            ]
            for module in modules_to_grant_access:
                cur.execute(
                    "INSERT INTO permissoes (usuario_id, modulo, acao) VALUES (%s, %s, %s)",
                    (new_user_id, module, "Consultar")
                )
            conn.commit()

            flash(
                "Usuário cadastrado com sucesso! Você já pode fazer login.", "success"
            )
            return redirect(url_for("login"))
        except psycopg2.errors.UniqueViolation:
            flash("Nome de usuário já existe. Por favor, escolha outro.", "danger")
            conn.rollback()
        except Exception as e:
            flash(f"Erro ao cadastrar usuário: {e}", "danger")
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    return render_template("register.html")


@app.route("/logout")
@login_required
def logout():
    log_user_action(
        "Acesso",
        "Autenticação",
        "Logout realizado.",
    )
    session.pop("user_id", None)
    session.pop("username", None)
    flash("Você foi desconectado.", "info")
    return redirect(url_for("login"))


# --- Rota Principal (Dashboard) ---
@app.route("/") # Redireciona a raiz para o login
def index():
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db_connection()
    cur = conn.cursor()
    atualizar_status_contas_a_receber(cur)
    atualizar_status_contas_a_pagar(cur)
    conn.commit()
    cur.execute(
        "SELECT COUNT(*), COALESCE(SUM(valor_previsto), 0) FROM contas_a_receber WHERE status_conta = 'Vencida'"
    )
    qtd_titulos_atrasados, valor_total_titulos_atrasados = cur.fetchone()
    valor_total_titulos_atrasados = float(valor_total_titulos_atrasados)
    cur.execute(
        "SELECT COUNT(*), COALESCE(SUM(valor_previsto), 0) FROM contas_a_pagar WHERE status_conta = 'Vencida'"
    )
    qtd_titulos_atrasados_pagar, valor_total_titulos_atrasados_pagar = cur.fetchone()
    valor_total_titulos_atrasados_pagar = float(valor_total_titulos_atrasados_pagar)
    cur.execute("SELECT COUNT(*) FROM imoveis")
    total_imoveis_ativos = cur.fetchone()[0]
    cur.execute(
        "SELECT COUNT(*) FROM contratos_aluguel WHERE status_contrato = 'Ativo'"
    )
    total_contratos_ativos = cur.fetchone()[0]
    cur.execute(
        "SELECT COUNT(DISTINCT imovel_id) FROM  contratos_aluguel WHERE status_contrato = 'Ativo'"
    )
    imoveis_com_contrato = cur.fetchone()[0]
    percent_imoveis_alugados = (
        imoveis_com_contrato / total_imoveis_ativos * 100
        if total_imoveis_ativos
        else 0
    )
    cur.execute(
        """
        SELECT c.id,
               c.nome_inquilino,
               c.data_fim,
               COALESCE(i.endereco, '') AS endereco,
               COALESCE(i.bairro, '') AS bairro
        FROM contratos_aluguel c
        LEFT JOIN imoveis i ON c.imovel_id = i.id
        WHERE c.status_contrato = 'Ativo'
          AND c.data_fim BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '15 days'
        ORDER BY c.data_fim ASC
        """
    )
    contratos_vencendo = cur.fetchall()
    hoje = date.today()
    avisos_contratos = []
    for contrato_id, nome_inquilino, data_fim, endereco, bairro in contratos_vencendo:
        if endereco and bairro:
            localizacao = f"{endereco} - {bairro}"
        else:
            localizacao = endereco or bairro or ''
        dias_restantes = (data_fim - hoje).days if data_fim else None
        avisos_contratos.append(
            {
                "id": contrato_id,
                "nome_inquilino": nome_inquilino,
                "data_fim": data_fim.strftime("%d/%m/%Y") if data_fim else '',
                "dias_restantes": dias_restantes,
                "localizacao": localizacao.strip(),
            }
        )
    cur.close()
    conn.close()
    meses_labels = [
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez",
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
    ]
    contratacoes = [2, 3, 1, 4, 2, 3, 1, 2, 4, 3, 2, 1]
    demissoes = [1, 0, 1, 0, 1, 0, 0, 1, 0, 1, 0, 1]

    saldo_caixas = db.session.query(func.coalesce(func.sum(ContaCaixa.saldo_atual), 0)).scalar() or 0
    saldo_bancos = db.session.query(func.coalesce(func.sum(ContaBanco.saldo_atual), 0)).scalar() or 0
    saldo_total = float(saldo_caixas) + float(saldo_bancos)
    alertas_saldo_negativo = ContaCaixa.query.filter(ContaCaixa.saldo_atual < 0).count() + \
        ContaBanco.query.filter(ContaBanco.saldo_atual < 0).count()
    conciliacoes_pendentes = Conciliacao.query.filter_by(status='pendente').count()

    return render_template(
        "dashboard.html",
        qtd_titulos_atrasados=qtd_titulos_atrasados,
        valor_total_titulos_atrasados=valor_total_titulos_atrasados,
        qtd_titulos_atrasados_pagar=qtd_titulos_atrasados_pagar,
        valor_total_titulos_atrasados_pagar=valor_total_titulos_atrasados_pagar,
        total_imoveis_ativos=total_imoveis_ativos,
        total_contratos_ativos=total_contratos_ativos,
        percent_imoveis_alugados=percent_imoveis_alugados,
        meses_labels=meses_labels,
        contratacoes=contratacoes,
        demissoes=demissoes,
        saldo_total=saldo_total,
        alertas_saldo_negativo=alertas_saldo_negativo,
        conciliacoes_pendentes=conciliacoes_pendentes,
        avisos_contratos=avisos_contratos,
    )


# --- Módulo de Cadastros ---


# 1.1. Cadastro de Fornecedores e Clientes (Pessoas)
@app.route("/pessoas")
@login_required
@permission_required("Cadastro Fornecedores/Clientes", "Consultar")
def pessoas_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    status_filter = request.args.get("status", "Ambos")

    base_query = "SELECT * FROM pessoas"
    conditions = []
    params = []

    if search_query:
        conditions.append(
            "(documento ILIKE %s OR razao_social_nome ILIKE %s OR nome_fantasia ILIKE %s)"
        )
        params.extend([
            f"%{search_query}%",
            f"%{search_query}%",
            f"%{search_query}%",
        ])

    if status_filter and status_filter.lower() != "ambos":
        conditions.append("status = %s")
        params.append(status_filter)

    if conditions:
        base_query += " WHERE " + " AND ".join(conditions)

    base_query += " ORDER BY data_cadastro DESC"
    cur.execute(base_query, tuple(params))
    pessoas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "pessoas/list.html",
        pessoas=pessoas,
        search_query=search_query,
        status_filter=status_filter,
    )


@app.route("/pessoas/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Fornecedores/Clientes", "Incluir")
def pessoas_add():
    if request.method == "POST":
        try:
            documento = (
                request.form["documento"]
                .replace(".", "")
                .replace("/", "")
                .replace("-", "")
            )
            razao_social_nome = request.form["razao_social_nome"]
            nome_fantasia = request.form.get("nome_fantasia")
            endereco = request.form.get("endereco")
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep", "").replace("-", "")
            telefone = request.form.get("telefone")
            contato = request.form.get("contato")
            nacionalidade = request.form.get("nacionalidade")
            estado_civil = request.form.get("estado_civil")
            profissao = request.form.get("profissao")
            rg = request.form.get("rg")
            observacao = request.form.get("observacao")
            responsavel_nome = request.form.get("responsavel_nome")
            responsavel_cpf = request.form.get("responsavel_cpf", "").strip()
            if responsavel_cpf:
                responsavel_cpf = "".join(ch for ch in responsavel_cpf if ch.isdigit())
            else:
                responsavel_cpf = None
            responsavel_endereco = request.form.get("responsavel_endereco")
            responsavel_bairro = request.form.get("responsavel_bairro")
            responsavel_cidade = request.form.get("responsavel_cidade")
            responsavel_estado = request.form.get("responsavel_estado")
            responsavel_uf = request.form.get("responsavel_uf")
            if responsavel_uf:
                responsavel_uf = responsavel_uf.upper()
            else:
                responsavel_uf = None
            responsavel_estado_civil = request.form.get("responsavel_estado_civil")
            is_cnpj = len(documento) == 14
            if not is_cnpj:
                responsavel_nome = None
                responsavel_cpf = None
                responsavel_endereco = None
                responsavel_bairro = None
                responsavel_cidade = None
                responsavel_estado = None
                responsavel_uf = None
                responsavel_estado_civil = None
            tipo = request.form["tipo"]
            status = request.form["status"]

            # Validação de CPF/CNPJ (simplificada, você precisaria de uma biblioteca real como "validate-docbr")
            if len(documento) == 11 and not documento.isdigit():
                 flash("CPF inválido. Deve conter apenas números.", "danger")
                 return render_template("pessoas/add_list.html", pessoa={})
            elif len(documento) == 14 and not documento.isdigit():
                 flash("CNPJ inválido. Deve conter apenas números.", "danger")
                 return render_template("pessoas/add_list.html", pessoa={})
            elif len(documento) != 11 and len(documento) != 14:
                flash(
                    "Documento deve ser um CPF (11 dígitos) ou CNPJ (14 dígitos).",
                    "danger",
                )
                return render_template("pessoas/add_list.html", pessoa={})
            
            # Exemplo de uso de validate-docbr (necessita instalação: pip install validate-docbr)
            # from validate_docbr import CPF, CNPJ
            # if len(documento) == 11:
            #     if not CPF().validate(documento):
            #         flash("CPF inválido.", "danger")
            #         return render_template("pessoas/add_list.html", pessoa={})
            # elif len(documento) == 14:
            #     if not CNPJ().validate(documento):
            #         flash("CNPJ inválido.", "danger")
            #         return render_template("pessoas/add_list.html", pessoa={})

            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO pessoas (documento, razao_social_nome, nome_fantasia, endereco, bairro, cidade, estado, cep, telefone, contato, nacionalidade, estado_civil, profissao, rg, observacao, responsavel_nome, responsavel_cpf, responsavel_endereco, responsavel_bairro, responsavel_cidade, responsavel_estado, responsavel_uf, responsavel_estado_civil, tipo, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    documento,
                    razao_social_nome,
                    nome_fantasia,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    telefone,
                    contato,
                    nacionalidade,
                    estado_civil,
                    profissao,
                    rg,
                    observacao,
                    responsavel_nome,
                    responsavel_cpf,
                    responsavel_endereco,
                    responsavel_bairro,
                    responsavel_cidade,
                    responsavel_estado,
                    responsavel_uf,
                    responsavel_estado_civil,
                    tipo,
                    status,
                ),
            )
            conn.commit()
            cur.close()
            conn.close()
            flash("Pessoa cadastrada com sucesso!", "success")
            return redirect(url_for("pessoas_list"))
        except psycopg2.errors.UniqueViolation:
            flash("Erro: Documento (CPF/CNPJ) já cadastrado.", "danger")
            conn.rollback()  # Garante que a transação seja desfeita em caso de erro
            return render_template("pessoas/add_list.html", pessoa=request.form)
        except Exception as e:
            flash(f"Erro ao cadastrar pessoa: {e}", "danger")
            return render_template("pessoas/add_list.html", pessoa=request.form)
    return render_template("pessoas/add_list.html", pessoa={})


@app.route("/pessoas/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Fornecedores/Clientes", "Editar")
def pessoas_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            documento = (
                request.form["documento"]
                .replace(".", "")
                .replace("/", "")
                .replace("-", "")
            )
            razao_social_nome = request.form["razao_social_nome"]
            nome_fantasia = request.form.get("nome_fantasia")
            endereco = request.form.get("endereco")
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep", "").replace("-", "")
            telefone = request.form.get("telefone")
            contato = request.form.get("contato")
            nacionalidade = request.form.get("nacionalidade")
            estado_civil = request.form.get("estado_civil")
            profissao = request.form.get("profissao")
            rg = request.form.get("rg")
            observacao = request.form.get("observacao")
            responsavel_nome = request.form.get("responsavel_nome")
            responsavel_cpf = request.form.get("responsavel_cpf", "").strip()
            if responsavel_cpf:
                responsavel_cpf = "".join(ch for ch in responsavel_cpf if ch.isdigit())
            else:
                responsavel_cpf = None
            responsavel_endereco = request.form.get("responsavel_endereco")
            responsavel_bairro = request.form.get("responsavel_bairro")
            responsavel_cidade = request.form.get("responsavel_cidade")
            responsavel_estado = request.form.get("responsavel_estado")
            responsavel_uf = request.form.get("responsavel_uf")
            if responsavel_uf:
                responsavel_uf = responsavel_uf.upper()
            else:
                responsavel_uf = None
            responsavel_estado_civil = request.form.get("responsavel_estado_civil")
            is_cnpj = len(documento) == 14
            if not is_cnpj:
                responsavel_nome = None
                responsavel_cpf = None
                responsavel_endereco = None
                responsavel_bairro = None
                responsavel_cidade = None
                responsavel_estado = None
                responsavel_uf = None
                responsavel_estado_civil = None
            tipo = request.form["tipo"]
            status = request.form["status"]

            # Validação de CPF/CNPJ (simplificada)
            if len(documento) == 11 and not documento.isdigit():
                 flash("CPF inválido. Deve conter apenas números.", "danger")
                 return render_template("pessoas/add_list.html", pessoa=request.form)
            elif len(documento) == 14 and not documento.isdigit():
                 flash("CNPJ inválido. Deve conter apenas números.", "danger")
                 return render_template("pessoas/add_list.html", pessoa={})
            elif len(documento) != 11 and len(documento) != 14:
                flash(
                    "Documento deve ser um CPF (11 dígitos) ou CNPJ (14 dígitos).",
                    "danger",
                )
                return render_template("pessoas/add_list.html", pessoa={})

            cur.execute(
                """
                UPDATE pessoas
                SET documento = %s, razao_social_nome = %s, nome_fantasia = %s, endereco = %s, bairro = %s, cidade = %s, estado = %s, cep = %s, telefone = %s, contato = %s, nacionalidade = %s, estado_civil = %s, profissao = %s, rg = %s, observacao = %s, responsavel_nome = %s, responsavel_cpf = %s, responsavel_endereco = %s, responsavel_bairro = %s, responsavel_cidade = %s, responsavel_estado = %s, responsavel_uf = %s, responsavel_estado_civil = %s, tipo = %s, status = %s
                WHERE id = %s
                """,
                (
                    documento,
                    razao_social_nome,
                    nome_fantasia,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    telefone,
                    contato,
                    nacionalidade,
                    estado_civil,
                    profissao,
                    rg,
                    observacao,
                    responsavel_nome,
                    responsavel_cpf,
                    responsavel_endereco,
                    responsavel_bairro,
                    responsavel_cidade,
                    responsavel_estado,
                    responsavel_uf,
                    responsavel_estado_civil,
                    tipo,
                    status,
                    id,
                ),
            )
            conn.commit()
            flash("Pessoa atualizada com sucesso!", "success")
            return redirect(url_for("pessoas_list"))
        except psycopg2.errors.UniqueViolation:
            flash(
                "Erro: Documento (CPF/CNPJ) já cadastrado para outra pessoa.", "danger"
            )
            conn.rollback()
            # Recarrega os dados da pessoa para preencher o formulário novamente
            cur.execute("SELECT * FROM pessoas WHERE id = %s", (id,))
            pessoa = cur.fetchone()
            return render_template("pessoas/add_list.html", pessoa=pessoa)
        except Exception as e:
            flash(f"Erro ao atualizar pessoa: {e}", "danger")
            # Recarrega os dados da pessoa para preencher o formulário novamente
            cur.execute("SELECT * FROM pessoas WHERE id = %s", (id,))
            pessoa = cur.fetchone()
            return render_template("pessoas/add_list.html", pessoa=pessoa)

    cur.execute("SELECT * FROM pessoas WHERE id = %s", (id,))
    pessoa = cur.fetchone()
    cur.close()
    conn.close()
    if pessoa is None:
        flash("Pessoa não encontrada.", "danger")
        return redirect(url_for("pessoas_list"))
    return render_template("pessoas/add_list.html", pessoa=pessoa)


@app.route("/pessoas/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Fornecedores/Clientes", "Excluir")
def pessoas_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM pessoas WHERE id = %s", (id,))
        conn.commit()
        flash("Pessoa excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir pessoa: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("pessoas_list"))


# --- Módulo de Gestão de Imóveis e Aluguéis ---


# 1.2. Cadastro de Imóveis
@app.route("/imoveis")
@login_required
@permission_required("Cadastro Imoveis", "Consultar")
def imoveis_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT i.*, EXISTS (
                SELECT 1 FROM contratos_aluguel c
                WHERE c.imovel_id = i.id AND c.status_contrato = 'Ativo'
            ) AS contrato_ativo
            FROM imoveis i
            WHERE endereco ILIKE %s OR bairro ILIKE %s OR cidade ILIKE %s OR inscricao_iptu ILIKE %s
            ORDER BY data_cadastro DESC
        """,
            (
                f"%{search_query}%",
                f"%{search_query}%",
                f"%{search_query}%",
                f"%{search_query}%",
            ),
        )
    else:
        cur.execute(
            """
            SELECT i.*, EXISTS (
                SELECT 1 FROM contratos_aluguel c
                WHERE c.imovel_id = i.id AND c.status_contrato = 'Ativo'
            ) AS contrato_ativo
            FROM imoveis i
            ORDER BY data_cadastro DESC
        """
        )
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "imoveis/list.html", imoveis=imoveis, search_query=search_query
    )


@app.route("/imoveis/mapa")
@login_required
@permission_required("Cadastro Imoveis", "Consultar")
def imoveis_mapa():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        """
        SELECT
            i.id,
            i.matricula,
            i.endereco,
            i.bairro,
            i.cidade,
            i.estado,
            i.inscricao_iptu,
            i.latitude,
            i.longitude,
            (c.id IS NOT NULL) AS contrato_ativo,
            p.razao_social_nome AS cliente_nome
        FROM imoveis i
        LEFT JOIN contratos_aluguel c
            ON c.imovel_id = i.id AND c.status_contrato = 'Ativo'
        LEFT JOIN pessoas p ON c.cliente_id = p.id
        WHERE i.latitude IS NOT NULL AND i.longitude IS NOT NULL
        """
    )
    imoveis = cur.fetchall()
    # Converte coordenadas para float quando possível para uso no JavaScript
    imoveis = [
        {
            **dict(row),
            "latitude": float(row["latitude"]) if row["latitude"] is not None else None,
            "longitude": float(row["longitude"]) if row["longitude"] is not None else None,
        }
        for row in imoveis
    ]
 
    # Totais de imóveis para exibição no mapa
    cur.execute(
        """
        SELECT
            COUNT(*) AS total_imoveis,
            COUNT(c.imovel_id) AS total_alugados
        FROM imoveis i
        LEFT JOIN (
            SELECT DISTINCT imovel_id
            FROM contratos_aluguel
            WHERE status_contrato = 'Ativo'
        ) c ON c.imovel_id = i.id
        """
    )
    totals = cur.fetchone()
    total_imoveis = totals["total_imoveis"]
    total_alugados = totals["total_alugados"]
    total_disponiveis = total_imoveis - total_alugados
    vacancia_percent = (
        (total_disponiveis / total_imoveis) * 100 if total_imoveis else 0
    )
    cur.close()
    conn.close()
    return render_template(
        "imoveis/mapa.html",
        imoveis=imoveis,
        total_imoveis=total_imoveis,
        total_alugados=total_alugados,
        total_disponiveis=total_disponiveis,
        vacancia_percent=vacancia_percent,
    )


@app.route("/imoveis/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Imoveis", "Incluir")
def imoveis_add():
    if request.method == "POST":
        try:
            tipo_imovel = request.form.get("tipo_imovel")
            endereco = request.form["endereco"]
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep")
            registro = request.form.get("registro")
            livro = request.form.get("livro")
            folha = request.form.get("folha")
            matricula = request.form.get("matricula")
            inscricao_iptu = request.form.get("inscricao_iptu")
            latitude = request.form.get("latitude") or None
            longitude = request.form.get("longitude") or None
            data_aquisicao_str = request.form.get("data_aquisicao")
            valor_imovel = parse_decimal(request.form.get("valor_imovel"))
            valor_previsto_aluguel = parse_decimal(
                request.form.get("valor_previsto_aluguel")
            ) or 0.0
            max_contratos = request.form.get("max_contratos") or 1
            destinacao = request.form.get("destinacao") or None
            observacao = request.form.get("observacao")

            data_aquisicao = (
                datetime.strptime(data_aquisicao_str, "%Y-%m-%d").date()
                if data_aquisicao_str
                else None
            )

            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO imoveis (tipo_imovel, endereco, bairro, cidade, estado, cep,
                                     registro, livro, folha, matricula, inscricao_iptu,
                                     latitude, longitude, data_aquisicao, valor_imovel,
                                     valor_previsto_aluguel, max_contratos, destinacao, observacao)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
                """,
                (
                    tipo_imovel,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    registro,
                    livro,
                    folha,
                    matricula,
                    inscricao_iptu,
                    latitude,
                    longitude,
                    data_aquisicao,
                    valor_imovel,
                    valor_previsto_aluguel,
                    max_contratos,
                    destinacao,
                    observacao,
                ),
            )
            imovel_id = cur.fetchone()[0]

            # Lidar com uploads de documentos
            if "anexos" in request.files:
                files = request.files.getlist("anexos")
                for file in files:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"], "imoveis_anexos", filename
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO imovel_anexos (imovel_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (
                                imovel_id,
                                filename,
                                filepath,
                                "documento",
                            ),
                        )

            # Upload de fotos (limite de 4)
            if "fotos" in request.files:
                fotos = request.files.getlist("fotos")[:4]
                for file in fotos:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"], "imoveis_anexos", filename
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO imovel_anexos (imovel_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (
                                imovel_id,
                                filename,
                                filepath,
                                "foto",
                            ),
                        )
            
            conn.commit()
            cur.close()
            conn.close()
            flash("Imóvel cadastrado com sucesso!", "success")
            return redirect(url_for("imoveis_list"))
        except Exception as e:
            if "conn" in locals():
                conn.rollback()
                cur.close()
                conn.close()
            flash(f"Erro ao cadastrar imóvel: {e}", "danger")
            return render_template("imoveis/add_list.html", imovel=request.form)
    return render_template("imoveis/add_list.html", imovel={})


@app.route("/imoveis/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Imoveis", "Editar")
def imoveis_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == "POST":
        try:
            tipo_imovel = request.form.get("tipo_imovel")
            endereco = request.form["endereco"]
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep")
            registro = request.form.get("registro")
            livro = request.form.get("livro")
            folha = request.form.get("folha")
            matricula = request.form.get("matricula")
            inscricao_iptu = request.form.get("inscricao_iptu")
            latitude = request.form.get("latitude") or None
            longitude = request.form.get("longitude") or None
            data_aquisicao_str = request.form.get("data_aquisicao")
            valor_imovel = parse_decimal(request.form.get("valor_imovel"))
            valor_previsto_aluguel = parse_decimal(
                request.form.get("valor_previsto_aluguel")
            ) or Decimal("0")
            max_contratos = request.form.get("max_contratos") or 1
            destinacao = request.form.get("destinacao") or None
            observacao = request.form.get("observacao")

            data_aquisicao = (
                datetime.strptime(data_aquisicao_str, "%Y-%m-%d").date()
                if data_aquisicao_str
                else None
            )

            cur.execute(
                """
                UPDATE imoveis
                SET tipo_imovel = %s, endereco = %s, bairro = %s, cidade = %s, estado = %s, cep = %s,
                    registro = %s, livro = %s, folha = %s, matricula = %s, inscricao_iptu = %s,
                    latitude = %s, longitude = %s, data_aquisicao = %s, valor_imovel = %s,
                    valor_previsto_aluguel = %s, max_contratos = %s, destinacao = %s, observacao = %s
                WHERE id = %s
                """,
                (
                    tipo_imovel,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    registro,
                    livro,
                    folha,
                    matricula,
                    inscricao_iptu,
                    latitude,
                    longitude,
                    data_aquisicao,
                    valor_imovel,
                    valor_previsto_aluguel,
                    max_contratos,
                    destinacao,
                    observacao,
                    id,
                ),
            )

            # Lidar com novos uploads de documentos
            if "anexos" in request.files:
                files = request.files.getlist("anexos")
                for file in files:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"], "imoveis_anexos", filename
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO imovel_anexos (imovel_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (id, filename, filepath, "documento"),
                        )

            # Upload de novas fotos (limite de 4 por envio)
            if "fotos" in request.files:
                fotos = request.files.getlist("fotos")[:4]
                for file in fotos:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"], "imoveis_anexos", filename
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO imovel_anexos (imovel_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (id, filename, filepath, "foto"),
                        )
            
            conn.commit()
            flash("Imóvel atualizado com sucesso!", "success")
            return redirect(url_for("imoveis_list"))
        except Exception as e:
            flash(f"Erro ao atualizar imóvel: {e}", "danger")
            conn.rollback()
            # Recarrega os dados do imóvel para preencher o formulário novamente
            cur.execute("SELECT * FROM imoveis WHERE id = %s", (id,))
            imovel = cur.fetchone()
            cur.execute("SELECT * FROM imovel_anexos WHERE imovel_id = %s", (id,))
            anexos = cur.fetchall()
            cur.close()
            conn.close()
            return render_template(
                "imoveis/add_list.html", imovel=imovel, anexos=anexos
            )

    cur.execute("SELECT * FROM imoveis WHERE id = %s", (id,))
    imovel = cur.fetchone()
    cur.execute("SELECT * FROM imovel_anexos WHERE imovel_id = %s", (id,))
    anexos = cur.fetchall()
    cur.close()
    conn.close()
    if imovel is None:
        flash("Imóvel não encontrado.", "danger")
        return redirect(url_for("imoveis_list"))
    return render_template("imoveis/add_list.html", imovel=imovel, anexos=anexos)


@app.route("/imoveis/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Imoveis", "Excluir")
def imoveis_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Primeiro, exclua os anexos relacionados para evitar violação de chave estrangeira
        cur.execute("DELETE FROM imovel_anexos WHERE imovel_id = %s", (id,))
        cur.execute("DELETE FROM imoveis WHERE id = %s", (id,))
        conn.commit()
        flash("Imóvel excluído com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir imóvel: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("imoveis_list"))


@app.route("/imoveis/anexo/delete/<int:anexo_id>", methods=["POST"])
@login_required
@permission_required(
    "Cadastro Imoveis", "Editar"
)  # Permissão para editar imóvel inclui exclusão de anexos
def imovel_anexo_delete(anexo_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            "SELECT caminho_arquivo, imovel_id FROM imovel_anexos WHERE id = %s",
            (anexo_id,),
        )
        anexo = cur.fetchone()
        
        if anexo:
            # Tenta remover o arquivo físico
            if os.path.exists(anexo["caminho_arquivo"]):
                os.remove(anexo["caminho_arquivo"])

            cur.execute("DELETE FROM imovel_anexos WHERE id = %s", (anexo_id,))
            conn.commit()
            flash("Anexo removido com sucesso!", "success")
            return redirect(url_for("imoveis_edit", id=anexo["imovel_id"]))
        else:
            flash("Anexo não encontrado.", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao remover anexo: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(
        url_for("imoveis_list")
    )  # Redireciona para a lista caso não encontre o anexo ou erro


# --- Avaliações de Imóveis ---
@app.route("/avaliacoes-imovel")
@login_required
@permission_required("Cadastro Imoveis", "Consultar")
def avaliacoes_imovel_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT a.*, i.endereco || ', ' || i.bairro AS endereco_imovel
            FROM avaliacoes_imovel a
            JOIN imoveis i ON a.imovel_id = i.id
            WHERE i.endereco ILIKE %s
            ORDER BY a.data_avaliacao DESC
            """,
            (f"%{search_query}%",),
        )
    else:
        cur.execute(
            """
            SELECT a.*, i.endereco || ', ' || i.bairro AS endereco_imovel
            FROM avaliacoes_imovel a
            JOIN imoveis i ON a.imovel_id = i.id
            ORDER BY a.data_avaliacao DESC
            """
        )
    avaliacoes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "avaliacoes_imovel/list.html", avaliacoes=avaliacoes, search_query=search_query
    )


@app.route("/avaliacoes-imovel/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Imoveis", "Incluir")
def avaliacoes_imovel_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = request.form["imovel_id"]
            data_str = request.form["data_avaliacao"]
            data_avaliacao = datetime.strptime(data_str + "-01", "%Y-%m-%d").date()
            valor_avaliacao = request.form["valor_avaliacao"]
            cur.execute(
                """
                INSERT INTO avaliacoes_imovel (imovel_id, data_avaliacao, valor_avaliacao)
                VALUES (%s, %s, %s)
                """,
                (imovel_id, data_avaliacao, valor_avaliacao),
            )
            conn.commit()
            flash("Avaliação cadastrada com sucesso!", "success")
            return redirect(url_for("avaliacoes_imovel_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar avaliação: {e}", "danger")
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("avaliacoes_imovel/add_list.html", avaliacao={}, imoveis=imoveis)


@app.route("/avaliacoes-imovel/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Imoveis", "Editar")
def avaliacoes_imovel_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = request.form["imovel_id"]
            data_str = request.form["data_avaliacao"]
            data_avaliacao = datetime.strptime(data_str + "-01", "%Y-%m-%d").date()
            valor_avaliacao = request.form["valor_avaliacao"]
            cur.execute(
                """
                UPDATE avaliacoes_imovel
                SET imovel_id = %s, data_avaliacao = %s, valor_avaliacao = %s
                WHERE id = %s
                """,
                (imovel_id, data_avaliacao, valor_avaliacao, id),
            )
            conn.commit()
            flash("Avaliação atualizada com sucesso!", "success")
            return redirect(url_for("avaliacoes_imovel_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar avaliação: {e}", "danger")
    cur.execute(
        """
        SELECT a.*, i.endereco || ', ' || i.bairro AS endereco_imovel
        FROM avaliacoes_imovel a
        JOIN imoveis i ON a.imovel_id = i.id
        WHERE a.id = %s
        """,
        (id,),
    )
    avaliacao = cur.fetchone()
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    if avaliacao is None:
        flash("Avaliação não encontrada.", "danger")
        return redirect(url_for("avaliacoes_imovel_list"))
    return render_template(
        "avaliacoes_imovel/add_list.html", avaliacao=avaliacao, imoveis=imoveis
    )


@app.route("/avaliacoes-imovel/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Imoveis", "Excluir")
def avaliacoes_imovel_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM avaliacoes_imovel WHERE id = %s", (id,))
        conn.commit()
        flash("Avaliação excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir avaliação: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("avaliacoes_imovel_list"))


# --- 1.3. Cadastro de Despesas ---
@app.route("/despesas")
@login_required
@permission_required("Cadastro Despesas", "Consultar")
def despesas_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT * FROM despesas_cadastro
            WHERE descricao ILIKE %s
            ORDER BY data_cadastro DESC
        """,
            (f"%{search_query}%",),
        )
    else:
        cur.execute("SELECT * FROM despesas_cadastro ORDER BY data_cadastro DESC")
    despesas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "despesas/list.html", despesas=despesas, search_query=search_query
    )


@app.route("/despesas/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Despesas", "Incluir")
def despesas_add():
    if request.method == "POST":
        descricao = request.form["descricao"]
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO despesas_cadastro (descricao) VALUES (%s)", (descricao,)
            )
            conn.commit()
            flash("Despesa cadastrada com sucesso!", "success")
            return redirect(url_for("despesas_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar despesa: {e}", "danger")
            return render_template("despesas/add_list.html", despesa=request.form)
        finally:
            cur.close()
            conn.close()
    return render_template("despesas/add_list.html", despesa={})


@app.route("/despesas/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Despesas", "Editar")
def despesas_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        descricao = request.form["descricao"]
        try:
            cur.execute(
                "UPDATE despesas_cadastro SET descricao = %s WHERE id = %s",
                (descricao, id),
            )
            conn.commit()
            flash("Despesa atualizada com sucesso!", "success")
            return redirect(url_for("despesas_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar despesa: {e}", "danger")
            cur.execute("SELECT * FROM despesas_cadastro WHERE id = %s", (id,))
            despesa = cur.fetchone()
            return render_template("despesas/add_list.html", despesa=despesa)
    cur.execute("SELECT * FROM despesas_cadastro WHERE id = %s", (id,))
    despesa = cur.fetchone()
    cur.close()
    conn.close()
    if despesa is None:
        flash("Despesa não encontrada.", "danger")
        return redirect(url_for("despesas_list"))
    return render_template("despesas/add_list.html", despesa=despesa)


@app.route("/despesas/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Despesas", "Excluir")
def despesas_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM despesas_cadastro WHERE id = %s", (id,))
        conn.commit()
        flash("Despesa excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir despesa: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("despesas_list"))


# --- 1.4. Cadastro de Origens ---
@app.route("/origens")
@login_required
@permission_required("Cadastro Origens", "Consultar")
def origens_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT * FROM origens_cadastro
            WHERE descricao ILIKE %s
            ORDER BY data_cadastro DESC
        """,
            (f"%{search_query}%",),
        )
    else:
        cur.execute("SELECT * FROM origens_cadastro ORDER BY data_cadastro DESC")
    origens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "origens/list.html", origens=origens, search_query=search_query
    )


@app.route("/origens/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Origens", "Incluir")
def origens_add():
    if request.method == "POST":
        descricao = request.form["descricao"]
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO origens_cadastro (descricao) VALUES (%s)", (descricao,)
            )
            conn.commit()
            flash("Origem cadastrada com sucesso!", "success")
            return redirect(url_for("origens_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar origem: {e}", "danger")
            return render_template("origens/add_list.html", origem=request.form)
        finally:
            cur.close()
            conn.close()
    return render_template("origens/add_list.html", origem={})


@app.route("/origens/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Origens", "Editar")
def origens_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        descricao = request.form["descricao"]
        try:
            cur.execute(
                "UPDATE origens_cadastro SET descricao = %s WHERE id = %s",
                (descricao, id),
            )
            conn.commit()
            flash("Origem atualizada com sucesso!", "success")
            return redirect(url_for("origens_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar origem: {e}", "danger")
            cur.execute("SELECT * FROM origens_cadastro WHERE id = %s", (id,))
            origem = cur.fetchone()
            return render_template("origens/add_list.html", origem=origem)
    cur.execute("SELECT * FROM origens_cadastro WHERE id = %s", (id,))
    origem = cur.fetchone()
    cur.close()
    conn.close()
    if origem is None:
        flash("Origem não encontrada.", "danger")
        return redirect(url_for("origens_list"))
    return render_template("origens/add_list.html", origem=origem)


@app.route("/origens/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Origens", "Excluir")
def origens_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM origens_cadastro WHERE id = %s", (id,))
        conn.commit()
        flash("Origem excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir origem: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("origens_list"))


# --- 1.5. Cadastro de Receitas ---
@app.route("/receitas")
@login_required
@permission_required("Cadastro Receitas", "Consultar")
def receitas_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT * FROM receitas_cadastro
            WHERE descricao ILIKE %s
            ORDER BY data_cadastro DESC
        """,
            (f"%{search_query}%",),
        )
    else:
        cur.execute("SELECT * FROM receitas_cadastro ORDER BY data_cadastro DESC")
    receitas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "receitaas/list.html", receitas=receitas, search_query=search_query
    )


@app.route("/receitas/add", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Receitas", "Incluir")
def receitas_add():
    if request.method == "POST":
        descricao = request.form["descricao"]
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO receitas_cadastro (descricao) VALUES (%s)", (descricao,)
            )
            conn.commit()
            flash("Receita cadastrada com sucesso!", "success")
            return redirect(url_for("receitas_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar receita: {e}", "danger")
            return render_template("receitaas/add_list.html", receita=request.form)
        finally:
            cur.close()
            conn.close()
    return render_template("receitaas/add_list.html", receita={})


@app.route("/receitas/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Cadastro Receitas", "Editar")
def receitas_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        descricao = request.form["descricao"]
        try:
            cur.execute(
                "UPDATE receitas_cadastro SET descricao = %s WHERE id = %s",
                (descricao, id),
            )
            conn.commit()
            flash("Receita atualizada com sucesso!", "success")
            return redirect(url_for("receitas_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar receita: {e}", "danger")
            cur.execute("SELECT * FROM receitas_cadastro WHERE id = %s", (id,))
            receita = cur.fetchone()
            return render_template("receitaas/add_list.html", receita=receita)
    cur.execute("SELECT * FROM receitas_cadastro WHERE id = %s", (id,))
    receita = cur.fetchone()
    cur.close()
    conn.close()
    if receita is None:
        flash("Receita não encontrada.", "danger")
        return redirect(url_for("receitas_list"))
    return render_template("receitaas/add_list.html", receita=receita)


@app.route("/receitas/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Cadastro Receitas", "Excluir")
def receitas_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM receitas_cadastro WHERE id = %s", (id,))
        conn.commit()
        flash("Receita excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir receita: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("receitas_list"))


# --- Rotas para outros módulos (placeholders existentes) ---
@app.route("/contratos")
@login_required
@permission_required("Gestao Contratos", "Consultar")
def contratos_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT c.*, i.endereco || ', ' || i.bairro AS endereco_imovel
            FROM contratos_aluguel c
            JOIN imoveis i ON c.imovel_id = i.id
            WHERE c.nome_inquilino ILIKE %s OR i.endereco ILIKE %s
            ORDER BY c.data_cadastro DESC
        """,
            (f"%{search_query}%", f"%{search_query}%"),
        )
    else:
        cur.execute(
            """
            SELECT c.*, i.endereco || ', ' || i.bairro AS endereco_imovel
            FROM contratos_aluguel c
            JOIN imoveis i ON c.imovel_id = i.id
            ORDER BY c.data_cadastro DESC
        """
        )
    contratos = cur.fetchall()
    # Carrega modelos de contrato para o modal de seleção
    cur.execute("SELECT id, nome FROM contrato_modelos ORDER BY nome")
    modelos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "contratos/list.html", contratos=contratos, modelos=modelos, search_query=search_query
    )


@app.route("/contratos/imovel/<int:imovel_id>/disponibilidade")
@login_required
@permission_required("Gestao Contratos", "Incluir")
def contratos_imovel_disponibilidade(imovel_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        """
        SELECT id
        FROM contratos_aluguel
        WHERE imovel_id = %s AND status_contrato = 'Ativo'
        """,
        (imovel_id,),
    )
    contrato = cur.fetchone()
    cur.close()
    conn.close()
    if contrato:
        return jsonify({"disponivel": False, "contrato_id": contrato["id"]})
    return jsonify({"disponivel": True})


def gerar_contas_a_receber_contrato(
    cur,
    *,
    contrato_id,
    cliente_id,
    finalidade,
    data_inicio,
    quantidade_parcelas,
    valor_parcela,
    vencimento_mesmo_dia=None,
    dias_intervalo=None,
    quantidade_calcao=0,
    valor_calcao=None,
    incluir_calcao=True,
    titulo_prefix="",
):
    """Create contas_a_receber rows for a contract using the existing rules."""
    if finalidade == "Comodato":
        return
    try:
        total_parcelas = int(quantidade_parcelas or 0)
    except (TypeError, ValueError):
        total_parcelas = 0
    if total_parcelas <= 0:
        return
    try:
        valor_decimal = (
            valor_parcela
            if isinstance(valor_parcela, Decimal)
            else Decimal(str(valor_parcela).replace(",", "."))
        )
    except (InvalidOperation, ValueError, TypeError):
        return
    if valor_decimal <= 0:
        return
    if isinstance(data_inicio, datetime):
        data_base = data_inicio.date()
    elif isinstance(data_inicio, date):
        data_base = data_inicio
    else:
        try:
            data_base = datetime.strptime(str(data_inicio), "%Y-%m-%d").date()
        except ValueError:
            return

    due_day = None
    interval_days = None
    if vencimento_mesmo_dia:
        try:
            due_day = max(1, min(31, int(vencimento_mesmo_dia)))
        except (TypeError, ValueError):
            due_day = None
    if due_day is None and dias_intervalo:
        try:
            interval_days = int(dias_intervalo)
        except (TypeError, ValueError):
            interval_days = None
    if due_day is None and interval_days is None:
        due_day = data_base.day

    def ensure_receita(descricao, aliases=None):
        candidatos = [descricao]
        if aliases:
            candidatos.extend(aliases)
        for candidato in candidatos:
            cur.execute(
                "SELECT id FROM receitas_cadastro WHERE descricao = %s",
                (candidato,),
            )
            resultado = cur.fetchone()
            if resultado:
                return resultado[0]
        cur.execute(
            "INSERT INTO receitas_cadastro (descricao) VALUES (%s) RETURNING id",
            (descricao,),
        )
        return cur.fetchone()[0]

    receita_id = ensure_receita("ALUGUEL")

    def build_titulo(numero):
        if titulo_prefix:
            return f"{contrato_id}-{titulo_prefix}{numero}/{total_parcelas}"
        return f"{contrato_id}-{numero}/{total_parcelas}"

    def add_months(start, months):
        month = start.month - 1 + months
        year = start.year + month // 12
        month = month % 12 + 1
        day = min(start.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)

    if due_day is not None:
        for numero in range(1, total_parcelas + 1):
            target_month = add_months(data_base, numero)
            ultimo_dia = calendar.monthrange(target_month.year, target_month.month)[1]
            vencimento = target_month.replace(day=min(due_day, ultimo_dia))
            cur.execute(
                """
                INSERT INTO contas_a_receber (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    titulo,
                    data_vencimento,
                    valor_previsto,
                    valor_pendente
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    build_titulo(numero),
                    vencimento,
                    valor_decimal,
                    valor_decimal,
                ),
            )
    else:
        if not interval_days or interval_days <= 0:
            interval_days = 30
        vencimento = data_base
        for numero in range(1, total_parcelas + 1):
            vencimento = vencimento + timedelta(days=interval_days)
            cur.execute(
                """
                INSERT INTO contas_a_receber (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    titulo,
                    data_vencimento,
                    valor_previsto,
                    valor_pendente
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    build_titulo(numero),
                    vencimento,
                    valor_decimal,
                    valor_decimal,
                ),
            )

    if not incluir_calcao:
        return
    try:
        qtd_calcao = int(quantidade_calcao or 0)
    except (TypeError, ValueError):
        qtd_calcao = 0
    if qtd_calcao <= 0:
        return
    try:
        valor_calcao_decimal = (
            valor_calcao
            if isinstance(valor_calcao, Decimal)
            else Decimal(str(valor_calcao).replace(",", "."))
        )
    except (InvalidOperation, ValueError, TypeError):
        return
    if valor_calcao_decimal <= 0:
        return

    calcao_receita_id = ensure_receita(
        "CALCOES",
        aliases=['CALÇÕES', 'CAL�OES'],
    )
    for numero in range(1, qtd_calcao + 1):
        titulo_calcao = (
            f"C{contrato_id}-{titulo_prefix}{numero}/{qtd_calcao}"
            if titulo_prefix
            else f"C{contrato_id}-{numero}/{qtd_calcao}"
        )
        venc_calcao = data_base + timedelta(days=30 * (numero - 1))
        cur.execute(
            """
            INSERT INTO contas_a_receber (
                contrato_id,
                receita_id,
                cliente_id,
                titulo,
                data_vencimento,
                valor_previsto,
                valor_pendente
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                contrato_id,
                calcao_receita_id,
                cliente_id,
                titulo_calcao,
                venc_calcao,
                valor_calcao_decimal,
                valor_calcao_decimal,
            ),
        )


@app.route("/contratos/add", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Incluir")
def contratos_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = request.form["imovel_id"]
            cliente_id = request.form["cliente_id"]
            data_inicio = datetime.strptime(request.form["data_inicio"], "%Y-%m-%d").date()
            data_fim = datetime.strptime(request.form["data_fim"], "%Y-%m-%d").date()
            finalidade = request.form["finalidade"]
            # Campos financeiros podem ser omitidos quando Comodato (inputs desabilitados)
            quantidade_parcelas = request.form.get("quantidade_parcelas")
            valor_parcela = request.form.get("valor_parcela")
            quantidade_calcao = int(request.form.get("quantidade_calcao") or 0)
            valor_calcao = request.form.get("valor_calcao")
            status_contrato = request.form["status_contrato"]
            observacao = request.form.get("observacao")
            vencimento_mesmo_dia = request.form.get("vencimento_mesmo_dia")
            dias_intervalo = request.form.get("dias_intervalo")
            if finalidade != "Comodato":
                if (vencimento_mesmo_dia and dias_intervalo) or (
                    not vencimento_mesmo_dia and not dias_intervalo
                ):
                    raise ValueError(
                        "Informe apenas um dos campos: Vencimento no mesmo dia ou Dias e intervalo"
                    )

            # Para contratos de Comodato não há contas a receber; força valores neutros
            if finalidade == "Comodato":
                quantidade_parcelas = 0
                valor_parcela = 0
                quantidade_calcao = 0
                valor_calcao = None
            else:
                # Garante que venham valores válidos para o banco
                quantidade_parcelas = int(quantidade_parcelas)
                # aceita string numérica; o driver converte para NUMERIC
                valor_parcela = valor_parcela or 0

            cur.execute(
                """
                SELECT id
                FROM contratos_aluguel
                WHERE imovel_id = %s AND status_contrato = 'Ativo'
                """,
                (imovel_id,),
            )
            contrato_ativo = cur.fetchone()
            if contrato_ativo:
                conn.rollback()
                flash(
                    f"Imovel ja possui contrato ativo (#{contrato_ativo['id']}). Selecione outro imovel.",
                    "danger",
                )
                cur.execute("SELECT id, endereco FROM imoveis ORDER BY endereco")
                imoveis = cur.fetchall()
                cur.execute(
                    "SELECT * FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
                )
                clientes = cur.fetchall()
                cur.close()
                conn.close()
                return render_template(
                    "contratos/add_list.html",
                    contrato=request.form,
                    imoveis=imoveis,
                    clientes=clientes,
                    anexos=[],
                )

            cur.execute(
                "SELECT razao_social_nome, endereco, bairro, cidade, estado, cep, telefone FROM pessoas WHERE id = %s",
                (cliente_id,),
            )
            cli = cur.fetchone()
            nome_inquilino = cli["razao_social_nome"] if cli else ""
            endereco_inquilino = cli["endereco"] if cli else ""
            bairro_inquilino = cli["bairro"] if cli else ""
            cidade_inquilino = cli["cidade"] if cli else ""
            estado_inquilino = cli["estado"] if cli else ""
            cep_inquilino = cli["cep"] if cli else ""
            telefone_inquilino = cli["telefone"] if cli else ""

            cur.execute(
                """
                INSERT INTO contratos_aluguel (
                    imovel_id, cliente_id, nome_inquilino, endereco_inquilino,
                    bairro_inquilino, cidade_inquilino, estado_inquilino,
                    cep_inquilino, telefone_inquilino, finalidade, data_inicio, data_fim,
                    quantidade_parcelas, valor_parcela, quantidade_calcao,
                    valor_calcao, status_contrato, observacao
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    imovel_id,
                    cliente_id,
                    nome_inquilino,
                    endereco_inquilino,
                    bairro_inquilino,
                    cidade_inquilino,
                    estado_inquilino,
                    cep_inquilino,
                    telefone_inquilino,
                    finalidade,
                    data_inicio,
                    data_fim,
                    quantidade_parcelas,
                    valor_parcela,
                    quantidade_calcao,
                    valor_calcao,
                    status_contrato,
                    observacao,
                ),
            )
            contrato_id = cur.fetchone()[0]

            gerar_contas_a_receber_contrato(
                cur,
                contrato_id=contrato_id,
                cliente_id=cliente_id,
                finalidade=finalidade,
                data_inicio=data_inicio,
                quantidade_parcelas=quantidade_parcelas,
                valor_parcela=valor_parcela,
                vencimento_mesmo_dia=vencimento_mesmo_dia,
                dias_intervalo=dias_intervalo,
                quantidade_calcao=quantidade_calcao,
                valor_calcao=valor_calcao,
            )
            if "anexos" in request.files:
                files = request.files.getlist("anexos")
                for file in files:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"],
                            "contratos_anexos",
                            filename,
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO contrato_anexos (contrato_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (contrato_id, filename, filepath, "documento"),
                        )

            conn.commit()
            flash("Contrato cadastrado com sucesso!", "success")
            return redirect(url_for("contratos_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar contrato: {e}", "danger")
            cur.execute("SELECT id, endereco FROM imoveis ORDER BY endereco")
            imoveis = cur.fetchall()
            cur.execute(
                "SELECT * FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
            )
            clientes = cur.fetchall()
            return render_template(
                "contratos/add_list.html",
                contrato=request.form,
                imoveis=imoveis,
                clientes=clientes,
                anexos=[],
            )

    cur.execute("SELECT id, endereco FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.execute(
        "SELECT * FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    clientes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "contratos/add_list.html", contrato={}, imoveis=imoveis, clientes=clientes, anexos=[]
    )


@app.route("/contratos/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = request.form["imovel_id"]
            cliente_id = request.form["cliente_id"]
            data_inicio = datetime.strptime(request.form["data_inicio"], "%Y-%m-%d").date()
            data_fim = datetime.strptime(request.form["data_fim"], "%Y-%m-%d").date()
            finalidade = request.form["finalidade"]
            quantidade_parcelas = request.form.get("quantidade_parcelas")
            valor_parcela = request.form.get("valor_parcela")
            quantidade_calcao = int(request.form.get("quantidade_calcao") or 0)
            valor_calcao = request.form.get("valor_calcao")
            status_contrato = request.form["status_contrato"]
            observacao = request.form.get("observacao")

            # Ajustes para Comodato
            if finalidade == "Comodato":
                quantidade_parcelas = 0
                valor_parcela = 0
                quantidade_calcao = 0
                valor_calcao = None
            else:
                quantidade_parcelas = int(quantidade_parcelas)
                valor_parcela = valor_parcela or 0

            cur.execute(
                "SELECT razao_social_nome, endereco, bairro, cidade, estado, cep, telefone FROM pessoas WHERE id = %s",
                (cliente_id,),
            )
            cli = cur.fetchone()
            nome_inquilino = cli["razao_social_nome"] if cli else ""
            endereco_inquilino = cli["endereco"] if cli else ""
            bairro_inquilino = cli["bairro"] if cli else ""
            cidade_inquilino = cli["cidade"] if cli else ""
            estado_inquilino = cli["estado"] if cli else ""
            cep_inquilino = cli["cep"] if cli else ""
            telefone_inquilino = cli["telefone"] if cli else ""

            cur.execute(
                """
                UPDATE contratos_aluguel
                SET imovel_id = %s, cliente_id = %s, nome_inquilino = %s,
                    endereco_inquilino = %s, bairro_inquilino = %s,
                    cidade_inquilino = %s, estado_inquilino = %s,
                    cep_inquilino = %s, telefone_inquilino = %s,
                    finalidade = %s,
                    data_inicio = %s, data_fim = %s,
                    quantidade_parcelas = %s, valor_parcela = %s,
                    quantidade_calcao = %s, valor_calcao = %s,
                    status_contrato = %s, observacao = %s
                WHERE id = %s
                """,
                (
                    imovel_id,
                    cliente_id,
                    nome_inquilino,
                    endereco_inquilino,
                    bairro_inquilino,
                    cidade_inquilino,
                    estado_inquilino,
                    cep_inquilino,
                    telefone_inquilino,
                    finalidade,
                    data_inicio,
                    data_fim,
                    quantidade_parcelas,
                    valor_parcela,
                    quantidade_calcao,
                    valor_calcao,
                    status_contrato,
                    observacao,
                    id,
                ),
            )

            if finalidade != "Comodato" and quantidade_calcao > 0 and valor_calcao:
                cur.execute(
                    "SELECT COUNT(*) FROM contas_a_receber WHERE contrato_id = %s AND titulo LIKE 'C%%'",
                    (id,),
                )
                existing_calcao = cur.fetchone()[0]
                if existing_calcao == 0:
                    cur.execute(
                        "SELECT id FROM receitas_cadastro WHERE descricao = %s",
                        ("CALÇOES",),
                    )
                    calcao_result = cur.fetchone()
                    if calcao_result:
                        calcao_receita_id = calcao_result[0]
                    else:
                        cur.execute(
                            "INSERT INTO receitas_cadastro (descricao) VALUES (%s) RETURNING id",
                            ("CALÇOES",),
                        )
                        calcao_receita_id = cur.fetchone()[0]

                    for numero in range(1, quantidade_calcao + 1):
                        titulo_calcao = f"C{id}-{numero}/{quantidade_calcao}"
                        venc_calcao = data_inicio + timedelta(days=30 * (numero - 1))
                        cur.execute(
                            """
                            INSERT INTO contas_a_receber (
                                contrato_id, receita_id, cliente_id, titulo,
                                data_vencimento, valor_previsto, valor_pendente
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                id,
                                calcao_receita_id,
                                cliente_id,
                                titulo_calcao,
                                venc_calcao,
                                valor_calcao,
                                valor_calcao,
                            ),
                        )

            if "anexos" in request.files:
                files = request.files.getlist("anexos")
                for file in files:
                    if file and allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(
                            app.config["UPLOAD_FOLDER"],
                            "contratos_anexos",
                            filename,
                        )
                        file.save(filepath)
                        cur.execute(
                            "INSERT INTO contrato_anexos (contrato_id, nome_arquivo, caminho_arquivo, tipo_anexo) VALUES (%s, %s, %s, %s)",
                            (id, filename, filepath, "documento"),
                        )

            conn.commit()
            flash("Contrato atualizado com sucesso!", "success")
            return redirect(url_for("contratos_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar contrato: {e}", "danger")

    cur.execute(
        """
        SELECT c.*, i.endereco || ', ' || i.bairro AS endereco_imovel
        FROM contratos_aluguel c
        JOIN imoveis i ON c.imovel_id = i.id
        WHERE c.id = %s
        """,
        (id,),
    )
    contrato = cur.fetchone()
    cur.execute("SELECT * FROM contrato_anexos WHERE contrato_id = %s", (id,))
    anexos = cur.fetchall()
    cur.execute("SELECT id, endereco FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.execute(
        "SELECT * FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    clientes = cur.fetchall()
    cur.close()
    conn.close()
    if contrato is None:
        flash("Contrato não encontrado.", "danger")
        return redirect(url_for("contratos_list"))
    return render_template(
        "contratos/add_list.html",
        contrato=contrato,
        imoveis=imoveis,
        clientes=clientes,
        anexos=anexos,
    )


@app.route("/contratos/<int:contrato_id>/renovar", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_renovar(contrato_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    form_data = {}
    try:
        cur.execute(
            """
            SELECT c.*, i.endereco || ', ' || i.bairro AS endereco_imovel
            FROM contratos_aluguel c
            JOIN imoveis i ON c.imovel_id = i.id
            WHERE c.id = %s
            """,
            (contrato_id,),
        )
        contrato = cur.fetchone()
        if contrato is None:
            flash("Contrato não encontrado.", "danger")
            return redirect(url_for("contratos_list"))

        if request.method == "POST":
            form_data = {
                "nova_data_inicio": (request.form.get("nova_data_inicio") or "").strip(),
                "nova_data_fim": (request.form.get("nova_data_fim") or "").strip(),
                "novo_valor": (request.form.get("novo_valor") or "").strip(),
                "observacao": (request.form.get("observacao") or "").strip(),
            }
            erros = []

            try:
                nova_data_inicio = datetime.strptime(form_data["nova_data_inicio"], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                erros.append("Informe uma data inicial válida.")
                nova_data_inicio = None

            try:
                nova_data_fim = datetime.strptime(form_data["nova_data_fim"], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                erros.append("Informe uma data final válida.")
                nova_data_fim = None

            if nova_data_inicio and nova_data_fim and nova_data_fim < nova_data_inicio:
                erros.append("A data final do contrato deve ser igual ou posterior à data inicial.")

            try:
                valor_raw = form_data["novo_valor"].replace(" ", "")
                if "," in valor_raw and "." in valor_raw:
                    if valor_raw.rfind(",") > valor_raw.rfind("."):
                        valor_str = valor_raw.replace(".", "").replace(",", ".")
                    else:
                        valor_str = valor_raw.replace(",", "")
                else:
                    valor_str = valor_raw.replace(",", ".")
                novo_valor = Decimal(valor_str)
                if novo_valor <= 0:
                    erros.append("O valor do contrato deve ser maior que zero.")
            except (InvalidOperation, AttributeError):
                novo_valor = None
                erros.append("Informe um valor válido para o contrato.")

            if erros:
                for mensagem in erros:
                    flash(mensagem, "warning")
                cur.execute(
                    """
                    SELECT r.*, u.nome_usuario AS usuario_nome
                    FROM contrato_renovacoes r
                    LEFT JOIN usuarios u ON u.id = r.usuario_id
                    WHERE r.contrato_id = %s
                    ORDER BY r.data_renovacao DESC
                    """,
                    (contrato_id,),
                )
                renovacoes = cur.fetchall()
                return render_template(
                    "contratos/renovar.html",
                    contrato=contrato,
                    renovacoes=renovacoes,
                    form_data=form_data,
                )

            try:
                cur.execute(
                    """
                    INSERT INTO contrato_renovacoes (
                        contrato_id,
                        data_inicio_anterior,
                        data_fim_anterior,
                        valor_parcela_anterior,
                        data_inicio_novo,
                        data_fim_novo,
                        valor_parcela_novo,
                        observacao,
                        usuario_id
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        contrato_id,
                        contrato["data_inicio"],
                        contrato["data_fim"],
                        contrato["valor_parcela"],
                        nova_data_inicio,
                        nova_data_fim,
                        novo_valor,
                        form_data["observacao"] or None,
                        session.get("user_id"),
                    ),
                )
                renovacao_id = cur.fetchone()[0]

                quantidade_parcelas_atual = contrato.get("quantidade_parcelas")
                try:
                    quantidade_parcelas_int = int(quantidade_parcelas_atual or 0)
                except (TypeError, ValueError):
                    quantidade_parcelas_int = 0
                if quantidade_parcelas_int <= 0 and nova_data_inicio and nova_data_fim:
                    quantidade_parcelas_int = max(
                        1,
                        (nova_data_fim.year - nova_data_inicio.year) * 12
                        + (nova_data_fim.month - nova_data_inicio.month)
                        + 1,
                    )
                elif quantidade_parcelas_int <= 0:
                    quantidade_parcelas_int = 1

                cur.execute(
                    """
                    UPDATE contratos_aluguel
                    SET data_inicio = %s,
                        data_fim = %s,
                        valor_parcela = %s,
                        quantidade_parcelas = %s,
                        status_contrato = 'Ativo'
                    WHERE id = %s
                    """,
                    (
                        nova_data_inicio,
                        nova_data_fim,
                        novo_valor,
                        quantidade_parcelas_int,
                        contrato_id,
                    ),
                )

                vencimento_mesmo_dia_param = None
                dias_intervalo_param = None
                cur.execute(
                    """
                    SELECT data_vencimento
                    FROM contas_a_receber
                    WHERE contrato_id = %s
                      AND (titulo IS NULL OR titulo NOT LIKE 'C%%')
                    ORDER BY data_vencimento ASC
                    LIMIT 2
                    """,
                    (contrato_id,),
                )
                datas_vencimento = [row[0] for row in cur.fetchall() if row[0]]
                if len(datas_vencimento) >= 2:
                    if datas_vencimento[0].day == datas_vencimento[1].day:
                        vencimento_mesmo_dia_param = datas_vencimento[0].day
                    else:
                        diff_days = (datas_vencimento[1] - datas_vencimento[0]).days
                        if diff_days > 0:
                            dias_intervalo_param = diff_days
                elif len(datas_vencimento) == 1:
                    vencimento_mesmo_dia_param = datas_vencimento[0].day
                if not vencimento_mesmo_dia_param and not dias_intervalo_param:
                    vencimento_mesmo_dia_param = nova_data_inicio.day

                titulo_prefix = f"R{renovacao_id}-" if renovacao_id else ""

                gerar_contas_a_receber_contrato(
                    cur,
                    contrato_id=contrato_id,
                    cliente_id=contrato["cliente_id"],
                    finalidade=contrato["finalidade"],
                    data_inicio=nova_data_inicio,
                    quantidade_parcelas=quantidade_parcelas_int,
                    valor_parcela=novo_valor,
                    vencimento_mesmo_dia=vencimento_mesmo_dia_param,
                    dias_intervalo=dias_intervalo_param,
                    incluir_calcao=False,
                    titulo_prefix=titulo_prefix,
                )

                conn.commit()
                flash("Contrato renovado com sucesso!", "success")
                return redirect(url_for("contratos_list"))
            except Exception as exc:
                conn.rollback()
                flash(f"Erro ao renovar contrato: {exc}", "danger")
                cur.execute(
                    """
                    SELECT r.*, u.nome_usuario AS usuario_nome
                    FROM contrato_renovacoes r
                    LEFT JOIN usuarios u ON u.id = r.usuario_id
                    WHERE r.contrato_id = %s
                    ORDER BY r.data_renovacao DESC
                    """,
                    (contrato_id,),
                )
                renovacoes = cur.fetchall()
                return render_template(
                    "contratos/renovar.html",
                    contrato=contrato,
                    renovacoes=renovacoes,
                    form_data=form_data,
                )

        form_data = {
            "nova_data_inicio": contrato["data_inicio"].isoformat() if contrato["data_inicio"] else "",
            "nova_data_fim": contrato["data_fim"].isoformat() if contrato["data_fim"] else "",
            "novo_valor": (
                format(contrato["valor_parcela"], ".2f")
                if contrato["valor_parcela"] is not None
                else ""
            ),
            "observacao": "",
        }

        cur.execute(
            """
            SELECT r.*, u.nome_usuario AS usuario_nome
            FROM contrato_renovacoes r
            LEFT JOIN usuarios u ON u.id = r.usuario_id
            WHERE r.contrato_id = %s
            ORDER BY r.data_renovacao DESC
            """,
            (contrato_id,),
        )
        renovacoes = cur.fetchall()

        return render_template(
            "contratos/renovar.html",
            contrato=contrato,
            renovacoes=renovacoes,
            form_data=form_data,
        )
    finally:
        cur.close()
        conn.close()




@app.route("/contratos/<int:contrato_id>/renovacoes/<int:renovacao_id>/excluir", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_renovacao_excluir(contrato_id, renovacao_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            """
            SELECT *
            FROM contrato_renovacoes
            WHERE id = %s AND contrato_id = %s
            """,
            (renovacao_id, contrato_id),
        )
        renovacao = cur.fetchone()
        if renovacao is None:
            flash("Renovacao nao encontrada.", "warning")
            return redirect(url_for("contratos_renovar", contrato_id=contrato_id))

        cur.execute(
            """
            SELECT id
            FROM contrato_renovacoes
            WHERE contrato_id = %s
            ORDER BY data_renovacao DESC, id DESC
            LIMIT 1
            """,
            (contrato_id,),
        )
        ultima = cur.fetchone()
        if not ultima or ultima["id"] != renovacao_id:
            flash("So e possivel excluir a ultima renovacao registrada.", "warning")
            return redirect(url_for("contratos_renovar", contrato_id=contrato_id))

        cur.execute(
            """
            UPDATE contratos_aluguel
            SET data_inicio = %s,
                data_fim = %s,
                valor_parcela = %s,
                status_contrato = 'Ativo'
            WHERE id = %s
            """,
            (
                renovacao["data_inicio_anterior"],
                renovacao["data_fim_anterior"],
                renovacao["valor_parcela_anterior"],
                contrato_id,
            ),
        )

        cur.execute(
            "DELETE FROM contrato_renovacoes WHERE id = %s",
            (renovacao_id,),
        )

        conn.commit()
        flash("Renovacao excluida e dados do contrato restaurados.", "success")
    except Exception as exc:
        conn.rollback()
        flash(f"Erro ao excluir renovacao: {exc}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(url_for("contratos_renovar", contrato_id=contrato_id))

@app.route("/contratos/<int:id>/titulos", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_titulos(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute("SELECT id FROM contratos_aluguel WHERE id = %s", (id,))
        if cur.fetchone() is None:
            return jsonify({"error": "Contrato nao encontrado."}), 404

        atualizar_status_contas_a_receber(cur)
        conn.commit()

        cur.execute(
            """
            SELECT id, titulo, data_vencimento, valor_previsto, valor_pago, valor_pendente, status_conta
            FROM contas_a_receber
            WHERE contrato_id = %s AND status_conta IN ('Aberta','Vencida')
            ORDER BY data_vencimento, id
            """,
            (id,),
        )
        titulos = cur.fetchall()
        response = []
        for titulo in titulos:
            valor_previsto = Decimal(titulo["valor_previsto"])
            valor_pago = (
                Decimal(titulo["valor_pago"])
                if titulo["valor_pago"] is not None
                else Decimal("0")
            )
            valor_pendente = (
                Decimal(titulo["valor_pendente"])
                if titulo["valor_pendente"] is not None
                else valor_previsto - valor_pago
            )
            data_vencimento = titulo["data_vencimento"]
            response.append(
                {
                    "id": titulo["id"],
                    "titulo": titulo["titulo"],
                    "status": titulo["status_conta"],
                    "data_vencimento": data_vencimento.isoformat() if data_vencimento else None,
                    "data_vencimento_formatada": data_vencimento.strftime("%d/%m/%Y")
                    if data_vencimento
                    else None,
                    "valor_previsto": format(valor_previsto, ".2f"),
                    "valor_pago": format(valor_pago, ".2f"),
                    "valor_pendente": format(valor_pendente, ".2f"),
                }
            )
        return jsonify({"titulos": response})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/contratos/encerrar/<int:id>", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_encerrar(id):
    data_encerramento_str = request.form.get("data_encerramento")
    if not data_encerramento_str:
        flash("Informe a data de encerramento.", "warning")
        return redirect(url_for("contratos_list"))

    try:
        data_encerramento = datetime.strptime(data_encerramento_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Data de encerramento invalida.", "danger")
        return redirect(url_for("contratos_list"))

    prestacao_raw = request.form.get("prestacao_payload")
    prestacao_payload = None
    if prestacao_raw:
        try:
            prestacao_payload = json.loads(prestacao_raw)
        except json.JSONDecodeError:
            flash("Dados da prestação de contas inválidos.", "danger")
            return redirect(url_for("contratos_list"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        contrato = fetch_contrato_info(cur, id)
        if contrato is None:
            raise ValueError("Contrato nao encontrado.")

        prestacao_result = None
        if prestacao_payload is not None:
            prestacao_result = processar_prestacao(
                cur,
                contrato,
                data_encerramento,
                prestacao_payload,
            )
        else:
            atualizar_status_contas_a_receber(cur)

        cur.execute(
            """
            UPDATE contratos_aluguel
            SET status_contrato = %s, data_fim = %s
            WHERE id = %s
            """,
            ("Encerrado", data_encerramento, id),
        )
        conn.commit()
        if prestacao_result:
            flash("Prestação de contas gerada e contrato encerrado com sucesso!", "success")
        else:
            flash("Contrato encerrado com sucesso!", "success")
    except ValueError as e:
        conn.rollback()
        flash(str(e), "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao encerrar contrato: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contratos_list"))

# ------------------ Modelos de Contrato (CRUD + Geração) ------------------

@app.route("/contratos/modelos")
@login_required
@permission_required("Gestao Contratos", "Consultar")
def contrato_modelos_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            "SELECT * FROM contrato_modelos WHERE nome ILIKE %s ORDER BY data_atualizacao DESC",
            (f"%{search_query}%",),
        )
    else:
        cur.execute("SELECT * FROM contrato_modelos ORDER BY data_atualizacao DESC")
    modelos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("contratos/modelos/list.html", modelos=modelos, search_query=search_query)


@app.route("/contratos/modelos/add", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Incluir")
def contrato_modelos_add():
    if request.method == "POST":
        nome = request.form.get("nome").strip()
        conteudo_html = request.form.get("conteudo_html") or ""
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO contrato_modelos (nome, conteudo_html) VALUES (%s, %s)",
                (nome, conteudo_html),
            )
            conn.commit()
            flash("Modelo criado com sucesso!", "success")
            return redirect(url_for("contrato_modelos_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao criar modelo: {e}", "danger")
            return render_template("contratos/modelos/add_list.html", modelo={"nome": nome, "conteudo_html": conteudo_html})
        finally:
            cur.close()
            conn.close()
    return render_template("contratos/modelos/add_list.html", modelo={})


@app.route("/contratos/modelos/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contrato_modelos_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        nome = request.form.get("nome").strip()
        conteudo_html = request.form.get("conteudo_html") or ""
        try:
            cur.execute(
                "UPDATE contrato_modelos SET nome = %s, conteudo_html = %s, data_atualizacao = NOW() WHERE id = %s",
                (nome, conteudo_html, id),
            )
            conn.commit()
            flash("Modelo atualizado com sucesso!", "success")
            return redirect(url_for("contrato_modelos_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar modelo: {e}", "danger")
    cur.execute("SELECT * FROM contrato_modelos WHERE id = %s", (id,))
    modelo = cur.fetchone()
    cur.close()
    conn.close()
    if not modelo:
        flash("Modelo não encontrado.", "danger")
        return redirect(url_for("contrato_modelos_list"))
    return render_template("contratos/modelos/add_list.html", modelo=modelo)


@app.route("/contratos/modelos/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Excluir")
def contrato_modelos_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM contrato_modelos WHERE id = %s", (id,))
        conn.commit()
        flash("Modelo excluído com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir modelo: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contrato_modelos_list"))


@app.route("/contratos/preview/<int:contrato_id>")
@login_required
@permission_required("Gestao Contratos", "Consultar")
def contrato_preview(contrato_id):
    modelo_id = request.args.get("modelo_id")
    if not modelo_id:
        flash("Selecione um modelo para gerar o contrato.", "warning")
        return redirect(url_for("contratos_list"))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM contrato_modelos WHERE id = %s", (modelo_id,))
    modelo = cur.fetchone()
    if not modelo:
        cur.close()
        conn.close()
        flash("Modelo não encontrado.", "danger")
        return redirect(url_for("contratos_list"))
    ctx = build_contrato_context(cur, contrato_id)
    cur.close()
    conn.close()
    conteudo = render_placeholders(modelo["conteudo_html"], ctx)
    return render_template("contratos/modelos/preview.html", conteudo=conteudo, contrato_id=contrato_id)


@app.route("/contratos/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Excluir")
def contratos_delete(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            "SELECT id, caminho_arquivo FROM contrato_anexos WHERE contrato_id = %s",
            (id,),
        )
        anexos = cur.fetchall()
        for anexo in anexos:
            if os.path.exists(anexo["caminho_arquivo"]):
                os.remove(anexo["caminho_arquivo"])
        cur.execute("DELETE FROM contrato_anexos WHERE contrato_id = %s", (id,))
        cur.execute("DELETE FROM contratos_aluguel WHERE id = %s", (id,))
        conn.commit()
        flash("Contrato excluído com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir contrato: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contratos_list"))


@app.route("/contratos/anexo/delete/<int:anexo_id>", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contrato_anexo_delete(anexo_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            "SELECT caminho_arquivo, contrato_id FROM contrato_anexos WHERE id = %s",
            (anexo_id,),
        )
        anexo = cur.fetchone()
        if anexo:
            if os.path.exists(anexo["caminho_arquivo"]):
                os.remove(anexo["caminho_arquivo"])
            cur.execute("DELETE FROM contrato_anexos WHERE id = %s", (anexo_id,))
            conn.commit()
            flash("Anexo removido com sucesso!", "success")
            return redirect(url_for("contratos_edit", id=anexo["contrato_id"]))
        else:
            flash("Anexo não encontrado.", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao remover anexo: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contratos_list"))


# --- Reajustes de Contrato ---
@app.route("/reajustes", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Consultar")
def reajustes_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT r.*, c.nome_inquilino, c.valor_parcela
            FROM reajustes_contrato r
            JOIN contratos_aluguel c ON r.contrato_id = c.id
            WHERE CAST(r.contrato_id AS TEXT) ILIKE %s OR c.nome_inquilino ILIKE %s
            ORDER BY r.data_alteracao DESC
            """,
            (f"%{search_query}%", f"%{search_query}%"),
        )
    else:
        cur.execute(
            """
            SELECT r.*, c.nome_inquilino, c.valor_parcela
            FROM reajustes_contrato r
            JOIN contratos_aluguel c ON r.contrato_id = c.id
            ORDER BY r.data_alteracao DESC
            """
        )
    reajustes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "reajustes_contrato/list.html", reajustes=reajustes, search_query=search_query
    )


@app.route("/reajustes/add", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Incluir")
def reajustes_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            contrato_id = request.form["contrato_id"]
            data_alteracao = datetime.strptime(request.form["data_alteracao"], "%Y-%m-%d").date()
            percentual_reajuste = float(request.form["percentual_reajuste"])
            cur.execute(
                "SELECT valor_parcela FROM contratos_aluguel WHERE id = %s",
                (contrato_id,),
            )
            contrato = cur.fetchone()
            if not contrato:
                flash("Contrato não encontrado.", "danger")
                cur.execute(
                    "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel ORDER BY id"
                )
                contratos = cur.fetchall()
                return render_template(
                    "reajustes_contrato/add_list.html",
                    reajuste=request.form,
                    contratos=contratos,
                )
            valor_atual = float(contrato["valor_parcela"])
            novo_valor = round(valor_atual * (1 + percentual_reajuste / 100), 2)
            observacao = request.form.get("observacao")
            cur.execute(
                """
                INSERT INTO reajustes_contrato (contrato_id, data_alteracao, percentual_reajuste, novo_valor_parcela, observacao)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (contrato_id, data_alteracao, percentual_reajuste, novo_valor, observacao),
            )
            cur.execute(
                "UPDATE contratos_aluguel SET valor_parcela = %s WHERE id = %s",
                (novo_valor, contrato_id),
            )
            conn.commit()
            flash("Reajuste cadastrado com sucesso!", "success")
            return redirect(url_for("reajustes_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar reajuste: {e}", "danger")
            cur.execute(
                "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel ORDER BY id"
            )
            contratos = cur.fetchall()
            return render_template(
                "reajustes_contrato/add_list.html",
                reajuste=request.form,
                contratos=contratos,
            )
        finally:
            cur.close()
            conn.close()
    cur.execute(
        "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel ORDER BY id"
    )
    contratos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "reajustes_contrato/add_list.html", reajuste={}, contratos=contratos
    )


@app.route("/reajustes/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def reajustes_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            contrato_id = request.form["contrato_id"]
            data_alteracao = datetime.strptime(request.form["data_alteracao"], "%Y-%m-%d").date()
            percentual_reajuste = float(request.form["percentual_reajuste"])
            cur.execute(
                "SELECT valor_parcela FROM contratos_aluguel WHERE id = %s",
                (contrato_id,),
            )
            contrato = cur.fetchone()
            if not contrato:
                flash("Contrato não encontrado.", "danger")
                cur.execute(
                    "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel ORDER BY id"
                )
                contratos = cur.fetchall()
                return render_template(
                    "reajustes_contrato/add_list.html",
                    reajuste=request.form,
                    contratos=contratos,
                )
            valor_atual = float(contrato["valor_parcela"])
            novo_valor = round(valor_atual * (1 + percentual_reajuste / 100), 2)
            observacao = request.form.get("observacao")
            cur.execute(
                """
                UPDATE reajustes_contrato
                SET contrato_id = %s, data_alteracao = %s, percentual_reajuste = %s, novo_valor_parcela = %s, observacao = %s
                WHERE id = %s
                """,
                (contrato_id, data_alteracao, percentual_reajuste, novo_valor, observacao, id),
            )
            cur.execute(
                "UPDATE contratos_aluguel SET valor_parcela = %s WHERE id = %s",
                (novo_valor, contrato_id),
            )
            conn.commit()
            flash("Reajuste atualizado com sucesso!", "success")
            return redirect(url_for("reajustes_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar reajuste: {e}", "danger")
    cur.execute(
        "SELECT * FROM reajustes_contrato WHERE id = %s",
        (id,),
    )
    reajuste = cur.fetchone()
    cur.execute(
        "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel ORDER BY id"
    )
    contratos = cur.fetchall()
    cur.close()
    conn.close()
    if reajuste is None:
        flash("Reajuste não encontrado.", "danger")
        return redirect(url_for("reajustes_list"))
    return render_template(
        "reajustes_contrato/add_list.html", reajuste=reajuste, contratos=contratos
    )


@app.route("/reajustes/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Excluir")
def reajustes_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM reajustes_contrato WHERE id = %s", (id,))
        conn.commit()
        flash("Reajuste excluído com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir reajuste: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("reajustes_list"))


@app.route("/reajustes/contrato/<int:contrato_id>")
@login_required
def contrato_info(contrato_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        "SELECT id, nome_inquilino, valor_parcela FROM contratos_aluguel WHERE id = %s",
        (contrato_id,),
    )
    contrato = cur.fetchone()
    cur.close()
    conn.close()
    if contrato:
        return {
            "id": contrato["id"],
            "nome_inquilino": contrato["nome_inquilino"],
            "valor_parcela": float(contrato["valor_parcela"]),
        }
    return {}, 404


# --- Módulo Financeiro ---
@app.route("/contas-a-receber")
@login_required
@permission_required("Financeiro", "Consultar")
def contas_a_receber_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_receber(cur)
    conn.commit()
    # Coleta filtros (GET)
    venc_inicio = request.args.get("venc_inicio")
    venc_fim = request.args.get("venc_fim")
    cliente_id = request.args.get("cliente_id")
    receita_id = request.args.get("receita_id")
    imovel_id = request.args.get("imovel_id")
    status_conta = request.args.get("status_conta")

    where = []
    params = []
    if venc_inicio:
        where.append("cr.data_vencimento >= %s")
        params.append(venc_inicio)
    if venc_fim:
        where.append("cr.data_vencimento <= %s")
        params.append(venc_fim)
    if cliente_id:
        where.append("cr.cliente_id = %s")
        params.append(cliente_id)
    if receita_id:
        where.append("cr.receita_id = %s")
        params.append(receita_id)
    if imovel_id:
        where.append("ca.imovel_id = %s")
        params.append(imovel_id)
    if status_conta:
        # Enum: força cast para status_conta_enum
        where.append("cr.status_conta = %s::status_conta_enum")
        params.append(status_conta)

    sql = (
        """
        SELECT cr.*, p.razao_social_nome AS cliente, r.descricao AS receita,
               ca.imovel_id AS imovel_id,
               NULLIF(CONCAT_WS(' - ', i.endereco, i.bairro), '') AS imovel_descricao
        FROM contas_a_receber cr
        JOIN pessoas p ON cr.cliente_id = p.id
        JOIN receitas_cadastro r ON cr.receita_id = r.id
        LEFT JOIN contratos_aluguel ca ON cr.contrato_id = ca.id
        LEFT JOIN imoveis i ON ca.imovel_id = i.id
        """
    )
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY cr.data_vencimento DESC"

    cur.execute(sql, tuple(params))
    contas = cur.fetchall()
    # Listas para filtros (usar mesma conexão antes de fechar)
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    clientes = cur.fetchall()
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco, bairro")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    contas_caixa = ContaCaixa.query.all()
    contas_banco = ContaBanco.query.all()

    filtros = {
        "venc_inicio": venc_inicio or "",
        "venc_fim": venc_fim or "",
        "cliente_id": (cliente_id or ""),
        "receita_id": (receita_id or ""),
        "imovel_id": (imovel_id or ""),
        "status_conta": (status_conta or ""),
    }

    return render_template(
        "financeiro/contas_a_receber/list.html",
        contas=contas,
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
        clientes=clientes,
        receitas=receitas,
        imoveis=imoveis,
        filtros=filtros,
    )


@app.route("/contas-a-receber/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def contas_a_receber_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            contrato_id = request.form.get("contrato_id") or None
            receita_id = request.form["receita_id"]
            cliente_id = request.form["cliente_id"]
            titulo = request.form.get("titulo") or None
            data_vencimento = request.form["data_vencimento"]
            valor_previsto = parse_decimal(request.form["valor_previsto"])  # BRL mask-aware
            data_pagamento = request.form.get("data_pagamento") or None
            valor_pago = parse_decimal(request.form.get("valor_pago"))
            valor_prev_dec = valor_previsto or Decimal('0')
            valor_pago_dec = valor_pago or Decimal('0')
            valor_pendente = valor_prev_dec - valor_pago_dec
            valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
            valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
            valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
            observacao = request.form.get("observacao")
            origem_id = request.form.get("origem_id") or None
            status_conta = calcular_status_conta(
                data_vencimento, data_pagamento, contrato_id, cur
            )
            if valor_pago is not None:
                if valor_pendente <= 0:
                    status_conta = "Paga"
                    valor_pendente = Decimal('0')
                elif valor_pago_dec > 0:
                    status_conta = "Parcial"

            cur.execute(
                """
                INSERT INTO contas_a_receber (
                    contrato_id, receita_id, cliente_id, titulo,
                    data_vencimento, valor_previsto, data_pagamento, valor_pago,
                    valor_pendente, valor_desconto, valor_multa, valor_juros, observacao,
                    status_conta, origem_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    titulo,
                    data_vencimento,
                    valor_previsto,
                    data_pagamento,
                    valor_pago,
                    valor_pendente,
                    valor_desconto,
                    valor_multa,
                    valor_juros,
                    observacao,
                    status_conta,
                    origem_id,
                ),
            )
            conn.commit()
            flash("Conta a receber cadastrada com sucesso!", "success")
            return redirect(url_for("contas_a_receber_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar conta: {e}", "danger")
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    clientes = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/contas_a_receber/add_list.html",
        conta=request.form,
        receitas=receitas,
        clientes=clientes,
        origens=origens,
    )


@app.route("/contas-a-receber/view/<int:id>")
@login_required
@permission_required("Financeiro", "Consultar")
def contas_a_receber_view(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_receber(cur)
    conn.commit()
    cur.execute(
        """
        SELECT cr.*, p.razao_social_nome AS cliente, r.descricao AS receita, o.descricao AS origem
        FROM contas_a_receber cr
        JOIN pessoas p ON cr.cliente_id = p.id
        JOIN receitas_cadastro r ON cr.receita_id = r.id
        LEFT JOIN origens_cadastro o ON cr.origem_id = o.id
        WHERE cr.id = %s
        """,
        (id,),
    )
    conta = cur.fetchone()
    cur.close()
    conn.close()
    if not conta:
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_receber_list"))
    return render_template(
        "financeiro/contas_a_receber/view.html",
        conta=conta,
    )


@app.route("/contas-a-receber/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def contas_a_receber_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            contrato_id = request.form.get("contrato_id") or None
            receita_id = request.form["receita_id"]
            cliente_id = request.form["cliente_id"]
            titulo = request.form.get("titulo") or None
            data_vencimento = request.form["data_vencimento"]
            valor_previsto = parse_decimal(request.form["valor_previsto"])  # BRL mask-aware
            data_pagamento = request.form.get("data_pagamento") or None
            valor_pago = parse_decimal(request.form.get("valor_pago"))
            valor_prev_dec = valor_previsto or Decimal('0')
            valor_pago_dec = valor_pago or Decimal('0')
            valor_pendente = valor_prev_dec - valor_pago_dec
            valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
            valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
            valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
            observacao = request.form.get("observacao")
            origem_id = request.form.get("origem_id") or None
            status_conta = calcular_status_conta(
                data_vencimento, data_pagamento, contrato_id, cur
            )
            if valor_pago is not None:
                if valor_pendente <= 0:
                    status_conta = "Paga"
                    valor_pendente = Decimal('0')
                elif valor_pago_dec > 0:
                    status_conta = "Parcial"

            cur.execute(
                """
                UPDATE contas_a_receber
                SET contrato_id=%s, receita_id=%s, cliente_id=%s, titulo=%s,
                    data_vencimento=%s, valor_previsto=%s, data_pagamento=%s,
                    valor_pago=%s, valor_pendente=%s, valor_desconto=%s, valor_multa=%s, valor_juros=%s,
                    observacao=%s, status_conta=%s, origem_id=%s
                WHERE id=%s
                """,
                (
                    contrato_id,
                    receita_id,
                    cliente_id,
                    titulo,
                    data_vencimento,
                    valor_previsto,
                    data_pagamento,
                    valor_pago,
                    valor_pendente,
                    valor_desconto,
                    valor_multa,
                    valor_juros,
                    observacao,
                    status_conta,
                    origem_id,
                    id,
                ),
            )
            conn.commit()
            flash("Conta a receber atualizada com sucesso!", "success")
            return redirect(url_for("contas_a_receber_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar conta: {e}", "danger")
    atualizar_status_contas_a_receber(cur)
    conn.commit()
    cur.execute("SELECT * FROM contas_a_receber WHERE id = %s", (id,))
    conta = cur.fetchone()
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Cliente', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    clientes = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.close()
    conn.close()
    if conta is None:
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_receber_list"))
    return render_template(
        "financeiro/contas_a_receber/add_list.html",
        conta=conta,
        receitas=receitas,
        clientes=clientes,
        origens=origens,
    )


@app.route("/contas-a-receber/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Excluir")
def contas_a_receber_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM contas_a_receber WHERE id = %s", (id,))
        conn.commit()
        flash("Conta excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir conta: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contas_a_receber_list"))


@app.route("/contas-a-receber/pagar/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def contas_a_receber_pagar(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_receber(cur)
    conn.commit()
    cur.execute(
        """
        SELECT cr.*, p.razao_social_nome AS cliente_nome
        FROM contas_a_receber cr
        JOIN pessoas p ON cr.cliente_id = p.id
        WHERE cr.id = %s
        """,
        (id,),
    )
    conta = cur.fetchone()
    if not conta:
        cur.close()
        conn.close()
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_receber_list"))
    # Evita pagamento duplicado ou de títulos cancelados
    if conta["status_conta"] not in ("Aberta", "Vencida", "Parcial"):
        cur.close()
        conn.close()
        flash("Esta conta não pode ser paga novamente.", "warning")
        return redirect(url_for("contas_a_receber_list"))
    try:
        conta_tipo = request.form["conta_tipo"]
        conta_id = int(request.form["conta_id"])
        valor_previsto = parse_decimal(request.form.get("valor_previsto")) or parse_decimal(conta["valor_previsto"])
        valor_pagamento = parse_decimal(request.form.get("valor_pago")) or valor_previsto
        valor_pago_atual = parse_decimal(conta["valor_pago"]) or Decimal("0")
        total_pago = valor_pago_atual + valor_pagamento
        valor_pendente = parse_decimal(conta["valor_previsto"]) - total_pago
        status = "Paga" if valor_pendente <= 0 else "Parcial"
        valor_pendente = Decimal('0') if valor_pendente < 0 else valor_pendente
        valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
        valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
        valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
        data_movimento = request.form.get("data_movimento") or datetime.today().date()
        titulo = conta["titulo"] or ""
        historico_padrao = (
            f"{titulo} - {conta['cliente_nome']}" if titulo else conta["cliente_nome"]
        )
        historico = request.form.get("historico") or historico_padrao

        cur.execute(
            """UPDATE contas_a_receber SET data_pagamento=%s, valor_pago=%s, valor_pendente=%s, valor_desconto=%s,
                valor_multa=%s, valor_juros=%s, status_conta=%s WHERE id=%s""",
            (
                data_movimento,
                total_pago,
                valor_pendente,
                valor_desconto,
                valor_multa,
                valor_juros,
                status,
                id,
            ),
        )
        conn.commit()
        
        valor_total = valor_pagamento + valor_juros + valor_multa - valor_desconto
        data = {
            "conta_origem_id": conta_id,
            "conta_origem_tipo": conta_tipo,
            "tipo": "entrada",
            "valor": valor_total,
            "historico": historico,
            "receita_id": conta["receita_id"],
            "data_movimento": data_movimento,
            "valor_previsto": valor_previsto,
            "valor_pago": valor_pagamento,
            "valor_desconto": valor_desconto,
            "valor_multa": valor_multa,
            "valor_juros": valor_juros,
            # associa o lançamento à conta a receber para permitir reversão
            "documento": f"CR-{id}",
        }
        criar_movimento(data)
        cur.close()
        conn.close()
        flash("Pagamento registrado com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        flash(f"Erro ao registrar pagamento: {e}", "danger")
    return redirect(url_for("contas_a_receber_list"))


@app.route("/contas-a-pagar")
@login_required
@permission_required("Financeiro", "Consultar")
def contas_a_pagar_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_pagar(cur)
    conn.commit()

    # Filtros
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    fornecedor_id = request.args.get("fornecedor_id")
    despesa_id = request.args.get("despesa_id")
    imovel_id = request.args.get("imovel_id")
    status_conta = request.args.get("status_conta")

    query = (
        "SELECT cp.*, p.razao_social_nome AS fornecedor, d.descricao AS despesa, "
        "       i.endereco AS imovel_endereco, i.bairro AS imovel_bairro "
        "FROM contas_a_pagar cp "
        "JOIN pessoas p ON cp.fornecedor_id = p.id "
        "JOIN despesas_cadastro d ON cp.despesa_id = d.id "
        "LEFT JOIN imoveis i ON cp.imovel_id = i.id "
    )
    where = []
    params = []

    if data_inicio and data_fim:
        where.append("cp.data_vencimento BETWEEN %s AND %s")
        params.extend([data_inicio, data_fim])
    elif data_inicio:
        where.append("cp.data_vencimento >= %s")
        params.append(data_inicio)
    elif data_fim:
        where.append("cp.data_vencimento <= %s")
        params.append(data_fim)

    if fornecedor_id:
        where.append("cp.fornecedor_id = %s")
        params.append(fornecedor_id)

    if despesa_id:
        where.append("cp.despesa_id = %s")
        params.append(despesa_id)
    if imovel_id:
        where.append("cp.imovel_id = %s")
        params.append(imovel_id)

    if status_conta:
        where.append("cp.status_conta = %s::status_conta_enum")
        params.append(status_conta)

    if where:
        query += " WHERE " + " AND ".join(where)

    query += " ORDER BY cp.data_vencimento DESC"

    cur.execute(query, params)
    contas = cur.fetchall()

    # Opções para filtros
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Fornecedor', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco, bairro")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()

    contas_caixa = ContaCaixa.query.all()
    contas_banco = ContaBanco.query.all()
    filtros = {
        "data_inicio": data_inicio or "",
        "data_fim": data_fim or "",
        "fornecedor_id": fornecedor_id or "",
        "despesa_id": despesa_id or "",
        "imovel_id": imovel_id or "",
        "status_conta": status_conta or "",
    }
    return render_template(
        "financeiro/contas_a_pagar/list.html",
        contas=contas,
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
        fornecedores=fornecedores,
        despesas=despesas,
        imoveis=imoveis,
        filtros=filtros,
    )


@app.route("/contas-a-pagar/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def contas_a_pagar_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            despesa_id = request.form["despesa_id"]
            fornecedor_id = request.form["fornecedor_id"]
            titulo = request.form["titulo"]
            data_vencimento = request.form["data_vencimento"]
            competencia_str = request.form.get("competencia")
            competencia = (
                datetime.strptime(competencia_str, "%m/%Y").date().replace(day=1)
                if competencia_str
                else None
            )
            valor_previsto = parse_decimal(request.form["valor_previsto"])  # BRL mask-aware
            data_pagamento = request.form.get("data_pagamento") or None
            valor_pago = parse_decimal(request.form.get("valor_pago"))
            valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
            valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
            valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
            observacao = request.form.get("observacao")
            imovel_id = parse_int(request.form.get("imovel_id"))
            origem_id = parse_int(request.form.get("origem_id"))
            status_conta = calcular_status_conta(
                data_vencimento, data_pagamento, None, cur
            )

            cur.execute(
                """
                INSERT INTO contas_a_pagar (
                    despesa_id, fornecedor_id, titulo, data_vencimento,
                    competencia, valor_previsto, data_pagamento, valor_pago,
                    valor_desconto, valor_multa, valor_juros, observacao,
                    imovel_id, status_conta, origem_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    despesa_id,
                    fornecedor_id,
                    titulo,
                    data_vencimento,
                    competencia,
                    valor_previsto,
                    data_pagamento,
                    valor_pago,
                    valor_desconto,
                    valor_multa,
                    valor_juros,
                    observacao,
                    imovel_id,
                    status_conta,
                    origem_id,
                ),
            )
            conn.commit()
            flash("Conta a pagar cadastrada com sucesso!", "success")
            return redirect(url_for("contas_a_pagar_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar conta: {e}", "danger")
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Fornecedor', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/contas_a_pagar/add_list.html",
        conta=request.form,
        despesas=despesas,
        fornecedores=fornecedores,
        origens=origens,
        imoveis=imoveis,
    )


@app.route("/contas-a-pagar/view/<int:id>")
@login_required
@permission_required("Financeiro", "Consultar")
def contas_a_pagar_view(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_pagar(cur)
    conn.commit()
    cur.execute(
        """
        SELECT cp.*, p.razao_social_nome AS fornecedor, d.descricao AS despesa,
               o.descricao AS origem, (i.endereco || ', ' || i.bairro) AS imovel
        FROM contas_a_pagar cp
        JOIN pessoas p ON cp.fornecedor_id = p.id
        JOIN despesas_cadastro d ON cp.despesa_id = d.id
        LEFT JOIN origens_cadastro o ON cp.origem_id = o.id
        LEFT JOIN imoveis i ON cp.imovel_id = i.id
        WHERE cp.id = %s
        """,
        (id,),
    )
    conta = cur.fetchone()
    cur.close()
    conn.close()
    if not conta:
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_pagar_list"))
    return render_template(
        "financeiro/contas_a_pagar/view.html",
        conta=conta,
    )


@app.route("/contas-a-pagar/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def contas_a_pagar_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            despesa_id = request.form["despesa_id"]
            fornecedor_id = request.form["fornecedor_id"]
            titulo = request.form["titulo"]
            data_vencimento = request.form["data_vencimento"]
            competencia_str = request.form.get("competencia")
            competencia = (
                datetime.strptime(competencia_str, "%m/%Y").date().replace(day=1)
                if competencia_str
                else None
            )
            valor_previsto = parse_decimal(request.form["valor_previsto"])  # BRL mask-aware
            data_pagamento = request.form.get("data_pagamento") or None
            valor_pago = parse_decimal(request.form.get("valor_pago"))
            valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
            valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
            valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
            observacao = request.form.get("observacao")
            imovel_id = parse_int(request.form.get("imovel_id"))
            origem_id = parse_int(request.form.get("origem_id"))
            status_conta = calcular_status_conta(
                data_vencimento, data_pagamento, None, cur
            )

            cur.execute(
                """
                UPDATE contas_a_pagar
                SET despesa_id=%s, fornecedor_id=%s, titulo=%s, data_vencimento=%s,
                    competencia=%s, valor_previsto=%s, data_pagamento=%s, valor_pago=%s,
                    valor_desconto=%s, valor_multa=%s, valor_juros=%s,
                    observacao=%s, imovel_id=%s, status_conta=%s, origem_id=%s
                WHERE id=%s
                """,
                (
                    despesa_id,
                    fornecedor_id,
                    titulo,
                    data_vencimento,
                    competencia,
                    valor_previsto,
                    data_pagamento,
                    valor_pago,
                    valor_desconto,
                    valor_multa,
                    valor_juros,
                    observacao,
                    imovel_id,
                    status_conta,
                    origem_id,
                    id,
                ),
            )
            conn.commit()
            flash("Conta a pagar atualizada com sucesso!", "success")
            return redirect(url_for("contas_a_pagar_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar conta: {e}", "danger")
    cur.execute("SELECT * FROM contas_a_pagar WHERE id = %s", (id,))
    conta = cur.fetchone()
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Fornecedor', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.execute("SELECT id, endereco, bairro FROM imoveis ORDER BY endereco")
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    if conta is None:
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_pagar_list"))
    return render_template(
        "financeiro/contas_a_pagar/add_list.html",
        conta=conta,
        despesas=despesas,
        fornecedores=fornecedores,
        origens=origens,
        imoveis=imoveis,
    )


@app.route("/contas-a-pagar/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Excluir")
def contas_a_pagar_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM contas_a_pagar WHERE id = %s", (id,))
        conn.commit()
        flash("Conta excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir conta: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contas_a_pagar_list"))


@app.route("/contas-a-pagar/replicar/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def contas_a_pagar_replicar(id):
    quantidade = int(request.form.get("quantidade", 0))
    same_day = int(request.form.get("same_day", 0) or 0)
    days_interval = int(request.form.get("days_interval", 0) or 0)
    if quantidade < 1:
        flash("Quantidade inválida.", "warning")
        return redirect(url_for("contas_a_pagar_list"))
    if (same_day > 0 and days_interval > 0) or (same_day == 0 and days_interval == 0):
        flash(
            "Informe apenas 'Vencimento no mesmo dia' ou 'Dias e intervalo'.",
            "warning",
        )
        return redirect(url_for("contas_a_pagar_list"))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM contas_a_pagar WHERE id = %s", (id,))
    conta = cur.fetchone()
    if not conta:
        cur.close()
        conn.close()
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_pagar_list"))
    try:
        if same_day > 0:
            base_date = conta["data_vencimento"]
            last_day = calendar.monthrange(base_date.year, base_date.month)[1]
            base_day = min(same_day, last_day)
            base_date = base_date.replace(day=base_day)
            for i in range(1, quantidade + 1):
                novo_vencimento = add_months(base_date, i)
                nova_competencia = (
                    add_months(conta["competencia"], i)
                    if conta["competencia"]
                    else None
                )
                match = re.match(r"^(.*?)-(\s*)(\d+)\s*/\s*(\d+)$", conta["titulo"])
                if match:
                    prefixo, espaco, numero, total = match.groups()
                    novo_titulo = f"{prefixo}-{espaco}{int(numero) + i}/{total}"
                else:
                    novo_titulo = f"{conta['titulo']}-{i}"
                status_conta = calcular_status_conta(
                    novo_vencimento.strftime("%Y-%m-%d"), None, None, cur
                )
                cur.execute(
                    """
                    INSERT INTO contas_a_pagar (
                        despesa_id, fornecedor_id, titulo, data_vencimento,
                        competencia, valor_previsto, data_pagamento, valor_pago, valor_desconto,
                        valor_multa, valor_juros, observacao, imovel_id,
                        status_conta, origem_id
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        conta["despesa_id"],
                        conta["fornecedor_id"],
                        novo_titulo,
                        novo_vencimento,
                        nova_competencia,
                        conta["valor_previsto"],
                        None,
                        None,
                        0,
                        0,
                        0,
                        conta["observacao"],
                        conta["imovel_id"],
                        status_conta,
                        conta["origem_id"],
                    ),
                )
        else:
            novo_vencimento = conta["data_vencimento"]
            for i in range(1, quantidade + 1):
                novo_vencimento = novo_vencimento + timedelta(days=days_interval)
                nova_competencia = (
                    add_months(conta["competencia"], i)
                    if conta["competencia"]
                    else None
                )
                match = re.match(r"^(.*?)-(\s*)(\d+)\s*/\s*(\d+)$", conta["titulo"])
                if match:
                    prefixo, espaco, numero, total = match.groups()
                    novo_titulo = f"{prefixo}-{espaco}{int(numero) + i}/{total}"
                else:
                    novo_titulo = f"{conta['titulo']}-{i}"
                status_conta = calcular_status_conta(
                    novo_vencimento.strftime("%Y-%m-%d"), None, None, cur
                )
                cur.execute(
                    """
                    INSERT INTO contas_a_pagar (
                        despesa_id, fornecedor_id, titulo, data_vencimento,
                        competencia, valor_previsto, data_pagamento, valor_pago, valor_desconto,
                        valor_multa, valor_juros, observacao, imovel_id,
                        status_conta, origem_id
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        conta["despesa_id"],
                        conta["fornecedor_id"],
                        novo_titulo,
                        novo_vencimento,
                        nova_competencia,
                        conta["valor_previsto"],
                        None,
                        None,
                        0,
                        0,
                        0,
                        conta["observacao"],
                        conta["imovel_id"],
                        status_conta,
                        conta["origem_id"],
                    ),
                )
        conn.commit()
        flash("Títulos replicados com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao replicar títulos: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("contas_a_pagar_list"))


@app.route("/contas-a-pagar/pagar/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def contas_a_pagar_pagar(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    atualizar_status_contas_a_pagar(cur)
    conn.commit()
    cur.execute(
        """
        SELECT cp.*, p.razao_social_nome AS fornecedor
        FROM contas_a_pagar cp
        JOIN pessoas p ON cp.fornecedor_id = p.id
        WHERE cp.id = %s
        """,
        (id,),
    )
    conta = cur.fetchone()
    if not conta:
        cur.close()
        conn.close()
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("contas_a_pagar_list"))
    if conta["status_conta"] not in ("Aberta", "Vencida"):
        cur.close()
        conn.close()
        flash("Esta conta não pode ser paga novamente.", "warning")
        return redirect(url_for("contas_a_pagar_list"))
    try:
        conta_tipo = request.form["conta_tipo"]
        conta_id = int(request.form["conta_id"])
        valor_previsto = parse_decimal(request.form.get("valor_previsto")) or parse_decimal(conta["valor_previsto"])
        valor_pago = parse_decimal(request.form.get("valor_pago")) or valor_previsto
        valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
        valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
        valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
        data_movimento = request.form.get("data_movimento") or datetime.today().date()
        titulo = conta["titulo"] or ""
        historico_padrao = f"{titulo} - {conta['fornecedor']}" if titulo else conta["fornecedor"]
        historico = request.form.get("historico") or historico_padrao

        cur.execute(
            """UPDATE contas_a_pagar SET data_pagamento=%s, valor_pago=%s, valor_desconto=%s,
                valor_multa=%s, valor_juros=%s, status_conta='Paga' WHERE id=%s""",
            (
                data_movimento,
                valor_pago,
                valor_desconto,
                valor_multa,
                valor_juros,
                id,
            ),
        )
        conn.commit()

        valor_total = valor_pago + valor_juros + valor_multa - valor_desconto
        data = {
            "conta_origem_id": conta_id,
            "conta_origem_tipo": conta_tipo,
            "tipo": "saida",
            "valor": valor_total,
            "historico": historico,
            "despesa_id": conta["despesa_id"],
            "data_movimento": data_movimento,
            "valor_previsto": valor_previsto,
            "valor_pago": valor_pago,
            "valor_desconto": valor_desconto,
            "valor_multa": valor_multa,
            "valor_juros": valor_juros,
            "documento": f"CP-{id}",
        }
        criar_movimento(data)
        cur.close()
        conn.close()
        flash("Pagamento registrado com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        flash(f"Erro ao registrar pagamento: {e}", "danger")
    return redirect(url_for("contas_a_pagar_list"))


# --- Módulo Caixa e Banco ---

# --- Ordens de Pagamento ---


@app.route("/ordens-pagamento")
@login_required
@permission_required("Financeiro", "Consultar")
def ordens_pagamento_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        """
        SELECT op.*, p.razao_social_nome AS fornecedor_nome, p.documento AS fornecedor_documento,
               i.endereco AS imovel_endereco
          FROM ordens_pagamento op
          JOIN pessoas p ON p.id = op.fornecedor_id
          LEFT JOIN imoveis i ON i.id = op.imovel_id
         ORDER BY op.data_emissao DESC, op.id DESC
        """
    )
    ordens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("financeiro/ordens_pagamento/list.html", ordens=ordens)


@app.route("/ordens-pagamento/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def ordens_pagamento_add():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = parse_int(request.form.get("imovel_id"))
            fornecedor_id = int(request.form.get("fornecedor_id"))
            forma_pagamento = request.form.get("forma_pagamento")
            data_vencimento = request.form.get("data_vencimento") or None
            descricao_servico = request.form.get("descricao_servico")
            valor_servico = parse_decimal(request.form.get("valor_servico")) or Decimal("0")
            desconto_manual = parse_decimal(request.form.get("desconto_manual")) or Decimal("0")
            observacoes = request.form.get("observacoes")
            despesa_id = parse_int(request.form.get("despesa_id"))
            origem_id = parse_int(request.form.get("origem_id"))

            if not despesa_id:
                raise ValueError("Selecione uma despesa para a ordem de pagamento.")

            codigos = request.form.getlist("item_codigo[]")
            descrs = request.form.getlist("item_descricao[]")
            qtes = request.form.getlist("item_quantidade[]")
            units = request.form.getlist("item_valor_unitario[]")
            descs = request.form.getlist("item_desconto[]")

            itens = []
            subtotal_produtos = Decimal("0")
            desconto_produtos = Decimal("0")
            for idx in range(len(codigos)):
                codigo = (codigos[idx] or "").strip()
                descricao = (descrs[idx] or "").strip()
                qtd = parse_decimal(qtes[idx]) or Decimal("0")
                unit = parse_decimal(units[idx]) or Decimal("0")
                desconto = parse_decimal(descs[idx]) or Decimal("0")
                total_item = (qtd * unit)
                subtotal_produtos += total_item
                desconto_produtos += desconto
                if any([codigo, descricao]) or total_item > 0:
                    itens.append({
                        "codigo": codigo,
                        "descricao": descricao,
                        "quantidade": qtd,
                        "valor_unitario": unit,
                        "valor_desconto": desconto,
                        "valor_total": total_item,
                    })

            subtotal_servicos = valor_servico
            total = subtotal_servicos + subtotal_produtos - desconto_produtos - desconto_manual

            cur.execute(
                """
                INSERT INTO ordens_pagamento (
                    imovel_id, fornecedor_id, forma_pagamento, data_vencimento,
                    descricao_servico, valor_servico, desconto_manual,
                    subtotal_servicos, subtotal_produtos, desconto_produtos, total, observacoes,
                    despesa_id, origem_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                """,
                (
                    imovel_id,
                    fornecedor_id,
                    forma_pagamento,
                    data_vencimento,
                    descricao_servico,
                    valor_servico,
                    desconto_manual,
                    subtotal_servicos,
                    subtotal_produtos,
                    desconto_produtos,
                    total,
                    observacoes,
                    despesa_id,
                    origem_id,
                ),
            )
            ordem_id = cur.fetchone()[0]
            for it in itens:
                cur.execute(
                    """
                    INSERT INTO ordem_pagamento_itens (
                        ordem_id, codigo_produto, descricao_produto, quantidade,
                        valor_unitario, valor_desconto, valor_total
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        ordem_id,
                        it["codigo"],
                        it["descricao"],
                        it["quantidade"],
                        it["valor_unitario"],
                        it["valor_desconto"],
                        it["valor_total"],
                    ),
                )
            # Parcelas (títulos)
            parc_nums = request.form.getlist("parcela_numero[]")
            parc_vencs = request.form.getlist("parcela_vencimento[]")
            parc_vals = request.form.getlist("parcela_valor[]")
            for idx in range(max(len(parc_vencs), len(parc_vals))):
                try:
                    n = parse_int(parc_nums[idx]) if idx < len(parc_nums) else (idx + 1)
                except Exception:
                    n = idx + 1
                venc = parc_vencs[idx] if idx < len(parc_vencs) else None
                val = parse_decimal(parc_vals[idx]) if idx < len(parc_vals) else None
                if venc and (val is not None):
                    cur.execute(
                        """
                        INSERT INTO ordem_pagamento_parcelas (ordem_id, numero, data_vencimento, valor)
                        VALUES (%s,%s,%s,%s)
                        """,
                        (ordem_id, n, venc, val),
                    )
            conn.commit()
            cur.close()
            conn.close()
            flash("Ordem de Pagamento salva com sucesso!", "success")
            return redirect(url_for("ordens_pagamento_view", id=ordem_id))
        except Exception as e:
            conn.rollback()
            cur.close()
            conn.close()
            flash(f"Erro ao salvar a Ordem de Pagamento: {e}", "danger")
            return redirect(url_for("ordens_pagamento_add"))

    # GET
    cur.execute(
        "SELECT id, endereco, bairro FROM imoveis ORDER BY endereco"
    )
    imoveis = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome, documento FROM pessoas WHERE tipo IN ('Fornecedor','Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/ordens_pagamento/add_edit.html",
        ordem=None,
        itens=[],
        imoveis=imoveis,
        fornecedores=fornecedores,
        despesas=despesas,
        origens=origens,
    )


@app.route("/ordens-pagamento/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def ordens_pagamento_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            imovel_id = parse_int(request.form.get("imovel_id"))
            fornecedor_id = int(request.form.get("fornecedor_id"))
            forma_pagamento = request.form.get("forma_pagamento")
            data_vencimento = request.form.get("data_vencimento") or None
            descricao_servico = request.form.get("descricao_servico")
            valor_servico = parse_decimal(request.form.get("valor_servico")) or Decimal("0")
            desconto_manual = parse_decimal(request.form.get("desconto_manual")) or Decimal("0")
            observacoes = request.form.get("observacoes")
            despesa_id = parse_int(request.form.get("despesa_id"))
            origem_id = parse_int(request.form.get("origem_id"))

            if not despesa_id:
                raise ValueError("Selecione uma despesa para a ordem de pagamento.")

            codigos = request.form.getlist("item_codigo[]")
            descrs = request.form.getlist("item_descricao[]")
            qtes = request.form.getlist("item_quantidade[]")
            units = request.form.getlist("item_valor_unitario[]")
            descs = request.form.getlist("item_desconto[]")

            itens = []
            subtotal_produtos = Decimal("0")
            desconto_produtos = Decimal("0")
            for idx in range(len(codigos)):
                codigo = (codigos[idx] or "").strip()
                descricao = (descrs[idx] or "").strip()
                qtd = parse_decimal(qtes[idx]) or Decimal("0")
                unit = parse_decimal(units[idx]) or Decimal("0")
                desconto = parse_decimal(descs[idx]) or Decimal("0")
                total_item = (qtd * unit)
                subtotal_produtos += total_item
                desconto_produtos += desconto
                if any([codigo, descricao]) or total_item > 0:
                    itens.append({
                        "codigo": codigo,
                        "descricao": descricao,
                        "quantidade": qtd,
                        "valor_unitario": unit,
                        "valor_desconto": desconto,
                        "valor_total": total_item,
                    })

            subtotal_servicos = valor_servico
            total = subtotal_servicos + subtotal_produtos - desconto_produtos - desconto_manual

            cur.execute(
                """
                UPDATE ordens_pagamento SET
                    imovel_id=%s, fornecedor_id=%s, forma_pagamento=%s, data_vencimento=%s,
                    descricao_servico=%s, valor_servico=%s, desconto_manual=%s,
                    subtotal_servicos=%s, subtotal_produtos=%s, desconto_produtos=%s, total=%s, observacoes=%s,
                    despesa_id=%s, origem_id=%s
                WHERE id=%s
                """,
                (
                    imovel_id,
                    fornecedor_id,
                    forma_pagamento,
                    data_vencimento,
                    descricao_servico,
                    valor_servico,
                    desconto_manual,
                    subtotal_servicos,
                    subtotal_produtos,
                    desconto_produtos,
                    total,
                    observacoes,
                    despesa_id,
                    origem_id,
                    id,
                ),
            )
            # Remove itens antigos e insere os novos
            cur.execute("DELETE FROM ordem_pagamento_itens WHERE ordem_id = %s", (id,))
            for it in itens:
                cur.execute(
                    """
                    INSERT INTO ordem_pagamento_itens (
                        ordem_id, codigo_produto, descricao_produto, quantidade,
                        valor_unitario, valor_desconto, valor_total
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        id,
                        it["codigo"],
                        it["descricao"],
                        it["quantidade"],
                        it["valor_unitario"],
                        it["valor_desconto"],
                        it["valor_total"],
                    ),
                )
            # Remove parcelas antigas e insere as novas
            cur.execute("DELETE FROM ordem_pagamento_parcelas WHERE ordem_id = %s", (id,))
            parc_nums = request.form.getlist("parcela_numero[]")
            parc_vencs = request.form.getlist("parcela_vencimento[]")
            parc_vals = request.form.getlist("parcela_valor[]")
            for idx in range(max(len(parc_vencs), len(parc_vals))):
                try:
                    n = parse_int(parc_nums[idx]) if idx < len(parc_nums) else (idx + 1)
                except Exception:
                    n = idx + 1
                venc = parc_vencs[idx] if idx < len(parc_vencs) else None
                val = parse_decimal(parc_vals[idx]) if idx < len(parc_vals) else None
                if venc and (val is not None):
                    cur.execute(
                        """
                        INSERT INTO ordem_pagamento_parcelas (ordem_id, numero, data_vencimento, valor)
                        VALUES (%s,%s,%s,%s)
                        """,
                        (id, n, venc, val),
                    )
            conn.commit()
            cur.close()
            conn.close()
            flash("Ordem de Pagamento atualizada com sucesso!", "success")
            return redirect(url_for("ordens_pagamento_view", id=id))
        except Exception as e:
            conn.rollback()
            cur.close()
            conn.close()
            flash(f"Erro ao atualizar a Ordem de Pagamento: {e}", "danger")
            return redirect(url_for("ordens_pagamento_edit", id=id))

    # GET
    cur.execute(
        """
        SELECT op.*, p.razao_social_nome AS fornecedor_nome, p.documento AS fornecedor_documento,
               i.endereco AS imovel_endereco, d.descricao AS despesa_descricao,
               o.descricao AS origem_descricao
          FROM ordens_pagamento op
          JOIN pessoas p ON p.id = op.fornecedor_id
          LEFT JOIN imoveis i ON i.id = op.imovel_id
          LEFT JOIN despesas_cadastro d ON d.id = op.despesa_id
          LEFT JOIN origens_cadastro o ON o.id = op.origem_id
         WHERE op.id = %s
        """,
        (id,),
    )
    ordem = cur.fetchone()
    if not ordem:
        cur.close()
        conn.close()
        flash("Ordem de Pagamento não encontrada.", "danger")
        return redirect(url_for("ordens_pagamento_list"))
    cur.execute(
        "SELECT * FROM ordem_pagamento_itens WHERE ordem_id = %s ORDER BY id",
        (id,),
    )
    itens = cur.fetchall()
    cur.execute(
        "SELECT * FROM ordem_pagamento_parcelas WHERE ordem_id = %s ORDER BY COALESCE(numero, 999999), data_vencimento, id",
        (id,),
    )
    parcelas = cur.fetchall()
    cur.execute(
        "SELECT id, endereco, bairro FROM imoveis ORDER BY endereco"
    )
    imoveis = cur.fetchall()
    cur.execute(
        "SELECT id, razao_social_nome, documento FROM pessoas WHERE tipo IN ('Fornecedor','Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, descricao FROM origens_cadastro ORDER BY descricao")
    origens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/ordens_pagamento/add_edit.html",
        ordem=ordem,
        itens=itens,
        parcelas=parcelas,
        imoveis=imoveis,
        fornecedores=fornecedores,
        despesas=despesas,
        origens=origens,
    )


@app.route("/ordens-pagamento/view/<int:id>")
@login_required
@permission_required("Financeiro", "Consultar")
def ordens_pagamento_view(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        """
        SELECT op.*, p.razao_social_nome AS fornecedor_nome, p.documento AS fornecedor_documento,
               i.endereco AS imovel_endereco, d.descricao AS despesa_descricao,
               o.descricao AS origem_descricao
          FROM ordens_pagamento op
          JOIN pessoas p ON p.id = op.fornecedor_id
          LEFT JOIN imoveis i ON i.id = op.imovel_id
          LEFT JOIN despesas_cadastro d ON d.id = op.despesa_id
          LEFT JOIN origens_cadastro o ON o.id = op.origem_id
         WHERE op.id = %s
        """,
        (id,),
    )
    ordem = cur.fetchone()
    if not ordem:
        cur.close()
        conn.close()
        flash("Ordem de Pagamento não encontrada.", "danger")
        return redirect(url_for("ordens_pagamento_list"))
    cur.execute(
        "SELECT * FROM ordem_pagamento_itens WHERE ordem_id = %s ORDER BY id",
        (id,),
    )
    itens = cur.fetchall()
    cur.execute(
        "SELECT * FROM ordem_pagamento_parcelas WHERE ordem_id = %s ORDER BY COALESCE(numero, 999999), data_vencimento, id",
        (id,),
    )
    parcelas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("financeiro/ordens_pagamento/view.html", ordem=ordem, itens=itens, parcelas=parcelas)


@app.route("/ordens-pagamento/<int:id>/gerar-contas", methods=["POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def ordens_pagamento_gerar_contas(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            """
            SELECT op.*, p.razao_social_nome AS fornecedor_nome
              FROM ordens_pagamento op
              JOIN pessoas p ON p.id = op.fornecedor_id
             WHERE op.id = %s
            """,
            (id,),
        )
        ordem = cur.fetchone()
        if not ordem:
            flash("Ordem de Pagamento não encontrada.", "danger")
            return redirect(url_for("ordens_pagamento_list"))

        ordem = dict(ordem)

        if ordem.get("despesa_id") is None:
            flash("Defina a despesa da ordem antes de gerar contas a pagar.", "warning")
            return redirect(url_for("ordens_pagamento_edit", id=id))

        cur.execute(
            "SELECT id FROM contas_a_pagar WHERE ordem_pagamento_id = %s",
            (id,),
        )
        if cur.fetchone():
            flash("Já existem contas a pagar vinculadas a esta ordem.", "info")
            return redirect(url_for("ordens_pagamento_view", id=id))

        cur.execute(
            """
            SELECT numero, data_vencimento, valor
              FROM ordem_pagamento_parcelas
             WHERE ordem_id = %s
             ORDER BY COALESCE(numero, 999999), data_vencimento, id
            """,
            (id,),
        )
        parcelas_db = cur.fetchall()
        parcelas = [dict(p) for p in parcelas_db]
        if not parcelas:
            data_venc = ordem["data_vencimento"] or date.today()
            parcelas = [
                {
                    "numero": 1,
                    "data_vencimento": data_venc,
                    "valor": ordem.get("total") or Decimal("0"),
                }
            ]
        total_parcelas = len(parcelas)

        observacao_base = (
            "Gerado automaticamente a partir da Ordem de Pagamento #{}".format(ordem["id"])
        )
        if ordem.get("observacoes"):
            observacao_base += f". Observações: {ordem['observacoes']}"

        titulo_base = (ordem.get("descricao_servico") or "").strip()
        if not titulo_base:
            titulo_base = f"Ordem de Pagamento {ordem['id']}"

        for idx, parcela in enumerate(parcelas, start=1):
            numero = parcela.get("numero") or idx
            data_vencimento = parcela["data_vencimento"]
            if isinstance(data_vencimento, str):
                data_vencimento = datetime.strptime(data_vencimento, "%Y-%m-%d").date()
            competencia = date(data_vencimento.year, data_vencimento.month, 1)
            status_conta = calcular_status_conta(
                data_vencimento.strftime("%Y-%m-%d"), None, None, cur
            )
            valor_previsto = parcela.get("valor")
            if valor_previsto is None:
                valor_previsto = Decimal("0")
            else:
                valor_previsto = Decimal(valor_previsto)

            if total_parcelas > 1:
                titulo = f"{titulo_base} - Parcela {numero}/{total_parcelas}"
            else:
                titulo = titulo_base

            cur.execute(
                """
                INSERT INTO contas_a_pagar (
                    despesa_id, fornecedor_id, titulo, data_vencimento,
                    competencia, valor_previsto, data_pagamento, valor_pago,
                    valor_desconto, valor_multa, valor_juros, observacao,
                    imovel_id, status_conta, origem_id, ordem_pagamento_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    ordem["despesa_id"],
                    ordem["fornecedor_id"],
                    titulo,
                    data_vencimento,
                    competencia,
                    valor_previsto,
                    None,
                    None,
                    Decimal("0"),
                    Decimal("0"),
                    Decimal("0"),
                    observacao_base,
                    ordem.get("imovel_id"),
                    status_conta,
                    ordem.get("origem_id"),
                    ordem["id"],
                ),
            )

        conn.commit()
        flash("Contas a pagar geradas com sucesso.", "success")
        return redirect(url_for("contas_a_pagar_list"))
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao gerar contas a pagar: {e}", "danger")
        return redirect(url_for("ordens_pagamento_view", id=id))
    finally:
        cur.close()
        conn.close()


@app.route("/contratos/<int:contrato_id>/prestacao/preview", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_prestacao_preview(contrato_id):
    data_encerramento = parse_date(request.args.get("data_encerramento"))
    if not data_encerramento:
        return jsonify({"error": "Data de encerramento inválida."}), 400

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        contrato = fetch_contrato_info(cur, contrato_id)
        if contrato is None:
            return jsonify({"error": "Contrato não encontrado."}), 404

        creditos = listar_creditos_calcao(cur, contrato_id)
        debitos = listar_debitos_para_prestacao(cur, contrato_id, data_encerramento)
        totais = calcular_totais_prestacao(creditos, debitos, [])

        creditos_serializados = []
        for item in creditos:
            data_pagamento = item["data_pagamento"]
            creditos_serializados.append(
                {
                    "id": item["id"],
                    "titulo": item["titulo"],
                    "valor": decimal_to_string(item["valor"]),
                    "valor_formatado": format_currency(item["valor"]),
                    "valor_previsto": decimal_to_string(item["valor_previsto"]),
                    "data_pagamento": data_pagamento.isoformat()
                    if data_pagamento
                    else None,
                    "data_pagamento_formatada": data_pagamento.strftime("%d/%m/%Y")
                    if data_pagamento
                    else None,
                }
            )

        debitos_serializados = []
        for item in debitos:
            data_vencimento = item["data_vencimento"]
            debitos_serializados.append(
                {
                    "id": item["id"],
                    "titulo": item["titulo"],
                    "valor_previsto": decimal_to_string(item["valor_previsto"]),
                    "valor_previsto_formatado": format_currency(item["valor_previsto"]),
                    "valor_pago": decimal_to_string(item["valor_pago"]),
                    "valor_pago_formatado": format_currency(item["valor_pago"]),
                    "valor_pendente": decimal_to_string(item["valor_pendente"]),
                    "valor_pendente_formatado": format_currency(item["valor_pendente"]),
                    "status": item["status"],
                    "data_vencimento": data_vencimento.isoformat()
                    if data_vencimento
                    else None,
                    "data_vencimento_formatada": data_vencimento.strftime("%d/%m/%Y")
                    if data_vencimento
                    else None,
                }
            )

        response = {
            "contrato": {
                "id": contrato["id"],
                "cliente_nome": contrato.get("cliente_nome"),
                "imovel_id": contrato.get("imovel_id"),
                "data_inicio": contrato.get("data_inicio").isoformat()
                if contrato.get("data_inicio")
                else None,
                "data_fim": contrato.get("data_fim").isoformat()
                if contrato.get("data_fim")
                else None,
            },
            "creditos": creditos_serializados,
            "debitos": debitos_serializados,
            "totais": {
                "total_creditos": decimal_to_string(totais["total_creditos"]),
                "total_creditos_formatado": format_currency(totais["total_creditos"]),
                "total_debitos": decimal_to_string(totais["total_debitos"]),
                "total_debitos_formatado": format_currency(totais["total_debitos"]),
                "total_despesas": decimal_to_string(totais["total_despesas"]),
                "total_despesas_formatado": format_currency(totais["total_despesas"]),
                "saldo_final": decimal_to_string(totais["saldo_final"]),
                "saldo_final_formatado": format_currency(totais["saldo_final"]),
            },
        }
        return jsonify(response)
    finally:
        cur.close()
        conn.close()


@app.route("/contratos/<int:contrato_id>/encerrar", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_encerrar_form(contrato_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        contrato = fetch_contrato_info(cur, contrato_id)
        if contrato is None:
            flash("Contrato não encontrado.", "danger")
            return redirect(url_for("contratos_list"))

        cur.execute(
            """
            SELECT id, data_encerramento, total_creditos, total_debitos,
                   total_despesas, saldo_final, criado_em, atualizado_em
            FROM prestacoes_contas
            WHERE contrato_id = %s
            ORDER BY criado_em DESC
            """,
            (contrato_id,),
        )
        prestacoes = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    data_sugerida = contrato.get("data_fim") or date.today()
    configuracao = {
        "selectedDebitos": [],
        "despesas": [],
        "creditosExtras": [],
        "observacoes": "",
    }
    return render_template(
        "contratos/prestacoes/form.html",
        contrato=contrato,
        prestacao=None,
        prestacoes=prestacoes,
        form_action=url_for("contratos_encerrar", id=contrato_id),
        preview_url=url_for("contratos_prestacao_preview", contrato_id=contrato_id),
        data_encerramento=data_sugerida.strftime("%Y-%m-%d"),
        configuracao=configuracao,
        is_edit=False,
    )


@app.route("/contratos/prestacoes/<int:prestacao_id>/editar", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_prestacao_editar(prestacao_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            """
            SELECT pc.*, c.id AS contrato_id, c.data_fim,
                   p.razao_social_nome AS cliente_nome, p.documento AS cliente_documento,
                   c.cliente_id, c.imovel_id,
                   i.endereco AS imovel_endereco, i.bairro AS imovel_bairro,
                   i.cidade AS imovel_cidade, i.estado AS imovel_estado
            FROM prestacoes_contas pc
            JOIN contratos_aluguel c ON pc.contrato_id = c.id
            JOIN pessoas p ON c.cliente_id = p.id
            JOIN imoveis i ON c.imovel_id = i.id
            WHERE pc.id = %s
            """,
            (prestacao_id,),
        )
        prestacao = cur.fetchone()
        if prestacao is None:
            flash("Prestação de contas não encontrada.", "danger")
            return redirect(url_for("contratos_list"))

        contrato_id = prestacao["contrato_id"]
        contrato = fetch_contrato_info(cur, contrato_id)

        cur.execute(
            """
            SELECT tipo, descricao, valor, conta_receber_id
            FROM prestacoes_contas_itens
            WHERE prestacao_id = %s
            ORDER BY id
            """,
            (prestacao_id,),
        )
        itens = cur.fetchall()

        cur.execute(
            """
            SELECT id, data_encerramento, total_creditos, total_debitos,
                   total_despesas, saldo_final, criado_em, atualizado_em
            FROM prestacoes_contas
            WHERE contrato_id = %s
            ORDER BY criado_em DESC
            """,
            (contrato_id,),
        )
        prestacoes = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    selected_debitos = [item["conta_receber_id"] for item in itens if item["tipo"] == "debito"]
    despesas = [
        {
            "descricao": item["descricao"],
            "valor": decimal_to_string(item["valor"]),
        }
        for item in itens
        if item["tipo"] == "despesa"
    ]
    creditos_extras = [
        {
            "descricao": item["descricao"],
            "valor": decimal_to_string(item["valor"]),
        }
        for item in itens
        if item["tipo"] == "credito_extra"
    ]
    configuracao = {
        "selectedDebitos": selected_debitos,
        "despesas": despesas,
        "creditosExtras": creditos_extras,
        "observacoes": prestacao["observacoes"] or "",
    }
    data_encerramento = prestacao["data_encerramento"].strftime("%Y-%m-%d")

    return render_template(
        "contratos/prestacoes/form.html",
        contrato=contrato,
        prestacao=prestacao,
        prestacoes=prestacoes,
        form_action=url_for("contratos_prestacao_atualizar", prestacao_id=prestacao_id),
        preview_url=url_for("contratos_prestacao_preview", contrato_id=contrato["id"]),
        data_encerramento=data_encerramento,
        configuracao=configuracao,
        is_edit=True,
    )


@app.route("/contratos/prestacoes/<int:prestacao_id>/atualizar", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_prestacao_atualizar(prestacao_id):
    data_encerramento = parse_date(request.form.get("data_encerramento"))
    if not data_encerramento:
        flash("Informe a data de encerramento.", "warning")
        return redirect(request.referrer or url_for("contratos_list"))

    prestacao_raw = request.form.get("prestacao_payload")
    if not prestacao_raw:
        flash("Informe os dados da prestação de contas.", "danger")
        return redirect(request.referrer or url_for("contratos_list"))
    try:
        prestacao_payload = json.loads(prestacao_raw)
    except json.JSONDecodeError:
        flash("Dados da prestação de contas inválidos.", "danger")
        return redirect(request.referrer or url_for("contratos_list"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            "SELECT contrato_id FROM prestacoes_contas WHERE id = %s",
            (prestacao_id,),
        )
        row = cur.fetchone()
        if row is None:
            raise ValueError("Prestação de contas não encontrada.")
        contrato_id = row["contrato_id"]
        contrato = fetch_contrato_info(cur, contrato_id)
        if contrato is None:
            raise ValueError("Contrato relacionado não encontrado.")

        reverter_prestacao(cur, prestacao_id)
        processar_prestacao(
            cur,
            contrato,
            data_encerramento,
            prestacao_payload,
            prestacao_id=prestacao_id,
        )
        conn.commit()
        flash("Prestação de contas atualizada com sucesso!", "success")
        return redirect(url_for("contratos_encerrar_form", contrato_id=contrato_id))
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except Exception as exc:
        conn.rollback()
        flash(f"Erro ao atualizar prestação de contas: {exc}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(request.referrer or url_for("contratos_list"))


@app.route("/contratos/prestacoes/<int:prestacao_id>/excluir", methods=["POST"])
@login_required
@permission_required("Gestao Contratos", "Editar")
def contratos_prestacao_excluir(prestacao_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            "SELECT contrato_id FROM prestacoes_contas WHERE id = %s",
            (prestacao_id,),
        )
        row = cur.fetchone()
        if row is None:
            raise ValueError("Prestação de contas não encontrada.")
        contrato_id = row["contrato_id"]
        reverter_prestacao(cur, prestacao_id)
        cur.execute(
            "DELETE FROM prestacoes_contas WHERE id = %s",
            (prestacao_id,),
        )
        conn.commit()
        flash("Prestação de contas excluída com sucesso!", "success")
        return redirect(url_for("contratos_encerrar_form", contrato_id=contrato_id))
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except Exception as exc:
        conn.rollback()
        flash(f"Erro ao excluir prestação de contas: {exc}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(request.referrer or url_for("contratos_list"))


@app.route("/contratos/prestacoes/<int:prestacao_id>/relatorio", methods=["GET"])
@login_required
@permission_required("Gestao Contratos", "Consultar")
def contratos_prestacao_relatorio(prestacao_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        cur.execute(
            """
            SELECT pc.*, c.id AS contrato_id, c.data_inicio, c.data_fim,
                   p.razao_social_nome AS cliente_nome, p.documento AS cliente_documento,
                   p.endereco AS cliente_endereco, p.cidade AS cliente_cidade,
                   p.estado AS cliente_estado, p.cep AS cliente_cep,
                   i.endereco AS imovel_endereco, i.cidade AS imovel_cidade,
                   i.estado AS imovel_estado
            FROM prestacoes_contas pc
            JOIN contratos_aluguel c ON pc.contrato_id = c.id
            JOIN pessoas p ON c.cliente_id = p.id
            JOIN imoveis i ON c.imovel_id = i.id
            WHERE pc.id = %s
            """,
            (prestacao_id,),
        )
        prestacao = cur.fetchone()
        if prestacao is None:
            flash("Prestação de contas não encontrada.", "danger")
            return redirect(url_for("contratos_list"))

        cur.execute(
            """
            SELECT tipo, descricao, valor, conta_receber_id
            FROM prestacoes_contas_itens
            WHERE prestacao_id = %s
            ORDER BY tipo, id
            """,
            (prestacao_id,),
        )
        itens = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    creditos = [item for item in itens if item["tipo"] == "credito"]
    creditos_extras = [item for item in itens if item["tipo"] == "credito_extra"]
    debitos = [item for item in itens if item["tipo"] == "debito"]
    despesas = [item for item in itens if item["tipo"] == "despesa"]
    total_creditos_extras = sum((Decimal(str(item["valor"])) for item in creditos_extras), Decimal("0"))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Prestação de Contas", ln=True, align="C")

    pdf.set_font("Arial", "", 10)
    pdf.ln(4)
    pdf.cell(0, 6, f"Contrato #{prestacao['contrato_id']} - Cliente: {prestacao['cliente_nome']}", ln=True)
    pdf.cell(0, 6, f"Documento: {prestacao['cliente_documento']}", ln=True)
    pdf.cell(
        0,
        6,
        f"Imóvel: {prestacao['imovel_endereco']} - {prestacao['imovel_cidade']}/{prestacao['imovel_estado']}",
        ln=True,
    )
    pdf.cell(
        0,
        6,
        f"Período do contrato: {prestacao['data_inicio'].strftime('%d/%m/%Y')}"
        f" a {prestacao['data_fim'].strftime('%d/%m/%Y')}",
        ln=True,
    )
    pdf.cell(
        0,
        6,
        f"Data da prestação: {prestacao['data_encerramento'].strftime('%d/%m/%Y')}",
        ln=True,
    )

    def add_section(titulo, registros):
        pdf.ln(6)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, titulo, ln=True)
        pdf.set_font("Arial", "", 10)
        if not registros:
            pdf.cell(0, 6, "Nenhum registro.", ln=True)
            return
        for registro in registros:
            descricao = registro["descricao"]
            valor = Decimal(str(registro["valor"]))
            pdf.cell(0, 6, f"- {descricao}: {format_currency(valor)}", ln=True)

    add_section("Créditos (Calção)", creditos)
    add_section("Creditos Extras", creditos_extras)
    add_section("Débitos", debitos)
    add_section("Despesas Extras", despesas)

    pdf.ln(8)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Totais", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"Total de créditos: {format_currency(prestacao['total_creditos'])}", ln=True)
    if total_creditos_extras > 0:
        pdf.cell(0, 6, f"Creditos extras: {format_currency(total_creditos_extras)}", ln=True)
    pdf.cell(0, 6, f"Total de débitos: {format_currency(prestacao['total_debitos'])}", ln=True)
    pdf.cell(0, 6, f"Despesas extras: {format_currency(prestacao['total_despesas'])}", ln=True)
    pdf.cell(0, 6, f"Saldo final: {format_currency(prestacao['saldo_final'])}", ln=True)

    if prestacao.get("observacoes"):
        pdf.ln(6)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, "Observações", ln=True)
        pdf.set_font("Arial", "", 10)
        for linha in prestacao["observacoes"].splitlines():
            pdf.multi_cell(0, 5, linha)

    pdf_bytes = pdf.output(dest="S").encode("latin-1")
    buffer = io.BytesIO(pdf_bytes)
    buffer.seek(0)
    filename = f"prestacao_contrato_{prestacao_id}.pdf"
    return send_file(buffer, mimetype="application/pdf", download_name=filename, as_attachment=True)


@app.route("/ordens-pagamento/print/<int:id>")
@login_required
@permission_required("Financeiro", "Consultar")
def ordens_pagamento_print(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(
        """
        SELECT op.*, p.razao_social_nome AS fornecedor_nome, p.documento AS fornecedor_documento,
               i.endereco AS imovel_endereco, d.descricao AS despesa_descricao,
               o.descricao AS origem_descricao
          FROM ordens_pagamento op
          JOIN pessoas p ON p.id = op.fornecedor_id
          LEFT JOIN imoveis i ON i.id = op.imovel_id
          LEFT JOIN despesas_cadastro d ON d.id = op.despesa_id
          LEFT JOIN origens_cadastro o ON o.id = op.origem_id
         WHERE op.id = %s
        """,
        (id,),
    )
    ordem = cur.fetchone()
    if not ordem:
        cur.close()
        conn.close()
        flash("Ordem de Pagamento não encontrada.", "danger")
        return redirect(url_for("ordens_pagamento_list"))
    cur.execute(
        "SELECT * FROM ordem_pagamento_itens WHERE ordem_id = %s ORDER BY id",
        (id,),
    )
    itens = cur.fetchall()
    cur.execute(
        "SELECT * FROM ordem_pagamento_parcelas WHERE ordem_id = %s ORDER BY COALESCE(numero, 999999), data_vencimento, id",
        (id,),
    )
    parcelas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("financeiro/ordens_pagamento/print.html", ordem=ordem, itens=itens, parcelas=parcelas)


# --- Módulo Caixa e Banco ---

@app.route("/ordens-pagamento/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Excluir")
def ordens_pagamento_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Itens são removidos por ON DELETE CASCADE
        cur.execute("DELETE FROM ordens_pagamento WHERE id = %s", (id,))
        conn.commit()
        flash("Ordem de Pagamento excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir a Ordem de Pagamento: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("ordens_pagamento_list"))

@app.route("/caixas")
@login_required
@permission_required("Financeiro", "Consultar")
def caixas_list():
    contas = ContaCaixa.query.all()
    return render_template("financeiro/caixas/list.html", contas=contas)


@app.route("/caixas/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def caixas_add():
    if request.method == "POST":
        nome = request.form["nome"]
        moeda = request.form.get("moeda", "BRL")
        saldo_inicial = request.form.get("saldo_inicial") or 0
        data_saldo = request.form.get("data_saldo_inicial")
        data_saldo_inicial = (
            datetime.strptime(data_saldo, "%Y-%m-%d").date() if data_saldo else None
        )
        conta = ContaCaixa(
            nome=nome,
            moeda=moeda,
            saldo_inicial=saldo_inicial,
            saldo_atual=saldo_inicial,
            data_saldo_inicial=data_saldo_inicial,
        )
        db.session.add(conta)
        db.session.commit()
        flash("Caixa cadastrado com sucesso!", "success")
        return redirect(url_for("caixas_list"))
    return render_template(
        "financeiro/caixas/add_list.html",
        conta=request.form,
        action_url=url_for("caixas_add"),
    )


@app.route("/caixas/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def caixas_edit(id):
    conta = ContaCaixa.query.get_or_404(id)
    if request.method == "POST":
        conta.nome = request.form["nome"]
        conta.moeda = request.form.get("moeda", "BRL")
        conta.saldo_inicial = request.form.get("saldo_inicial") or 0
        data_saldo = request.form.get("data_saldo_inicial")
        conta.data_saldo_inicial = (
            datetime.strptime(data_saldo, "%Y-%m-%d").date() if data_saldo else None
        )
        db.session.commit()
        flash("Conta de caixa atualizada com sucesso!", "success")
        return redirect(url_for("caixas_list"))
    return render_template(
        "financeiro/caixas/add_list.html",
        conta=conta,
        action_url=url_for("caixas_edit", id=id),
    )


@app.route("/bancos")
@login_required
@permission_required("Financeiro", "Consultar")
def bancos_list():
    contas = ContaBanco.query.all()
    return render_template("financeiro/bancos/list.html", contas=contas)


@app.route("/bancos/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def bancos_add():
    if request.method == "POST":
        conta = ContaBanco(
            banco=request.form["banco"],
            nome_banco=request.form.get("nome_banco"),
            agencia=request.form["agencia"],
            conta=request.form["conta"],
            tipo=request.form.get("tipo"),
            convenio=request.form.get("convenio"),
            carteira=request.form.get("carteira"),
            variacao=request.form.get("variacao"),
            contrato=request.form.get("contrato"),
            juros_mora=request.form.get("juros_mora"),
            multa=request.form.get("multa"),
            dias_protesto=request.form.get("dias_protesto"),
            especie_documento=request.form.get("especie_documento"),
            saldo_inicial=request.form.get("saldo_inicial") or 0,
            saldo_atual=request.form.get("saldo_inicial") or 0,
            data_saldo_inicial=(
                datetime.strptime(
                    request.form.get("data_saldo_inicial"), "%Y-%m-%d"
                ).date()
                if request.form.get("data_saldo_inicial")
                else None
            ),
        )
        db.session.add(conta)
        db.session.commit()
        flash("Conta bancária cadastrada com sucesso!", "success")
        return redirect(url_for("bancos_list"))
    return render_template(
        "financeiro/bancos/add_list.html",
        conta=request.form,
        action_url=url_for("bancos_add"),
    )


@app.route("/bancos/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def bancos_edit(id):
    conta = ContaBanco.query.get_or_404(id)
    if request.method == "POST":
        conta.banco = request.form["banco"]
        conta.nome_banco = request.form.get("nome_banco")
        conta.agencia = request.form["agencia"]
        conta.conta = request.form["conta"]
        conta.tipo = request.form.get("tipo")
        conta.convenio = request.form.get("convenio")
        conta.carteira = request.form.get("carteira")
        conta.variacao = request.form.get("variacao")
        conta.contrato = request.form.get("contrato")
        conta.juros_mora = request.form.get("juros_mora")
        conta.multa = request.form.get("multa")
        conta.dias_protesto = request.form.get("dias_protesto")
        conta.especie_documento = request.form.get("especie_documento")
        conta.saldo_inicial = request.form.get("saldo_inicial") or 0
        data_saldo = request.form.get("data_saldo_inicial")
        conta.data_saldo_inicial = (
            datetime.strptime(data_saldo, "%Y-%m-%d").date() if data_saldo else None
        )
        db.session.commit()
        flash("Conta bancária atualizada com sucesso!", "success")
        return redirect(url_for("bancos_list"))
    return render_template(
        "financeiro/bancos/add_list.html",
        conta=conta,
        action_url=url_for("bancos_edit", id=id),
    )


@app.route("/movimento/<tipo>/<int:conta_id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def movimento_novo(tipo, conta_id):
    conta = ContaCaixa.query.get(conta_id) if tipo == "caixa" else ContaBanco.query.get(conta_id)
    if not conta:
        flash("Conta não encontrada.", "danger")
        return redirect(url_for("caixas_list" if tipo == "caixa" else "bancos_list"))
    if request.method == "POST":
        valor_pago = parse_decimal(request.form.get("valor_pago")) or Decimal("0")
        valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
        valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
        valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
        valor_total = valor_pago + valor_juros + valor_multa - valor_desconto
        data = {
            "conta_origem_id": conta_id,
            "conta_origem_tipo": tipo,
            "tipo": request.form["tipo"],
            "valor": valor_total,
            "valor_pago": valor_pago,
            "valor_desconto": valor_desconto,
            "valor_multa": valor_multa,
            "valor_juros": valor_juros,
            "categoria": request.form.get("categoria"),
            "historico": request.form.get("historico"),
            "data_movimento": request.form.get("data_movimento") or datetime.today().date(),
        }
        if data["tipo"] == "saida":
            data["despesa_id"] = request.form.get("despesa_id") or None
        elif data["tipo"] == "entrada":
            data["receita_id"] = request.form.get("receita_id") or None
        criar_movimento(data)
        flash("Movimento registrado com sucesso!", "success")
        return redirect(url_for("caixas_list" if tipo == "caixa" else "bancos_list"))
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/caixas/lancamento.html",
        conta=conta,
        tipo=tipo,
        despesas=despesas,
        receitas=receitas,
        date_today=datetime.today().date().isoformat(),
    )

@app.route("/lancamentos/novo", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def lancamentos_novo():
    contas_caixa = ContaCaixa.query.all()
    contas_banco = ContaBanco.query.all()
    if request.method == "POST":
        conta_tipo = request.form["conta_tipo"]
        conta_id = int(request.form["conta_id"])
        valor_pago = parse_decimal(request.form.get("valor_pago")) or Decimal("0")
        valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
        valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
        valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
        valor_total = valor_pago + valor_juros + valor_multa - valor_desconto
        data = {
            "conta_origem_id": conta_id,
            "conta_origem_tipo": conta_tipo,
            "tipo": request.form["tipo"],
            "valor": valor_total,
            "valor_pago": valor_pago,
            "valor_desconto": valor_desconto,
            "valor_multa": valor_multa,
            "valor_juros": valor_juros,
            "categoria": request.form.get("categoria"),
            "historico": request.form.get("historico"),
            "data_movimento": request.form.get("data_movimento") or datetime.today().date(),
        }
        if data["tipo"] == "saida":
            data["despesa_id"] = request.form.get("despesa_id") or None
        elif data["tipo"] == "entrada":
            data["receita_id"] = request.form.get("receita_id") or None
        criar_movimento(data)
        flash("Movimento registrado com sucesso!", "success")
        return redirect(url_for("lancamentos_novo"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/lancamentos/add_edit.html",
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
        despesas=despesas,
        receitas=receitas,
        date_today=datetime.today().date().isoformat(),
        lancamento=None,
    )


@app.route("/lancamentos/view/<int:id>")
@login_required
@permission_required("Financeiro", "Consultar")
def lancamentos_view(id):
    lancamento = MovimentoFinanceiro.query.get(id)
    if not lancamento:
        flash("Lançamento não encontrado.", "danger")
        return redirect(url_for("lancamentos_list"))
    contas_caixa = {c.id: c.nome for c in ContaCaixa.query.all()}
    contas_banco = {b.id: f"{b.nome_banco} {b.conta}" for b in ContaBanco.query.all()}
    return render_template(
        "financeiro/lancamentos/view.html",
        lancamento=lancamento,
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
    )


@app.route("/lancamentos/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def lancamentos_edit(id):
    lancamento = MovimentoFinanceiro.query.get(id)
    if not lancamento:
        flash("Lançamento não encontrado.", "danger")
        return redirect(url_for("lancamentos_list"))
    contas_caixa = ContaCaixa.query.all()
    contas_banco = ContaBanco.query.all()
    if request.method == "POST":
        valor_pago = parse_decimal(request.form.get("valor_pago")) or Decimal("0")
        valor_desconto = parse_decimal(request.form.get("valor_desconto")) or Decimal("0")
        valor_multa = parse_decimal(request.form.get("valor_multa")) or Decimal("0")
        valor_juros = parse_decimal(request.form.get("valor_juros")) or Decimal("0")
        valor_total = valor_pago + valor_juros + valor_multa - valor_desconto
        data = {
            "conta_origem_id": int(request.form["conta_id"]),
            "conta_origem_tipo": request.form["conta_tipo"],
            "tipo": request.form["tipo"],
            "valor": valor_total,
            "valor_pago": valor_pago,
            "valor_desconto": valor_desconto,
            "valor_multa": valor_multa,
            "valor_juros": valor_juros,
            "categoria": request.form.get("categoria"),
            "historico": request.form.get("historico"),
            "data_movimento": request.form.get("data_movimento") or datetime.today().date(),
            "despesa_id": request.form.get("despesa_id") or None,
            "receita_id": request.form.get("receita_id") or None,
        }
        atualizar_movimento(lancamento, data)
        flash("Lançamento atualizado com sucesso!", "success")
        return redirect(url_for("lancamentos_list"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
    despesas = cur.fetchall()
    cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
    receitas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "financeiro/lancamentos/add_edit.html",
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
        despesas=despesas,
        receitas=receitas,
        lancamento=lancamento,
        date_today=lancamento.data_movimento.isoformat(),
    )


@app.route("/lancamentos/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Excluir")
def lancamentos_delete(id):
    lancamento = MovimentoFinanceiro.query.get(id)
    if not lancamento:
        flash("Lançamento não encontrado.", "danger")
    else:
        deletar_movimento(lancamento)
        flash("Lançamento excluído com sucesso!", "success")
    return redirect(url_for("lancamentos_list"))


@app.route("/lancamentos")
@login_required
@permission_required("Financeiro", "Consultar")
def lancamentos_list():
    movimentos = (
        MovimentoFinanceiro.query
        .order_by(MovimentoFinanceiro.data_movimento.desc())
        .all()
    )
    contas_caixa = {c.id: c.nome for c in ContaCaixa.query.all()}
    contas_banco = {
        b.id: f"{b.nome_banco} {b.conta}" for b in ContaBanco.query.all()
    }
    return render_template(
        "financeiro/lancamentos/list.html",
        movimentos=movimentos,
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
    )


@app.route("/bancos/importar/<int:conta_id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def banco_importar_cnab(conta_id):
    conta = ContaBanco.query.get(conta_id)
    if not conta:
        flash("Conta bancária não encontrada.", "danger")
        return redirect(url_for("bancos_list"))
    arquivo = request.files.get("arquivo")
    if not arquivo:
        flash("Selecione um arquivo CNAB.", "danger")
    else:
        resultados = importar_cnab(arquivo, conta_id, "banco")
        flash(f"{len(resultados)} lançamentos importados.", "success")
    return redirect(url_for("bancos_list"))


@app.route("/cobrancas", methods=["GET"])
@login_required
@permission_required("Financeiro", "Consultar")
def cobrancas_list():
    cobrancas = (
        db.session.query(Cobranca, ContaReceber, Pessoa)
        .join(ContaReceber, Cobranca.conta_id == ContaReceber.id)
        .join(Pessoa, ContaReceber.cliente_id == Pessoa.id)
        .order_by(Cobranca.data_cobranca.desc())
        .all()
    )
    return render_template("financeiro/cobrancas/list.html", cobrancas=cobrancas)


@app.route("/cobrancas/titulos", methods=["GET"])
@login_required
@permission_required("Financeiro", "Consultar")
def cobrancas_titulos():
    hoje = date.today()
    titulos = (
        db.session.query(ContaReceber, Pessoa)
        .join(Pessoa, ContaReceber.cliente_id == Pessoa.id)
        .filter(
            ContaReceber.status_conta.in_(["Aberta", "Parcial", "Vencida"]),
            ContaReceber.data_vencimento < hoje,
        )
        .order_by(ContaReceber.data_vencimento)
        .all()
    )
    return render_template("financeiro/cobrancas/titulos.html", titulos=titulos)


@app.route("/cobrancas/add", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Incluir")
def cobrancas_add():
    conta_id = request.args.get("conta_id") or request.form.get("conta_id")
    conta = ContaReceber.query.get(conta_id) if conta_id else None
    cliente = Pessoa.query.get(conta.cliente_id) if conta else None
    if request.method == "POST":
        cobrador = request.form.get("cobrador")
        contato = request.form.get("contato")
        data_cobranca = request.form.get("data_cobranca")
        historico = request.form.get("historico")
        data_prevista_pagamento = request.form.get("data_prevista_pagamento")
        if conta and data_cobranca:
            cobranca = Cobranca(
                conta_id=conta.id,
                cobrador=cobrador,
                contato=contato,
                data_cobranca=datetime.strptime(data_cobranca, "%Y-%m-%d").date(),
                historico=historico,
                data_prevista_pagamento=
                    datetime.strptime(data_prevista_pagamento, "%Y-%m-%d").date()
                    if data_prevista_pagamento
                    else None,
            )
            db.session.add(cobranca)
            db.session.commit()
            flash("Cobrança registrada com sucesso.", "success")
            return redirect(url_for("cobrancas_list"))
        else:
            flash("Selecione um título e informe a data da cobrança.", "danger")
    return render_template("financeiro/cobrancas/add.html", conta=conta, cliente=cliente)


@app.route("/cobrancas/view/<int:id>", methods=["GET"])
@login_required
@permission_required("Financeiro", "Consultar")
def cobrancas_view(id):
    cobranca = Cobranca.query.get(id)
    if not cobranca:
        flash("Cobrança não encontrada.", "danger")
        return redirect(url_for("cobrancas_list"))
    conta = ContaReceber.query.get(cobranca.conta_id)
    cliente = Pessoa.query.get(conta.cliente_id) if conta else None
    return render_template(
        "financeiro/cobrancas/view.html",
        cobranca=cobranca,
        conta=conta,
        cliente=cliente,
    )


@app.route("/cobrancas/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Financeiro", "Editar")
def cobrancas_edit(id):
    cobranca = Cobranca.query.get(id)
    if not cobranca:
        flash("Cobrança não encontrada.", "danger")
        return redirect(url_for("cobrancas_list"))
    conta = ContaReceber.query.get(cobranca.conta_id)
    cliente = Pessoa.query.get(conta.cliente_id) if conta else None
    if request.method == "POST":
        cobranca.cobrador = request.form.get("cobrador")
        cobranca.contato = request.form.get("contato")
        data_cobranca = request.form.get("data_cobranca")
        if data_cobranca:
            cobranca.data_cobranca = datetime.strptime(data_cobranca, "%Y-%m-%d").date()
        cobranca.historico = request.form.get("historico")
        data_prevista_pagamento = request.form.get("data_prevista_pagamento")
        cobranca.data_prevista_pagamento = (
            datetime.strptime(data_prevista_pagamento, "%Y-%m-%d").date()
            if data_prevista_pagamento
            else None
        )
        db.session.commit()
        flash("Cobrança atualizada com sucesso.", "success")
        return redirect(url_for("cobrancas_list"))
    return render_template(
        "financeiro/cobrancas/add.html",
        conta=conta,
        cliente=cliente,
        cobranca=cobranca,
    )


@app.route("/cobrancas/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Financeiro", "Excluir")
def cobrancas_delete(id):
    cobranca = Cobranca.query.get(id)
    if not cobranca:
        flash("Cobrança não encontrada.", "danger")
    else:
        db.session.delete(cobranca)
        db.session.commit()
        flash("Cobrança excluída com sucesso.", "success")
    return redirect(url_for("cobrancas_list"))


@app.route("/posicoes", methods=["GET"])
@login_required
@permission_required("Financeiro", "Consultar")
def posicoes_list():
    posicoes = PosicaoDiaria.query.order_by(PosicaoDiaria.data.desc()).all()
    contas_caixa = {c.id: c.nome for c in ContaCaixa.query.all()}
    contas_banco = {
        b.id: f"{b.nome_banco} {b.conta}" for b in ContaBanco.query.all()
    }
    return render_template(
        "financeiro/posicoes/list.html",
        posicoes=posicoes,
        contas_caixa=contas_caixa,
        contas_banco=contas_banco,
    )


@app.route("/posicoes/recalcular", methods=["POST"])
@login_required
@permission_required("Financeiro", "Editar")
def posicoes_recalcular():
    inicio_str = request.form.get("inicio")
    inicio = (
        datetime.strptime(inicio_str, "%Y-%m-%d").date() if inicio_str else None
    )
    quantidade = recalcular_posicoes(inicio)
    flash(f"{quantidade} posições recalculadas.", "success")
    return redirect(url_for("posicoes_list"))


# --- Módulo de Relatórios ---

@app.route("/relatorios/contas-a-pagar")
@login_required
def relatorios_contas_a_pagar():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    cur.execute(
        "SELECT id, razao_social_nome FROM pessoas WHERE tipo IN ('Fornecedor', 'Cliente/Fornecedor') ORDER BY razao_social_nome"
    )
    fornecedores = cur.fetchall()
    cur.execute(
        "SELECT id, tipo_imovel, endereco, cidade, estado FROM imoveis ORDER BY endereco"
    )
    imoveis = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "relatorios/contas_a_pagar/index.html", fornecedores=fornecedores, imoveis=imoveis
    )


@app.route("/relatorios/contas-a-receber")
@login_required
def relatorios_contas_a_receber():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, razao_social_nome
              FROM pessoas
             WHERE tipo IN ('Cliente', 'Cliente/Fornecedor')
             ORDER BY razao_social_nome
            """
        )
        clientes = cur.fetchall()

        cur.execute(
            """
            SELECT id, tipo_imovel, endereco, cidade, estado
              FROM imoveis
             ORDER BY endereco
            """
        )
        imoveis = cur.fetchall()

        cur.execute(
            """
            SELECT DISTINCT cliente_id, imovel_id
              FROM contratos_aluguel
             WHERE cliente_id IS NOT NULL AND imovel_id IS NOT NULL
            """
        )
        vinculos = cur.fetchall()

        cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
        receitas = cur.fetchall()
    finally:
        cur.close()
        conn.close()
        
    vinculos_json = json.dumps(
        [
            {"cliente_id": v["cliente_id"], "imovel_id": v["imovel_id"]}
            for v in vinculos
            if v["cliente_id"] and v["imovel_id"]
        ]
    )
    status_opcoes = ["Aberta", "Parcial", "Paga", "Vencida", "Cancelada"]

    return render_template(
        "relatorios/contas_a_receber/index.html",
        receitas=receitas,
        clientes=clientes,
        imoveis=imoveis,
        vinculos_json=vinculos_json,
        status_opcoes=status_opcoes,
    )


def consultar_relatorio_contas_receber(
    cur,
    data_inicio,
    data_fim,
    cliente_id=None,
    imovel_id=None,
    status_contas=None,
):
    query = [
        """
        SELECT cr.data_vencimento,
               cr.data_pagamento,
               r.descricao AS receita,
               cr.status_conta,
               COALESCE(cr.valor_previsto, 0) AS valor_previsto,
               COALESCE(cr.valor_multa, 0) AS valor_multa,
               COALESCE(cr.valor_juros, 0) AS valor_juros,
               COALESCE(cr.valor_desconto, 0) AS valor_desconto,
               COALESCE(cr.valor_previsto, 0)
             + COALESCE(cr.valor_multa, 0)
             + COALESCE(cr.valor_juros, 0)
             - COALESCE(cr.valor_desconto, 0) AS total,
               cr.cliente_id,
               ca.imovel_id
          FROM contas_a_receber cr
          LEFT JOIN receitas_cadastro r ON cr.receita_id = r.id
          LEFT JOIN contratos_aluguel ca ON cr.contrato_id = ca.id
         WHERE cr.data_vencimento BETWEEN %s AND %s
        """
    ]
    params = [data_inicio, data_fim]

    if cliente_id:
        query.append("AND cr.cliente_id = %s")
        params.append(cliente_id)

    if imovel_id:
        query.append("AND ca.imovel_id = %s")
        params.append(imovel_id)

    if status_contas:
        marcadores = ", ".join(["%s"] * len(status_contas))
        query.append(f"AND cr.status_conta IN ({marcadores})")
        params.extend(status_contas)

    query.append("ORDER BY cr.data_vencimento, receita")
    cur.execute("\n".join(query), params)
    return cur.fetchall()


def calcular_totais_contas_receber(dados):
    total_previsto = Decimal("0")
    total_multa = Decimal("0")
    total_juros = Decimal("0")
    total_desconto = Decimal("0")
    total_geral = Decimal("0")

    for linha in dados:
        valor_previsto = linha.get("valor_previsto") or Decimal("0")
        valor_multa = linha.get("valor_multa") or Decimal("0")
        valor_juros = linha.get("valor_juros") or Decimal("0")
        valor_desconto = linha.get("valor_desconto") or Decimal("0")
        total = linha.get("total") or (
            valor_previsto + valor_multa + valor_juros - valor_desconto
        )

        total_previsto += Decimal(valor_previsto)
        total_multa += Decimal(valor_multa)
        total_juros += Decimal(valor_juros)
        total_desconto += Decimal(valor_desconto)
        total_geral += Decimal(total)

    return {
        "valor_previsto": total_previsto,
        "multa": total_multa,
        "juros": total_juros,
        "desconto": total_desconto,
        "total": total_geral,
    }


def obter_descricoes_filtros_contas_receber(cur, cliente_id=None, imovel_id=None):
    cliente_nome = None
    imovel_descricao = None

    if cliente_id:
        cur.execute(
            "SELECT razao_social_nome FROM pessoas WHERE id = %s",
            (cliente_id,),
        )
        row = cur.fetchone()
        if row:
            cliente_nome = row["razao_social_nome"]

    if imovel_id:
        cur.execute(
            """
            SELECT tipo_imovel, endereco, cidade, estado
              FROM imoveis
             WHERE id = %s
            """,
            (imovel_id,),
        )
        row = cur.fetchone()
        if row:
            cidade = row["cidade"] or ""
            estado = row["estado"] or ""
            local = (
                f" {cidade}/{estado}".strip()
                if cidade or estado
                else ""
            )
            imovel_descricao = f"{row['tipo_imovel']} - {row['endereco']}{local}"

    return cliente_nome, imovel_descricao


def carregar_dados_relatorio_contas_receber(form):
    cliente_id = parse_int(form.get("cliente_id"))
    imovel_id = parse_int(form.get("imovel_id"))
    data_inicio = parse_date(form.get("vencimento_inicio"))
    data_fim = parse_date(form.get("vencimento_fim"))
    status_contas = []
    if hasattr(form, "getlist"):
        status_contas = [valor for valor in form.getlist("status_conta") if valor]

    if not data_inicio or not data_fim:
        return None, "Informe o intervalo de vencimento."

    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    try:
        dados = consultar_relatorio_contas_receber(
            cur,
            data_inicio,
            data_fim,
            cliente_id,
            imovel_id,
            status_contas,
        )
        totais = calcular_totais_contas_receber(dados)
        cliente_nome, imovel_descricao = obter_descricoes_filtros_contas_receber(
            cur, cliente_id, imovel_id
        )
    finally:
        cur.close()
        conn.close()

    resultado = {
        "dados": dados,
        "totais": totais,
        "cliente_id": cliente_id,
        "imovel_id": imovel_id,
        "cliente_nome": cliente_nome,
        "imovel_descricao": imovel_descricao,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "status_contas": status_contas,
    }

    return resultado, None


@app.route("/relatorios/contas-a-receber/detalhado", methods=["POST"])
@login_required
def relatorio_contas_a_receber_detalhado():
    resultado, erro = carregar_dados_relatorio_contas_receber(request.form)
    if erro:
        flash(erro, "warning")
        return redirect(url_for("relatorios_contas_a_receber"))

    periodo_txt = (
        resultado["data_inicio"].strftime("%d/%m/%Y")
        + " a "
        + resultado["data_fim"].strftime("%d/%m/%Y")
    )

    filtros = {
        "cliente_id": resultado["cliente_id"],
        "imovel_id": resultado["imovel_id"],
        "cliente_nome": resultado["cliente_nome"],
        "imovel_descricao": resultado["imovel_descricao"],
        "inicio": resultado["data_inicio"],
        "fim": resultado["data_fim"],
        "inicio_str": resultado["data_inicio"].isoformat(),
        "fim_str": resultado["data_fim"].isoformat(),
        "periodo": periodo_txt,
        "status_contas": resultado["status_contas"],
    }

    return render_template(
        "relatorios/contas_a_receber/relatorio_detalhado.html",
        dados=resultado["dados"],
        totais=resultado["totais"],
        filtros=filtros,
    )


@app.route("/relatorios/contas-a-receber/detalhado/excel", methods=["POST"])
@login_required
def relatorio_contas_a_receber_detalhado_excel():
    resultado, erro = carregar_dados_relatorio_contas_receber(request.form)
    if erro:
        flash(erro, "warning")
        return redirect(url_for("relatorios_contas_a_receber"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Receber"

    headers = [
        "Vencimento",
        "Pagamento",
        "Receita",
        "Status",
        "Valor Previsto",
        "Multa",
        "Juros",
        "Desconto",
        "Total",
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    row_index = 2
    for linha in resultado["dados"]:
        data_vencimento = (
            linha["data_vencimento"].strftime("%d/%m/%Y")
            if linha["data_vencimento"]
            else ""
        )
        ws.cell(row=row_index, column=1, value=data_vencimento)
        data_pagamento = (
            linha["data_pagamento"].strftime("%d/%m/%Y")
            if linha["data_pagamento"]
            else ""
        )
        ws.cell(row=row_index, column=2, value=data_pagamento)
        ws.cell(row=row_index, column=3, value=linha.get("receita") or "")
        ws.cell(row=row_index, column=4, value=linha.get("status_conta") or "")
        ws.cell(
            row=row_index,
            column=5,
            value=float(linha.get("valor_previsto") or 0),
        )
        ws.cell(
            row=row_index,
            column=6,
            value=float(linha.get("valor_multa") or 0),
        )
        ws.cell(
            row=row_index,
            column=7,
            value=float(linha.get("valor_juros") or 0),
        )
        ws.cell(
            row=row_index,
            column=8,
            value=float(linha.get("valor_desconto") or 0),
        )
        ws.cell(row=row_index, column=9, value=float(linha.get("total") or 0))

        for col in range(5, 10):
            ws.cell(row=row_index, column=col).number_format = "#,##0.00"

        row_index += 1

    total_row = row_index
    ws.cell(row=total_row, column=1, value="Totais")
    for col in range(2, 5):
        ws.cell(row=total_row, column=col, value="")
    ws.cell(
        row=total_row,
        column=5,
        value=float(resultado["totais"]["valor_previsto"]),
    ).number_format = "#,##0.00"
    ws.cell(
        row=total_row,
        column=6,
        value=float(resultado["totais"]["multa"]),
    ).number_format = "#,##0.00"
    ws.cell(
        row=total_row,
        column=7,
        value=float(resultado["totais"]["juros"]),
    ).number_format = "#,##0.00"
    ws.cell(
        row=total_row,
        column=8,
        value=float(resultado["totais"]["desconto"]),
    ).number_format = "#,##0.00"
    ws.cell(
        row=total_row,
        column=9,
        value=float(resultado["totais"]["total"]),
    ).number_format = "#,##0.00"

    column_widths = [15, 15, 40, 18, 20, 15, 15, 15, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "contas_a_receber_{inicio}_{fim}.xlsx".format(
        inicio=resultado["data_inicio"].strftime("%Y%m%d"),
        fim=resultado["data_fim"].strftime("%Y%m%d"),
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/relatorios/contas-a-receber/detalhado/pdf", methods=["POST"])
@login_required
def relatorio_contas_a_receber_detalhado_pdf():
    resultado, erro = carregar_dados_relatorio_contas_receber(request.form)
    if erro:
        flash(erro, "warning")
        return redirect(url_for("relatorios_contas_a_receber"))

    def to_latin(text):
        if not text:
            return ""
        return str(text).encode("latin-1", "replace").decode("latin-1")

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            largura_util = self.w - self.l_margin - self.r_margin
            self.cell(largura_util, 8, "Contas a Receber - Detalhado", 0, 1, "C")
            self.set_font("Arial", "", 10)
            periodo = getattr(self, "periodo", "")
            if periodo:
                self.cell(largura_util, 6, f"Período: {periodo}", 0, 1, "C")
            filtros = getattr(self, "filtros", {})
            cliente = filtros.get("cliente")
            imovel = filtros.get("imovel")
            if cliente:
                self.cell(largura_util, 5, to_latin(f"Cliente: {cliente}"), 0, 1, "C")
            if imovel:
                self.cell(largura_util, 5, to_latin(f"Imóvel: {imovel}"), 0, 1, "C")
            self.ln(2)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    headers = [
        ("Vencimento", 28),
        ("Pagamento", 28),
        ("Receita", 65),
        ("Status", 28),
        ("Valor Previsto", 28),
        ("Multa", 22),
        ("Juros", 22),
        ("Desconto", 25),
        ("Total", 28),
    ]

    def truncate_text(pdf_obj, texto, largura):
        if not texto:
            return ""
        base = to_latin(texto)
        if pdf_obj.get_string_width(base) <= largura:
            return base
        corte = base
        while corte and pdf_obj.get_string_width(corte + "...") > largura:
            corte = corte[:-1]
        return corte + "..."

    pdf = PDF(orientation="L")
    pdf.alias_nb_pages()
    pdf.periodo = (
        resultado["data_inicio"].strftime("%d/%m/%Y")
        + " a "
        + resultado["data_fim"].strftime("%d/%m/%Y")
    )
    pdf.filtros = {
        "cliente": resultado["cliente_nome"],
        "imovel": resultado["imovel_descricao"],
    }
    pdf.add_page()

    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    for titulo, largura in headers:
        pdf.cell(largura, 8, titulo, 1, 0, "C", True)
    pdf.ln(8)

    pdf.set_font("Arial", "", 9)
    for linha in resultado["dados"]:
        data_venc = (
            linha["data_vencimento"].strftime("%d/%m/%Y")
            if linha["data_vencimento"]
            else ""
        )
        data_pag = (
            linha["data_pagamento"].strftime("%d/%m/%Y")
            if linha["data_pagamento"]
            else ""
        )
        valores = [
            data_venc,
            data_pag,
            truncate_text(pdf, linha.get("receita") or "", headers[2][1] - 2),
            truncate_text(pdf, linha.get("status_conta") or "", headers[3][1] - 2),
            format_currency(linha.get("valor_previsto") or 0),
            format_currency(linha.get("valor_multa") or 0),
            format_currency(linha.get("valor_juros") or 0),
            format_currency(linha.get("valor_desconto") or 0),
            format_currency(linha.get("total") or 0),
        ]
        for (titulo, largura), valor in zip(headers, valores):
            align = "R" if titulo not in ("Vencimento", "Pagamento", "Receita", "Status") else "L"
            pdf.cell(largura, 7, valor, 1, 0, align)
        pdf.ln(7)

    pdf.set_font("Arial", "B", 9)
    pdf.cell(
        headers[0][1] + headers[1][1] + headers[2][1] + headers[3][1],
        7,
        "Totais",
        1,
        0,
        "R",
    )
    pdf.cell(
        headers[4][1],
        7,
        format_currency(resultado["totais"]["valor_previsto"]),
        1,
        0,
        "R",
    )
    pdf.cell(
        headers[5][1],
        7,
        format_currency(resultado["totais"]["multa"]),
        1,
        0,
        "R",
    )
    pdf.cell(
        headers[6][1],
        7,
        format_currency(resultado["totais"]["juros"]),
        1,
        0,
        "R",
    )
    pdf.cell(
        headers[7][1],
        7,
        format_currency(resultado["totais"]["desconto"]),
        1,
        0,
        "R",
    )
    pdf.cell(
        headers[8][1],
        7,
        format_currency(resultado["totais"]["total"]),
        1,
        1,
        "R",
    )

    pdf_bytes = pdf.output(dest="S").encode("latin1")

    filename = "contas_a_receber_{inicio}_{fim}.pdf".format(
        inicio=resultado["data_inicio"].strftime("%Y%m%d"),
        fim=resultado["data_fim"].strftime("%Y%m%d"),
    )

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/relatorios/contas-a-receber/recebimento-inquilino", methods=["POST"])
@login_required
def relatorio_recebimento_por_inquilino():
    """Relatório de recebimentos por inquilino, filtrado por data de recebimento.

    Colunas: Data | Cliente | CPF | Imóvel | Receita | Valor Previsto | Multa | Juros | Desconto | Total Recebido | Histórico
    Orientação: Paisagem
    """
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    receitas_ids = [int(x) for x in request.form.getlist("receitas_ids") if str(x).strip()]

    if not (data_inicio and data_fim):
        flash("Informe o período.", "warning")
        return redirect(url_for("relatorios_contas_a_receber"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)

    # Empresa para cabeçalho
    cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
    empresa = cur.fetchone()

    query = (
        """
        SELECT cr.id,
               cr.data_pagamento,
               p.razao_social_nome AS cliente,
               p.documento AS cpf,
               i.tipo_imovel, i.endereco, i.cidade, i.estado,
               r.descricao AS receita,
               cr.valor_previsto,
               COALESCE(cr.valor_multa, 0) AS valor_multa,
               COALESCE(cr.valor_juros, 0) AS valor_juros,
               COALESCE(cr.valor_desconto, 0) AS valor_desconto,
               COALESCE(cr.valor_pago, 0) AS valor_pago,
               mf.historico AS historico,
               mf.valor AS valor_mov
          FROM contas_a_receber cr
          JOIN pessoas p ON cr.cliente_id = p.id
     LEFT JOIN contratos_aluguel ca ON ca.id = cr.contrato_id
     LEFT JOIN imoveis i ON i.id = ca.imovel_id
     LEFT JOIN receitas_cadastro r ON r.id = cr.receita_id
     LEFT JOIN LATERAL (
                SELECT historico, valor
                  FROM movimento_financeiro
                 WHERE documento = 'CR-' || cr.id::text
                   AND tipo = 'entrada'
                   AND data_movimento = cr.data_pagamento
                 ORDER BY id DESC
                 LIMIT 1
            ) mf ON TRUE
         WHERE cr.data_pagamento IS NOT NULL
           AND cr.data_pagamento BETWEEN %s AND %s
        """
    )
    params = [data_inicio, data_fim]
    if receitas_ids:
        placeholders = ",".join(["%s"] * len(receitas_ids))
        query += f" AND cr.receita_id IN ({placeholders})"
        params.extend(receitas_ids)
    query += " ORDER BY cr.data_pagamento ASC, cr.id ASC"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Monta linhas já com campos prontos e totais
    linhas = []
    for r in rows:
        imovel = ""
        if r.get("tipo_imovel") or r.get("endereco"):
            tipo = r.get("tipo_imovel") or ""
            ender = r.get("endereco") or ""
            cid = r.get("cidade") or ""
            uf = r.get("estado") or ""
            if cid and uf:
                imovel = f"{tipo} / {ender} / {cid}/{uf}".strip()
            else:
                imovel = f"{tipo} / {ender}".strip(" /")
        total_recebido = r.get("valor_mov")
        if total_recebido is None:
            # Fallback quando não houver movimento associado
            total_recebido = (r.get("valor_pago") or 0) + (r.get("valor_juros") or 0) + (r.get("valor_multa") or 0) - (r.get("valor_desconto") or 0)
        linhas.append(
            {
                "data": r["data_pagamento"],
                "cliente": r["cliente"],
                "cpf": r["cpf"],
                "imovel": imovel,
                "receita": r["receita"],
                "valor_previsto": r["valor_previsto"],
                "multa": r["valor_multa"],
                "juros": r["valor_juros"],
                "desconto": r["valor_desconto"],
                "total_recebido": total_recebido,
                "historico": r.get("historico") or "",
            }
        )

    periodo_txt = (
        datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
        + " a "
        + datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
    )

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width / 3, 8, "", 0, 0, "L")
            self.cell(page_width / 3, 8, "Recebimento Por Inquilino", 0, 0, "C")
            self.set_font("Arial", "", 10)
            self.cell(page_width / 3, 8, getattr(self, "gerado_em", ""), 0, 1, "R")
            emp = getattr(self, "empresa", "")
            if emp:
                self.cell(0, 6, emp, 0, 1, "C")
            per = getattr(self, "periodo", "")
            if per:
                self.cell(0, 6, f"Período: {per}", 0, 1, "C")
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    # Larguras pensadas para A4 paisagem (~277mm úteis)
    headers = [
        ("Data", 20),
        ("Cliente", 35),
        ("CPF", 28),
        ("Imóvel", 38),
        ("Receita", 25),
        ("Valor Previsto", 22),
        ("Multa", 18),
        ("Juros", 18),
        ("Desconto", 18),
        ("Total Recebido", 25),
        ("Histórico", 30),
    ]

    def truncate_text(pdf, text, max_width):
        if not text:
            return ""
        s = str(text)
        if pdf.get_string_width(s) <= max_width:
            return s
        while s and pdf.get_string_width(s + "...") > max_width:
            s = s[:-1]
        return s + "..."

    pdf = PDF(orientation="L")
    pdf.gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.empresa = empresa["razao_social_nome"] if empresa else ""
    pdf.periodo = periodo_txt
    pdf.alias_nb_pages()
    pdf.add_page()

    # Cabeçalho da tabela
    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    for text, width in headers:
        pdf.cell(width, 8, text, 1, 0, "C", True)
    pdf.ln(8)

    pdf.set_font("Arial", "", 9)
    for l in linhas:
        pdf.cell(headers[0][1], 7, l["data"].strftime("%d/%m/%Y"), 1)
        pdf.cell(headers[1][1], 7, truncate_text(pdf, l["cliente"], headers[1][1] - 2), 1)
        pdf.cell(headers[2][1], 7, str(l["cpf"] or ""), 1)
        pdf.cell(headers[3][1], 7, truncate_text(pdf, l["imovel"], headers[3][1] - 2), 1)
        pdf.cell(headers[4][1], 7, truncate_text(pdf, l["receita"] or "", headers[4][1] - 2), 1)
        pdf.cell(headers[5][1], 7, format_currency(l["valor_previsto"]), 1, 0, "R")
        pdf.cell(headers[6][1], 7, format_currency(l["multa"]), 1, 0, "R")
        pdf.cell(headers[7][1], 7, format_currency(l["juros"]), 1, 0, "R")
        pdf.cell(headers[8][1], 7, format_currency(l["desconto"]), 1, 0, "R")
        pdf.cell(headers[9][1], 7, format_currency(l["total_recebido"]), 1, 0, "R")
        pdf.cell(headers[10][1], 7, truncate_text(pdf, l["historico"], headers[10][1] - 2), 1)
        pdf.ln(7)

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="recebimento_por_inquilino.pdf",
    )

@app.route("/relatorios/contas-a-pagar/por-periodo", methods=["POST"])
@login_required
def relatorio_contas_a_pagar_periodo():
    fornecedor_id = request.form.get("fornecedor_id")
    imovel_id = parse_int(request.form.get("imovel_id"))
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    status = request.form.get("status")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    atualizar_status_contas_a_pagar(cur)

    query = (
        "SELECT cp.titulo, p.razao_social_nome AS fornecedor, d.descricao AS despesa, "
        "cp.data_vencimento, cp.valor_previsto "
        "FROM contas_a_pagar cp "
        "JOIN pessoas p ON cp.fornecedor_id = p.id "
        "LEFT JOIN despesas_cadastro d ON cp.despesa_id = d.id "
        "WHERE cp.data_vencimento BETWEEN %s AND %s"
    )
    params = [data_inicio, data_fim]
    if fornecedor_id:
        query += " AND cp.fornecedor_id = %s"
        params.append(fornecedor_id)
    if imovel_id:
        query += " AND cp.imovel_id = %s"
        params.append(imovel_id)
    if status:
        query += " AND cp.status_conta = %s"
        params.append(status)
    query += " ORDER BY cp.data_vencimento ASC"
    cur.execute(query, params)
    contas = cur.fetchall()
    total_valor_previsto = sum(c["valor_previsto"] for c in contas)

    cur.execute(
        "SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1"
    )
    empresa = cur.fetchone()
    imovel = None
    if imovel_id:
        cur.execute(
            "SELECT tipo_imovel, endereco, cidade, estado FROM imoveis WHERE id = %s",
            (imovel_id,),
        )
        imovel = cur.fetchone()
    cur.close()
    conn.close()

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width / 3, 10, "", 0, 0, "L")
            self.cell(page_width / 3, 10, "Contas a Pagar", 0, 0, "C")
            self.set_font("Arial", "", 10)
            self.cell(page_width / 3, 10, getattr(self, "gerado_em", ""), 0, 0, "R")
            self.ln(5)
            self.cell(0, 10, getattr(self, "empresa", ""), 0, 1, "C")
            self.ln(5)
            imovel_info = getattr(self, "imovel_info", "")
            periodo = getattr(self, "periodo", "")
            if imovel_info:
                self.cell(0, 10, imovel_info, 0, 1, "L")
                periodo = getattr(self, "periodo", "")
            if periodo:
                self.cell(0, 10, periodo, 0, 1, "L")
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    pdf = PDF()
    pdf.empresa = empresa["razao_social_nome"] if empresa else ""
    pdf.gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    data_inicio_fmt = datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
    data_fim_fmt = datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
    pdf.periodo = f"Período de {data_inicio_fmt} até {data_fim_fmt}"
    if imovel:
        pdf.imovel_info = (
            f"Imóvel {imovel['tipo_imovel']} / {imovel['endereco']} / {imovel['cidade']}/{imovel['estado']}"
        )
    else:
        pdf.imovel_info = ""
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    headers = [
        ("Título\n ", 36),
        ("Fornecedor\n ", 68),
        ("Despesa\n ", 24),
        ("Data\nVencimento", 30),
        ("Valor\nPrevisto", 30),
    ]
    line_height = 8
    row_height = line_height * 2
    y_start = pdf.get_y()
    for text, width in headers:
        x_start = pdf.get_x()
        pdf.multi_cell(width, line_height, text, border=1, align="L", fill=True)
        pdf.set_xy(x_start + width, y_start)
    pdf.ln(row_height)

    def truncate_text(text, max_width):
        """Limita o texto para caber na largura especificada."""
        if not text:
            return ""
        text = str(text)
        if pdf.get_string_width(text) <= max_width:
            return text
        while pdf.get_string_width(text + "...") > max_width and text:
            text = text[:-1]
        return text + "..."

    pdf.set_font("Arial", "", 10)
    for c in contas:
        pdf.cell(36, 8, truncate_text(c["titulo"] or "", 34), 1)
        pdf.cell(68, 8, truncate_text(c["fornecedor"], 66), 1)
        pdf.cell(24, 8, truncate_text(c["despesa"] or "", 22), 1)
        pdf.cell(30, 8, c["data_vencimento"].strftime("%d/%m/%Y"), 1)
        pdf.cell(30, 8, format_currency(c["valor_previsto"]), 1, 1, "R")

    pdf.set_font("Arial", "B", 10)
    pdf.cell(158, 8, "Total", 1, 0, "R")
    pdf.cell(30, 8, format_currency(total_valor_previsto), 1, 1, "R")

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="contas_a_pagar_periodo.pdf",
    )


@app.route("/relatorios/financeiro")
@login_required
def relatorios_financeiro():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    cur.execute("SELECT id, nome FROM conta_caixa ORDER BY nome")
    caixas = cur.fetchall()
    cur.execute(
        "SELECT id, nome_banco, agencia, conta FROM conta_banco ORDER BY nome_banco"
    )
    bancos = cur.fetchall()
    cur.execute(
        "SELECT id, tipo_imovel, endereco, cidade, estado FROM imoveis ORDER BY endereco"
    )
    imoveis = cur.fetchall()
    cur.execute(
        "SELECT id, descricao FROM despesas_cadastro ORDER BY descricao"
    )
    despesas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "relatorios/financeiro/index.html",
        caixas=caixas,
        bancos=bancos,
        imoveis=imoveis,
        despesas=despesas,
    )


@app.route("/relatorios/financeiro/caixa-banco", methods=["POST"])
@login_required
def relatorio_financeiro_caixa_banco():
    tipo_conta = request.form.get("tipo_conta")
    conta_id = request.form.get("conta_id")
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)

    if tipo_conta == "caixa":
        cur.execute("SELECT nome FROM conta_caixa WHERE id = %s", (conta_id,))
        conta = cur.fetchone()
        conta_nome = conta["nome"] if conta else ""
    else:
        cur.execute(
            "SELECT nome_banco, agencia, conta FROM conta_banco WHERE id = %s",
            (conta_id,),
        )
        conta = cur.fetchone()
        if conta:
            conta_nome = (
                f"{conta['nome_banco']} Ag {conta['agencia']} Conta {conta['conta']}"
            )
        else:
            conta_nome = ""

    cur.execute(
        """
        SELECT COALESCE(SUM(CASE WHEN tipo = 'entrada' THEN valor ELSE -valor END),0) AS saldo
        FROM movimento_financeiro
        WHERE conta_origem_tipo = %s AND conta_origem_id = %s AND data_movimento < %s
        """,
        (tipo_conta, conta_id, data_inicio),
    )
    saldo_inicial = cur.fetchone()["saldo"]

    cur.execute(
        """
        SELECT id, data_movimento, historico, tipo, valor
        FROM movimento_financeiro
        WHERE conta_origem_tipo = %s AND conta_origem_id = %s
          AND data_movimento BETWEEN %s AND %s
        ORDER BY data_movimento ASC, id ASC
        """,
        (tipo_conta, conta_id, data_inicio, data_fim),
    )
    movimentos = cur.fetchall()
    cur.close()
    conn.close()

    entradas = Decimal("0")
    saidas = Decimal("0")
    saldo = Decimal(saldo_inicial)
    linhas = []
    for m in movimentos:
        entrada = m["valor"] if m["tipo"] == "entrada" else Decimal("0")
        saida = m["valor"] if m["tipo"] == "saida" else Decimal("0")
        entradas += entrada
        saidas += saida
        saldo += entrada - saida
        linhas.append(
            {
                "data": m["data_movimento"],
                "historico": m["historico"],
                "entrada": entrada,
                "saida": saida,
                "saldo": saldo,
            }
        )

    total_final = saldo_inicial + entradas - saidas

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            self.cell(0, 10, "Lançamentos Financeiros", 0, 0, "L")
            self.set_font("Arial", "", 10)
            self.cell(0, 10, getattr(self, "gerado_em", ""), 0, 1, "R")
            conta = getattr(self, "conta", "")
            if conta:
                self.cell(0, 10, conta, 0, 1, "L")
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    def nb_lines(pdf, w, txt):
        cw = pdf.current_font["cw"]
        if w == 0:
            w = pdf.w - pdf.r_margin - pdf.x
        wmax = (w - 2 * pdf.c_margin) * 1000 / pdf.font_size
        s = txt.replace("\r", "")
        nb = len(s)
        sep = -1
        i = 0
        j = 0
        l = 0
        nl = 1
        while i < nb:
            c = s[i]
            if c == "\n":
                i += 1
                sep = -1
                j = i
                l = 0
                nl += 1
                continue
            if c == " ":
                sep = i
            l += cw.get(c, 0)
            if l > wmax:
                if sep == -1:
                    if i == j:
                        i += 1
                else:
                    i = sep + 1
                sep = -1
                j = i
                l = 0
                nl += 1
            else:
                i += 1
        return nl

    pdf = PDF()
    pdf.gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.conta = conta_nome
    pdf.alias_nb_pages()
    pdf.add_page()

    headers = [
        ("Data", 25),
        ("Histórico", 75),
        ("Entrada (Valor)", 30),
        ("Saída (Valor)", 30),
        ("Saldo", 30),
    ]
    def draw_table_header():
        pdf.set_font("Arial", "B", 10)
        pdf.set_fill_color(200, 200, 200)
        for text, width in headers:
            pdf.cell(width, 8, text, 1, 0, "C", True)
        pdf.ln(8)
        pdf.set_font("Arial", "", 10)

    draw_table_header()

    line_height = 8
    for linha in linhas:
        historico_txt = str(linha["historico"] or "")
        lines = nb_lines(pdf, 75, historico_txt)
        cell_height = line_height * lines
        
        # Add new page and redraw header if the next row would overflow
        if pdf.get_y() + cell_height > pdf.page_break_trigger:
            pdf.add_page()
            draw_table_header()

        x = pdf.get_x()
        y = pdf.get_y()

        pdf.cell(25, cell_height, linha["data"].strftime("%d/%m/%Y"), 1)
        pdf.set_xy(x + 25, y)
        pdf.multi_cell(75, line_height, historico_txt, 1)
        pdf.set_xy(x + 100, y)
        pdf.cell(30, cell_height, format_currency(linha["entrada"]), 1, 0, "R")
        pdf.cell(30, cell_height, format_currency(linha["saida"]), 1, 0, "R")
        pdf.cell(30, cell_height, format_currency(linha["saldo"]), 1, 0, "R")
        pdf.ln(cell_height)

    pdf.set_font("Arial", "B", 10)
    pdf.cell(100, 8, "Total", 1, 0, "R")
    pdf.cell(30, 8, format_currency(entradas), 1, 0, "R")
    pdf.cell(30, 8, format_currency(saidas), 1, 0, "R")
    pdf.cell(30, 8, format_currency(total_final), 1, 1, "R")

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="lancamentos_financeiros.pdf",
    )


@app.route("/relatorios/financeiro/despesas-imovel", methods=["POST"])
@login_required
def relatorio_financeiro_despesas_imovel():
    imovel_id = request.form.get("imovel_id")
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    despesa_id = request.form.get("despesa_id")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)
    query = (
        "SELECT cp.data_pagamento, cp.titulo, p.razao_social_nome AS fornecedor, "
        "d.descricao AS despesa, cp.valor_pago "
        "FROM contas_a_pagar cp "
        "JOIN pessoas p ON cp.fornecedor_id = p.id "
        "LEFT JOIN despesas_cadastro d ON cp.despesa_id = d.id "
        "WHERE cp.imovel_id = %s AND cp.data_pagamento BETWEEN %s AND %s "
        "AND cp.status_conta = 'Paga'"
    )
    params = [imovel_id, data_inicio, data_fim]
    if despesa_id:
        query += " AND cp.despesa_id = %s"
        params.append(despesa_id)
    query += " ORDER BY cp.data_pagamento ASC"
    cur.execute(query, params)
    contas = cur.fetchall()
    total_valor = sum(c["valor_pago"] for c in contas)

    cur.execute(
        "SELECT tipo_imovel, endereco, cidade, estado FROM imoveis WHERE id = %s",
        (imovel_id,),
    )
    imovel = cur.fetchone()
    cur.execute(
        "SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1"
    )
    empresa = cur.fetchone()
    cur.close()
    conn.close()

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width / 3, 10, "", 0, 0, "L")
            self.cell(page_width / 3, 10, "Despesas Por Imóvel", 0, 0, "C")
            self.set_font("Arial", "", 10)
            self.cell(page_width / 3, 10, getattr(self, "gerado_em", ""), 0, 1, "R")
            self.cell(0, 10, getattr(self, "empresa", ""), 0, 1, "C")
            imovel_info = getattr(self, "imovel_info", "")
            if imovel_info:
                self.cell(0, 10, f"Imóvel: {imovel_info}", 0, 1, "L")
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    pdf = PDF()
    pdf.gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.empresa = empresa["razao_social_nome"] if empresa else ""
    if imovel:
        pdf.imovel_info = f"{imovel['tipo_imovel']} / {imovel['endereco']} / {imovel['cidade']}/{imovel['estado']}"
    else:
        pdf.imovel_info = ""
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    headers = [
        ("Data Pagamento", 25),
        ("Título", 45),
        ("Fornecedor", 50),
        ("Despesas", 40),
        ("Valor Pago", 30),
    ]
    for text, width in headers:
        pdf.cell(width, 8, text, 1, 0, "C", True)
    pdf.ln(8)

    pdf.set_font("Arial", "", 10)

    def truncate_text(text, max_width):
        if not text:
            return ""
        text = str(text)
        if pdf.get_string_width(text) <= max_width:
            return text
        while pdf.get_string_width(text + "...") > max_width and text:
            text = text[:-1]
        return text + "..."

    for c in contas:
        pdf.cell(25, 8, c["data_pagamento"].strftime("%d/%m/%Y"), 1)
        pdf.cell(45, 8, truncate_text(c["titulo"], 43), 1)
        pdf.cell(50, 8, truncate_text(c["fornecedor"], 48), 1)
        pdf.cell(40, 8, truncate_text(c["despesa"] or "", 38), 1)
        pdf.cell(30, 8, format_currency(c["valor_pago"]), 1, 1, "R")

    pdf.set_font("Arial", "B", 10)
    pdf.cell(160, 8, "Total", 1, 0, "R")
    pdf.cell(30, 8, format_currency(total_valor), 1, 1, "R")

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="despesas_por_imovel.pdf",
    )


@app.route("/relatorios/financeiro/fluxo-caixa", methods=["POST"])
@login_required
def relatorio_financeiro_fluxo_caixa():
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    caixas_ids = [int(x) for x in request.form.getlist("caixas_ids") if str(x).strip()]
    bancos_ids = [int(x) for x in request.form.getlist("bancos_ids") if str(x).strip()]

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)

    # Monta filtros opcionais por contas
    filtros = []
    params_periodo = [data_inicio, data_fim]
    if caixas_ids:
        placeholders_cx = ",".join(["%s"] * len(caixas_ids))
        filtros.append(f"(conta_origem_tipo = 'caixa' AND conta_origem_id IN ({placeholders_cx}))")
        params_periodo.extend(caixas_ids)
    if bancos_ids:
        placeholders_bk = ",".join(["%s"] * len(bancos_ids))
        filtros.append(f"(conta_origem_tipo = 'banco' AND conta_origem_id IN ({placeholders_bk}))")
        params_periodo.extend(bancos_ids)
    where_extras = f" AND ({' OR '.join(filtros)})" if filtros else ""

    # Saldo inicial via Posições Diárias (dia anterior ao início)
    data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    dia_anterior = (data_inicio_dt - timedelta(days=1)).isoformat()
    pos_filtros = []
    params_pos = [dia_anterior]
    if caixas_ids:
        placeholders_cx2 = ",".join(["%s"] * len(caixas_ids))
        pos_filtros.append(f"(conta_tipo = 'caixa' AND conta_id IN ({placeholders_cx2}))")
        params_pos.extend(caixas_ids)
    if bancos_ids:
        placeholders_bk2 = ",".join(["%s"] * len(bancos_ids))
        pos_filtros.append(f"(conta_tipo = 'banco' AND conta_id IN ({placeholders_bk2}))")
        params_pos.extend(bancos_ids)
    where_pos = f" AND ({' OR '.join(pos_filtros)})" if pos_filtros else ""
    query_pos = (
        "SELECT COALESCE(SUM(saldo),0) AS saldo FROM ("
        "  SELECT DISTINCT ON (conta_tipo, conta_id) conta_tipo, conta_id, saldo"
        "    FROM posicao_diaria"
        "   WHERE data <= %s" + where_pos +
        "   ORDER BY conta_tipo, conta_id, data DESC"
        ") t"
    )
    cur.execute(query_pos, tuple(params_pos))
    saldo_inicial = Decimal(cur.fetchone()["saldo"]) if cur.rowcount is not None else Decimal("0")

    # Entradas e saídas por dia no período
    query_periodo = (
        "SELECT DATE(data_movimento) AS dia, "
        "COALESCE(SUM(CASE WHEN tipo = 'entrada' THEN valor ELSE 0 END),0) AS entradas, "
        "COALESCE(SUM(CASE WHEN tipo = 'saida' THEN valor ELSE 0 END),0)   AS saidas "
        "FROM movimento_financeiro "
        "WHERE data_movimento BETWEEN %s AND %s" + where_extras + " GROUP BY 1 ORDER BY 1"
    )
    cur.execute(query_periodo, tuple(params_periodo))
    rows = cur.fetchall()

    # Nome da empresa para cabeçalho
    cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
    empresa = cur.fetchone()

    cur.close()
    conn.close()

    # Monta linhas com saldo acumulado
    linhas = []
    saldo = saldo_inicial
    total_entradas = Decimal("0")
    total_saidas = Decimal("0")
    for r in rows:
        dia = r["dia"]
        ent = Decimal(r["entradas"]) or Decimal("0")
        sai = Decimal(r["saidas"]) or Decimal("0")
        saldo = saldo + ent - sai
        total_entradas += ent
        total_saidas += sai
        linhas.append({"dia": dia, "entradas": ent, "saidas": sai, "saldo": saldo})

    total_final = saldo

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width / 3, 10, "", 0, 0, "L")
            self.cell(page_width / 3, 10, "Fluxo de Caixa", 0, 0, "C")
            self.set_font("Arial", "", 10)
            self.cell(page_width / 3, 10, getattr(self, "gerado_em", ""), 0, 1, "R")
            if getattr(self, "empresa", ""):
                self.cell(0, 10, self.empresa, 0, 1, "C")
            if getattr(self, "periodo", ""):
                self.cell(0, 10, self.periodo, 0, 1, "C")
            if getattr(self, "saldo_inicial_txt", ""):
                self.cell(0, 10, self.saldo_inicial_txt, 0, 1, "L")
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "", 10)
            self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.empresa = empresa["razao_social_nome"] if empresa else ""
    data_inicio_fmt = datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
    data_fim_fmt = datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
    pdf.periodo = f"Período: {data_inicio_fmt} a {data_fim_fmt}"
    pdf.saldo_inicial_txt = f"Saldo inicial do período: {format_currency(saldo_inicial)}"
    pdf.add_page()

    # Cabeçalhos da tabela
    pdf.set_font("Arial", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    headers = [("Data", 30), ("Entradas", 40), ("Saídas", 40), ("Saldo", 40)]
    for text, width in headers:
        pdf.cell(width, 8, text, 1, 0, "C", True)
    pdf.ln(8)

    pdf.set_font("Arial", "", 10)
    for linha in linhas:
        pdf.cell(30, 8, linha["dia"].strftime("%d/%m/%Y"), 1)
        pdf.cell(40, 8, format_currency(linha["entradas"]), 1, 0, "R")
        pdf.cell(40, 8, format_currency(linha["saidas"]), 1, 0, "R")
        pdf.cell(40, 8, format_currency(linha["saldo"]), 1, 1, "R")

    # Totais
    pdf.set_font("Arial", "B", 10)
    pdf.cell(30, 8, "Total", 1)
    pdf.cell(40, 8, format_currency(total_entradas), 1, 0, "R")
    pdf.cell(40, 8, format_currency(total_saidas), 1, 0, "R")
    pdf.cell(40, 8, format_currency(total_final), 1, 1, "R")

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="fluxo_de_caixa.pdf",
    )


@app.route("/relatorios/financeiro/fluxo-caixa/visualizar", methods=["POST"])
@login_required
def relatorio_financeiro_fluxo_caixa_visualizar():
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    caixas_ids = [int(x) for x in request.form.getlist("caixas_ids") if str(x).strip()]
    bancos_ids = [int(x) for x in request.form.getlist("bancos_ids") if str(x).strip()]

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=extras.DictCursor)

    # Monta filtros opcionais por contas
    filtros = []
    params_periodo = [data_inicio, data_fim]
    if caixas_ids:
        placeholders_cx = ",".join(["%s"] * len(caixas_ids))
        filtros.append(f"(conta_origem_tipo = 'caixa' AND conta_origem_id IN ({placeholders_cx}))")
        params_periodo.extend(caixas_ids)
    if bancos_ids:
        placeholders_bk = ",".join(["%s"] * len(bancos_ids))
        filtros.append(f"(conta_origem_tipo = 'banco' AND conta_origem_id IN ({placeholders_bk}))")
        params_periodo.extend(bancos_ids)
    where_extras = f" AND ({' OR '.join(filtros)})" if filtros else ""

    # Saldo inicial via Posições Diárias (dia anterior ao início)
    data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    dia_anterior = (data_inicio_dt - timedelta(days=1)).isoformat()
    pos_filtros = []
    params_pos = [dia_anterior]
    if caixas_ids:
        placeholders_cx2 = ",".join(["%s"] * len(caixas_ids))
        pos_filtros.append(f"(conta_tipo = 'caixa' AND conta_id IN ({placeholders_cx2}))")
        params_pos.extend(caixas_ids)
    if bancos_ids:
        placeholders_bk2 = ",".join(["%s"] * len(bancos_ids))
        pos_filtros.append(f"(conta_tipo = 'banco' AND conta_id IN ({placeholders_bk2}))")
        params_pos.extend(bancos_ids)
    where_pos = f" AND ({' OR '.join(pos_filtros)})" if pos_filtros else ""
    query_pos = (
        "SELECT COALESCE(SUM(saldo),0) AS saldo FROM ("
        "  SELECT DISTINCT ON (conta_tipo, conta_id) conta_tipo, conta_id, saldo"
        "    FROM posicao_diaria"
        "   WHERE data <= %s" + where_pos +
        "   ORDER BY conta_tipo, conta_id, data DESC"
        ") t"
    )
    cur.execute(query_pos, tuple(params_pos))
    saldo_inicial = Decimal(cur.fetchone()["saldo"]) if cur.rowcount is not None else Decimal("0")

    # Entradas e saídas por dia no período
    query_periodo = (
        "SELECT DATE(data_movimento) AS dia, "
        "COALESCE(SUM(CASE WHEN tipo = 'entrada' THEN valor ELSE 0 END),0) AS entradas, "
        "COALESCE(SUM(CASE WHEN tipo = 'saida' THEN valor ELSE 0 END),0)   AS saidas "
        "FROM movimento_financeiro "
        "WHERE data_movimento BETWEEN %s AND %s" + where_extras + " GROUP BY 1 ORDER BY 1"
    )
    cur.execute(query_periodo, tuple(params_periodo))
    rows = cur.fetchall()

    # Nome da empresa para cabeçalho
    cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
    empresa = cur.fetchone()

    cur.close()
    conn.close()

    # Monta linhas com saldo acumulado
    linhas = []
    saldo = saldo_inicial
    total_entradas = Decimal("0")
    total_saidas = Decimal("0")
    for r in rows:
        dia = r["dia"]
        ent = Decimal(r["entradas"]) or Decimal("0")
        sai = Decimal(r["saidas"]) or Decimal("0")
        saldo = saldo + ent - sai
        total_entradas += ent
        total_saidas += sai
        linhas.append({"dia": dia, "entradas": ent, "saidas": sai, "saldo": saldo})

    total_final = saldo

    # Formatações de período para exibição
    data_inicio_fmt = datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
    data_fim_fmt = datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
    periodo = f"Período: {data_inicio_fmt} a {data_fim_fmt}"

    return render_template(
        "relatorios/financeiro/fluxo_caixa.html",
        empresa=empresa["razao_social_nome"] if empresa else "",
        periodo=periodo,
        data_inicio=data_inicio,
        data_fim=data_fim,
        caixas_ids=caixas_ids,
        bancos_ids=bancos_ids,
        saldo_inicial=saldo_inicial,
        linhas=linhas,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        total_final=total_final,
    )

@app.route("/relatorios/gerencial")
@login_required
def relatorios_gerencial():
    return render_template("relatorios/gerencial/index.html")


# ------------------ DRE (Máscara & Relatório) ------------------


def _montar_arvore_nos(nos):
    """Monta uma árvore a partir de uma lista de nós (dicts)."""
    by_parent = {}
    for n in nos:
        by_parent.setdefault(n.get("parent_id"), []).append(n)
    for lst in by_parent.values():
        lst.sort(key=lambda x: (x.get("ordem") or 0, x.get("id") or 0))

    def attach(parent_id=None):
        out = []
        for n in by_parent.get(parent_id, []):
            n["filhos"] = attach(n["id"])
            out.append(n)
        return out

    return attach(None)


@app.route("/gerencial/dre/mascaras")
@login_required
@permission_required("Administracao Sistema", "Consultar")
def dre_mascaras_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM dre_mascaras ORDER BY ordem ASC, nome ASC")
    mascaras = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("gerencial/dre_mascaras/list.html", mascaras=mascaras)


@app.route("/gerencial/dre/mascaras/add", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Incluir")
def dre_mascaras_add():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        descricao = request.form.get("descricao")
        ativo = True if request.form.get("ativo") == "on" else False
        ordem = request.form.get("ordem", type=int)
        eh_formula = True if request.form.get("eh_formula") == "on" else False
        formula = request.form.get("formula")
        if not nome:
            flash("Informe o nome da máscara.", "danger")
            # Carrega tokens de máscaras existentes
            conn = get_db_connection()
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT id, nome FROM dre_mascaras ORDER BY ordem ASC, nome ASC")
            mascaras_tokens = cur.fetchall()
            cur.close()
            conn.close()
            return render_template("gerencial/dre_mascaras/add_edit.html", mascara={}, mascaras_tokens=mascaras_tokens)
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO dre_mascaras (nome, descricao, ordem, eh_formula, formula, ativo) VALUES (%s,%s,COALESCE(%s,0),%s,%s,%s) RETURNING id",
                (nome, descricao, ordem, eh_formula, formula, ativo),
            )
            mascara_id = cur.fetchone()[0]
            conn.commit()
            log_user_action(
                "Incluir",
                "Gerencial - DRE Máscaras",
                f"Máscara '{nome}' criada.",
                {
                    "mascara_id": mascara_id,
                    "eh_formula": eh_formula,
                    "ativo": ativo,
                },
            )
            flash("Máscara criada com sucesso!", "success")
            return redirect(url_for("dre_mascaras_list"))
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao criar máscara: {e}", "danger")
            # Carrega tokens de máscaras existentes
            try:
                cur2 = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                cur2.execute("SELECT id, nome FROM dre_mascaras ORDER BY ordem ASC, nome ASC")
                mascaras_tokens = cur2.fetchall()
                cur2.close()
            except Exception:
                mascaras_tokens = []
            return render_template("gerencial/dre_mascaras/add_edit.html", mascara={}, mascaras_tokens=mascaras_tokens)
        finally:
            cur.close()
            conn.close()
    # GET: Carrega tokens de máscaras existentes
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, nome FROM dre_mascaras ORDER BY ordem ASC, nome ASC")
    mascaras_tokens = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("gerencial/dre_mascaras/add_edit.html", mascara={}, mascaras_tokens=mascaras_tokens)


@app.route("/gerencial/dre/mascaras/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def dre_mascaras_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        descricao = request.form.get("descricao")
        ativo = True if request.form.get("ativo") == "on" else False
        ordem = request.form.get("ordem", type=int)
        eh_formula = True if request.form.get("eh_formula") == "on" else False
        formula = request.form.get("formula")
        if not nome:
            flash("Informe o nome da máscara.", "danger")
        else:
            try:
                cur.execute(
                    "UPDATE dre_mascaras SET nome=%s, descricao=%s, ordem=COALESCE(%s,0), eh_formula=%s, formula=%s, ativo=%s WHERE id=%s",
                    (nome, descricao, ordem, eh_formula, formula, ativo, id),
                )
                conn.commit()
                log_user_action(
                    "Editar",
                    "Gerencial - DRE Máscaras",
                    f"Máscara ID {id} atualizada.",
                    {
                        "mascara_id": id,
                        "eh_formula": eh_formula,
                        "ativo": ativo,
                    },
                )
                flash("Máscara atualizada com sucesso!", "success")
                return redirect(url_for("dre_mascaras_list"))
            except Exception as e:
                conn.rollback()
                flash(f"Erro ao atualizar máscara: {e}", "danger")
    cur.execute("SELECT * FROM dre_mascaras WHERE id=%s", (id,))
    mascara = cur.fetchone()
    # Carrega tokens de máscaras existentes (exceto a própria)
    cur.execute("SELECT id, nome FROM dre_mascaras WHERE id<>%s ORDER BY ordem ASC, nome ASC", (id,))
    mascaras_tokens = cur.fetchall()
    cur.close()
    conn.close()
    if not mascara:
        flash("Máscara não encontrada.", "danger")
        return redirect(url_for("dre_mascaras_list"))
    return render_template("gerencial/dre_mascaras/add_edit.html", mascara=mascara, mascaras_tokens=mascaras_tokens)


@app.route("/gerencial/dre/mascaras/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Excluir")
def dre_mascaras_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM dre_mascaras WHERE id=%s", (id,))
        deleted = cur.rowcount
        conn.commit()
        if deleted:
            log_user_action(
                "Excluir",
                "Gerencial - DRE Máscaras",
                f"Máscara ID {id} excluída.",
                {"mascara_id": id},
            )
        flash("Máscara excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir máscara: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("dre_mascaras_list"))


@app.route("/gerencial/dre/mascaras/<int:id>/builder", methods=["GET"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def dre_mascaras_builder(id):
    """Tela de montagem da máscara: gerencia nós e mapeamentos."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM dre_mascaras WHERE id=%s", (id,))
    mascara = cur.fetchone()
    if not mascara:
        cur.close()
        conn.close()
        flash("Máscara não encontrada.", "danger")
        return redirect(url_for("dre_mascaras_list"))

    cur.execute(
        "SELECT * FROM dre_nos WHERE mascara_id=%s ORDER BY parent_id NULLS FIRST, ordem ASC, id ASC",
        (id,),
    )
    nos = cur.fetchall()
    arvore = _montar_arvore_nos([dict(n) for n in nos])

    # Para formulários: pais possíveis (apenas grupos)
    cur.execute(
        "SELECT id, titulo FROM dre_nos WHERE mascara_id=%s AND tipo='grupo' ORDER BY ordem, titulo",
        (id,),
    )
    pais = cur.fetchall()

    # Dados para mapeamentos (caso um nó seja selecionado via query string)
    no_id = request.args.get("no_id", type=int)
    no = None
    receitas = despesas = []
    mapeadas_receitas = mapeadas_despesas = set()
    if no_id:
        cur.execute("SELECT * FROM dre_nos WHERE id=%s AND mascara_id=%s", (no_id, id))
        no = cur.fetchone()
        if no and no["tipo"] == "receita":
            cur.execute("SELECT id, descricao FROM receitas_cadastro ORDER BY descricao")
            receitas = cur.fetchall()
            cur.execute("SELECT receita_id FROM dre_no_receitas WHERE no_id=%s", (no_id,))
            mapeadas_receitas = {r[0] for r in cur.fetchall()}
        elif no and no["tipo"] == "despesa":
            cur.execute("SELECT id, descricao FROM despesas_cadastro ORDER BY descricao")
            despesas = cur.fetchall()
            cur.execute("SELECT despesa_id FROM dre_no_despesas WHERE no_id=%s", (no_id,))
            mapeadas_despesas = {r[0] for r in cur.fetchall()}

    cur.close()
    conn.close()

    return render_template(
        "gerencial/dre_mascaras/builder.html",
        mascara=mascara,
        arvore=arvore,
        pais=pais,
        no=no,
        receitas=receitas,
        despesas=despesas,
        mapeadas_receitas=mapeadas_receitas,
        mapeadas_despesas=mapeadas_despesas,
    )


@app.route("/gerencial/dre/mascaras/<int:id>/estrutura/delete", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Excluir")
def dre_mascaras_estrutura_delete(id):
    """Exclui toda a estrutura (nós) de uma máscara, mantendo a máscara."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM dre_nos WHERE mascara_id=%s", (id,))
        removidos = cur.rowcount
        conn.commit()
        if removidos:
            log_user_action(
                "Excluir",
                "Gerencial - DRE Máscaras",
                f"Estrutura da máscara ID {id} removida.",
                {"mascara_id": id, "nos_removidos": removidos},
            )
        flash("Estrutura da máscara excluída com sucesso.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir estrutura: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("dre_mascaras_builder", id=id))


@app.route("/gerencial/dre/nos/add", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def dre_nos_add():
    mascara_id = request.form.get("mascara_id", type=int)
    parent_id = request.form.get("parent_id", type=int)
    titulo = request.form.get("titulo", "").strip()
    tipo = request.form.get("tipo", "grupo")
    ordem = request.form.get("ordem", type=int)
    if not mascara_id or not titulo or tipo not in ("grupo", "receita", "despesa"):
        flash("Preencha os dados do novo item corretamente.", "danger")
        return redirect(url_for("dre_mascaras_builder", id=mascara_id))
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO dre_nos (mascara_id, parent_id, titulo, tipo, ordem) VALUES (%s,%s,%s,%s,COALESCE(%s,0)) RETURNING id",
            (mascara_id, parent_id, titulo, tipo, ordem),
        )
        no_id = cur.fetchone()[0]
        conn.commit()
        log_user_action(
            "Incluir",
            "Gerencial - DRE Máscaras",
            "Nó adicionado à estrutura do DRE.",
            {
                "mascara_id": mascara_id,
                "no_id": no_id,
                "tipo": tipo,
                "parent_id": parent_id,
            },
        )
        flash("Item adicionado.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao adicionar item: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("dre_mascaras_builder", id=mascara_id))


@app.route("/gerencial/dre/nos/<int:no_id>/delete", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Excluir")
def dre_nos_delete(no_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT mascara_id FROM dre_nos WHERE id=%s", (no_id,))
    row = cur.fetchone()
    mascara_id = row["mascara_id"] if row else None
    try:
        cur.execute("DELETE FROM dre_nos WHERE id=%s", (no_id,))
        removido = cur.rowcount
        conn.commit()
        if removido:
            log_user_action(
                "Excluir",
                "Gerencial - DRE Máscaras",
                f"Nó ID {no_id} removido.",
                {"mascara_id": mascara_id, "no_id": no_id},
            )
        flash("Item removido.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao remover item: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    if mascara_id:
        return redirect(url_for("dre_mascaras_builder", id=mascara_id))
    return redirect(url_for("dre_mascaras_list"))


@app.route("/gerencial/dre/nos/<int:no_id>/map", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def dre_nos_map(no_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT mascara_id, tipo FROM dre_nos WHERE id=%s", (no_id,))
    no = cur.fetchone()
    if not no:
        cur.close()
        conn.close()
        flash("Item não encontrado.", "danger")
        return redirect(url_for("dre_mascaras_list"))
    mascara_id = no["mascara_id"]
    selecionadas_ids = []
    try:
        if no["tipo"] == "receita":
            selecionadas = request.form.getlist("receitas")
            cur.execute("DELETE FROM dre_no_receitas WHERE no_id=%s", (no_id,))
            for rid in selecionadas:
                try:
                    rid_int = int(rid)
                except (TypeError, ValueError):
                    continue
                selecionadas_ids.append(rid_int)
                cur.execute(
                    "INSERT INTO dre_no_receitas (no_id, receita_id) VALUES (%s,%s)",
                    (no_id, rid_int),
                )
        elif no["tipo"] == "despesa":
            selecionadas = request.form.getlist("despesas")
            cur.execute("DELETE FROM dre_no_despesas WHERE no_id=%s", (no_id,))
            for did in selecionadas:
                try:
                    did_int = int(did)
                except (TypeError, ValueError):
                    continue
                selecionadas_ids.append(did_int)
                cur.execute(
                    "INSERT INTO dre_no_despesas (no_id, despesa_id) VALUES (%s,%s)",
                    (no_id, did_int),
                )
        conn.commit()
        log_user_action(
            "Editar",
            "Gerencial - DRE Máscaras",
            "Mapeamento de nó atualizado.",
            {
                "mascara_id": mascara_id,
                "no_id": no_id,
                "tipo": no["tipo"],
                "itens_mapeados": selecionadas_ids,
            },
        )
        flash("Mapeamentos atualizados.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao salvar mapeamentos: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("dre_mascaras_builder", id=mascara_id, no_id=no_id))


def _somar_leaf(cur, no_id, tipo, base, data_inicio, data_fim):
    total = Decimal("0")
    if tipo == "receita":
        if base == "caixa":
            # Contas a receber pagas
            cur.execute(
                """
                SELECT COALESCE(SUM(cr.valor_pago),0) AS total
                  FROM contas_a_receber cr
                  JOIN dre_no_receitas dmr ON dmr.receita_id = cr.receita_id AND dmr.no_id = %s
                 WHERE cr.data_pagamento IS NOT NULL
                   AND cr.data_pagamento BETWEEN %s AND %s
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")

            # Entradas diretas (Movimento Financeiro)
            cur.execute(
                """
                SELECT COALESCE(SUM(mf.valor),0) AS total
                  FROM movimento_financeiro mf
                  JOIN dre_no_receitas dmr ON dmr.receita_id = mf.receita_id AND dmr.no_id = %s
                 WHERE mf.tipo = 'entrada'
                   AND mf.data_movimento BETWEEN %s AND %s
                   AND (mf.documento IS NULL OR mf.documento NOT LIKE 'CR-%%')
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")
        else:  # competencia
            # Contas a receber por vencimento
            cur.execute(
                """
                SELECT COALESCE(SUM(cr.valor_previsto),0) AS total
                  FROM contas_a_receber cr
                  JOIN dre_no_receitas dmr ON dmr.receita_id = cr.receita_id AND dmr.no_id = %s
                 WHERE cr.data_vencimento BETWEEN %s AND %s
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")

            # Previsões diretas (Movimento Financeiro)
            cur.execute(
                """
                SELECT COALESCE(SUM(COALESCE(mf.valor_previsto, mf.valor)),0) AS total
                  FROM movimento_financeiro mf
                  JOIN dre_no_receitas dmr ON dmr.receita_id = mf.receita_id AND dmr.no_id = %s
                 WHERE mf.tipo = 'entrada'
                   AND mf.data_movimento BETWEEN %s AND %s
                   AND (mf.documento IS NULL OR mf.documento NOT LIKE 'CR-%%')
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")
    else:  # despesa
        if base == "caixa":
            # Contas a pagar pagas
            cur.execute(
                """
                SELECT COALESCE(SUM(cp.valor_pago),0) AS total
                  FROM contas_a_pagar cp
                  JOIN dre_no_despesas dmd ON dmd.despesa_id = cp.despesa_id AND dmd.no_id = %s
                 WHERE cp.data_pagamento IS NOT NULL
                   AND cp.data_pagamento BETWEEN %s AND %s
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")

            # Saídas diretas (Movimento Financeiro)
            cur.execute(
                """
                SELECT COALESCE(SUM(mf.valor),0) AS total
                  FROM movimento_financeiro mf
                  JOIN dre_no_despesas dmd ON dmd.despesa_id = mf.despesa_id AND dmd.no_id = %s
                 WHERE mf.tipo = 'saida'
                   AND mf.data_movimento BETWEEN %s AND %s
                   AND (mf.documento IS NULL OR mf.documento NOT LIKE 'CP-%%')
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")
        else:
            # Contas a pagar por vencimento
            cur.execute(
                """
                SELECT COALESCE(SUM(cp.valor_previsto),0) AS total
                  FROM contas_a_pagar cp
                  JOIN dre_no_despesas dmd ON dmd.despesa_id = cp.despesa_id AND dmd.no_id = %s
                 WHERE cp.data_vencimento BETWEEN %s AND %s
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")

            # Previsões diretas (Movimento Financeiro)
            cur.execute(
                """
                SELECT COALESCE(SUM(COALESCE(mf.valor_previsto, mf.valor)),0) AS total
                  FROM movimento_financeiro mf
                  JOIN dre_no_despesas dmd ON dmd.despesa_id = mf.despesa_id AND dmd.no_id = %s
                 WHERE mf.tipo = 'saida'
                   AND mf.data_movimento BETWEEN %s AND %s
                   AND (mf.documento IS NULL OR mf.documento NOT LIKE 'CP-%%')
                """,
                (no_id, data_inicio, data_fim),
            )
            row = cur.fetchone()
            total += Decimal(str(row["total"])) if row and row["total"] is not None else Decimal("0")

    return total


def _calcular_totais(cur, arvore, base, data_inicio, data_fim):
    """Calcula o total de cada nó da árvore e retorna a mesma estrutura com 'valor'."""
    total = Decimal("0")
    for n in arvore:
        if n["tipo"] == "grupo":
            n["filhos"], subtotal = _calcular_totais(cur, n.get("filhos", []), base, data_inicio, data_fim)
            n["valor"] = subtotal
        elif n["tipo"] in ("receita", "despesa"):
            soma = _somar_leaf(cur, n["id"], n["tipo"], base, data_inicio, data_fim)
            # Despesas negativas para facilitar leitura do resultado
            n["valor"] = soma if n["tipo"] == "receita" else -soma
        else:
            n["valor"] = Decimal("0")
        total += n["valor"]
    return arvore, total


def _prune_zero_nodes(nos):
    """Remove nós com valor 0 e sem filhos relevantes."""
    pruned = []
    for n in nos:
        filhos = _prune_zero_nodes(n.get("filhos", []))
        n["filhos"] = filhos
        valor = n.get("valor") or Decimal("0")
        if (valor != 0) or filhos:
            pruned.append(n)
    return pruned


@app.route("/relatorios/gerencial/dre", methods=["GET", "POST"])
@login_required
def relatorio_dre():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, nome, eh_formula, formula FROM dre_mascaras WHERE ativo=true ORDER BY ordem ASC, nome ASC")
    mascaras = cur.fetchall()

    # Gera lista de períodos (últimos 24 meses, incluindo mês atual)
    hoje = date.today()
    periodos_lista = []  # [{"value": "YYYY-MM", "label": "MM/YYYY"}, ...]
    ano = hoje.year
    mes = hoje.month
    for k in range(24):
        m = mes - k
        y = ano
        while m <= 0:
            m += 12
            y -= 1
        periodos_lista.append({"value": f"{y:04d}-{m:02d}", "label": f"{m:02d}/{y:04d}"})

    resultados = None
    base = (request.form.get("base") or "caixa").lower()
    periodos_sel = request.form.getlist("periodos")  # ["YYYY-MM", ...]
    hide_zeros = request.form.get("hide_zeros") in ("1", "on", "true", "True") or (request.method == "GET")

    comparativo = None
    if request.method == "POST" and periodos_sel:
        cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
        empresa = cur.fetchone()
        resultados = []  # lista por período
        # Map de alias por nome normalizado -> id
        def _norm(s):
            return re.sub(r"[^0-9a-zA-Z]+", "", (s or "").lower())
        alias_to_id = {_norm(m["nome"]): m["id"] for m in mascaras}
        for per in periodos_sel:
            try:
                y, m = per.split("-")
                y = int(y)
                m = int(m)
                di = date(y, m, 1)
                last_day = calendar.monthrange(y, m)[1]
                df = date(y, m, last_day)
            except Exception:
                continue

            grupo = {
                "periodo_value": per,
                "periodo_label": f"{m:02d}/{y:04d}",
                "data_inicio": di.strftime("%Y-%m-%d"),
                "data_fim": df.strftime("%Y-%m-%d"),
                "itens": [],  # por máscara
            }

            # Primeiro, calcula máscaras base (não fórmula) e guarda por id
            totals_by_id = {}
            item_by_id = {}
            for mask in mascaras:
                if not mask.get("eh_formula"):
                    cur.execute(
                        "SELECT * FROM dre_nos WHERE mascara_id=%s ORDER BY parent_id NULLS FIRST, ordem ASC, id ASC",
                        (mask["id"],),
                    )
                    nos = [dict(n) for n in cur.fetchall()]
                    arvore = _montar_arvore_nos(nos)
                    arvore_val, total = _calcular_totais(cur, arvore, base, di, df)
                    if hide_zeros:
                        arvore_val = _prune_zero_nodes(arvore_val)
                        _, total = _calcular_totais(cur, arvore_val, base, di, df)
                    totals_by_id[mask["id"]] = total
                    item_by_id[mask["id"]] = {
                        "mascara": mask,
                        "arvore": arvore_val,
                        "total": total,
                        "empresa": empresa["razao_social_nome"] if empresa else "",
                        "periodo": di.strftime("%d/%m/%Y") + " a " + df.strftime("%d/%m/%Y"),
                        "data_inicio": di.strftime("%Y-%m-%d"),
                        "data_fim": df.strftime("%Y-%m-%d"),
                    }

            # Depois, calcula máscaras por fórmula e guarda por id
            def _eval_formula(expr, totals_by_id, alias_to_id):
                if not expr:
                    return Decimal("0")
                # Substitui tokens #ID e nomes
                def repl_token(m):
                    tok = m.group(0)
                    if tok.startswith('#') and tok[1:].isdigit():
                        mid = int(tok[1:])
                        val = totals_by_id.get(mid, Decimal("0"))
                        return str(val)
                    norm = re.sub(r"[^0-9a-zA-Z]+", "", tok.lower())
                    mid = alias_to_id.get(norm)
                    if mid is not None:
                        val = totals_by_id.get(mid, Decimal("0"))
                        return str(val)
                    return tok  # mantém operadores/parênteses

                expr_num = re.sub(r"#\d+|[A-Za-zÀ-ÿ0-9_]+", repl_token, expr)

                # Avaliador seguro com Decimal
                tokens = re.findall(r"\d+(?:\.\d+)?|[()+\-*/]", expr_num)
                # Shunting-yard
                prec = {'+':1,'-':1,'*':2,'/':2}
                output = []
                ops = []
                for t in tokens:
                    if re.match(r"\d", t):
                        output.append(Decimal(t))
                    elif t in "+-*/":
                        while ops and ops[-1] in "+-*/" and prec[ops[-1]] >= prec[t]:
                            output.append(ops.pop())
                        ops.append(t)
                    elif t == '(':
                        ops.append(t)
                    elif t == ')':
                        while ops and ops[-1] != '(':
                            output.append(ops.pop())
                        if ops and ops[-1] == '(':
                            ops.pop()
                while ops:
                    output.append(ops.pop())
                # Avalia RPN
                stack = []
                for t in output:
                    if isinstance(t, str):
                        b = stack.pop() if stack else Decimal("0")
                        a = stack.pop() if stack else Decimal("0")
                        if t == '+': stack.append(a + b)
                        elif t == '-': stack.append(a - b)
                        elif t == '*': stack.append(a * b)
                        elif t == '/': stack.append(b and (a / b) or Decimal("0"))
                    else:
                        stack.append(t)
                return stack[-1] if stack else Decimal("0")

            for mask in mascaras:
                if mask.get("eh_formula"):
                    total = _eval_formula(mask.get("formula") or "", totals_by_id, alias_to_id)
                    item_by_id[mask["id"]] = {
                        "mascara": mask,
                        "arvore": [],
                        "total": total,
                        "empresa": empresa["razao_social_nome"] if empresa else "",
                        "periodo": di.strftime("%d/%m/%Y") + " a " + df.strftime("%d/%m/%Y"),
                        "data_inicio": di.strftime("%Y-%m-%d"),
                        "data_fim": df.strftime("%Y-%m-%d"),
                    }

            # Por fim, preenche itens na ordem cadastrada
            grupo["itens"] = []
            for mask in mascaras:
                it = item_by_id.get(mask["id"]) 
                if it:
                    grupo["itens"].append(it)

            resultados.append(grupo)

        # Monta estrutura comparativa (períodos em colunas)
        periodos_comp = [
            {
                "value": g["periodo_value"],
                "label": g["periodo_label"],
                "data_inicio": g.get("data_inicio"),
                "data_fim": g.get("data_fim"),
            }
            for g in resultados
        ]
        rows = []
        for m in mascaras:
            valores = []
            for g in resultados:
                it = next((it for it in g["itens"] if it["mascara"]["id"] == m["id"]), None)
                valores.append(it["total"] if it else Decimal("0"))
            variacoes = []
            for i, v in enumerate(valores):
                if i == 0:
                    variacoes.append(None)
                else:
                    prev = valores[i - 1]
                    try:
                        variacoes.append((v - prev) / (abs(prev) if prev != 0 else Decimal("1")))
                    except Exception:
                        variacoes.append(None)
            rows.append({"mascara": m, "valores": valores, "variacoes": variacoes})
        comparativo = {"periodos": periodos_comp, "rows": rows}

    cur.close()
    conn.close()

    return render_template(
        "relatorios/gerencial/dre.html",
        mascaras=mascaras,
        resultados=resultados,
        base=base,
        periodos_lista=periodos_lista,
        periodos_sel=periodos_sel,
        hide_zeros=hide_zeros,
        comparativo=comparativo,
    )


@app.route("/gerencial/dre/nos/reorder", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def dre_nos_reorder():
    """Atualiza ordem e pai de uma lista de nós após drag-and-drop."""
    data = request.get_json(silent=True) or {}
    mascara_id = data.get("mascara_id")
    parent_id = data.get("parent_id")
    ids = data.get("ids") or []

    # Normaliza parent_id
    if parent_id in ("", "null", None):
        parent_id = None
    try:
        mascara_id = int(mascara_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "mascara_id inválido"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    ids_processados = []
    try:
        for ordem, no_id in enumerate(ids):
            try:
                no_id_int = int(no_id)
            except (TypeError, ValueError):
                continue
            ids_processados.append(no_id_int)
            cur.execute(
                "UPDATE dre_nos SET parent_id=%s, ordem=%s WHERE id=%s AND mascara_id=%s",
                (parent_id, ordem, no_id_int, mascara_id),
            )
        conn.commit()
        log_user_action(
            "Editar",
            "Gerencial - DRE Máscaras",
            "Nós da estrutura foram reordenados.",
            {
                "mascara_id": mascara_id,
                "parent_id": parent_id,
                "ordem": ids_processados,
            },
        )
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/relatorios/gerencial/dre/pdf", methods=["POST"])
@login_required
def relatorio_dre_pdf():
    """Gera PDF do DRE com base nos parâmetros informados."""
    base = (request.form.get("base") or "caixa").lower()
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    mascara_id = request.form.get("mascara_id", type=int)
    hide_zeros = request.form.get("hide_zeros") in ("1", "on", "true", "True")

    if not (mascara_id and data_inicio and data_fim):
        flash("Parâmetros inválidos para gerar o PDF.", "danger")
        return redirect(url_for("relatorio_dre"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Busca máscara
    cur.execute("SELECT * FROM dre_mascaras WHERE id=%s", (mascara_id,))
    mascara_sel = cur.fetchone()
    # Se máscara for por fórmula, calcula total pela fórmula; caso contrário, usa estrutura
    arvore_val = []
    if mascara_sel and mascara_sel.get('eh_formula'):
        # Carrega todas as máscaras ativas para mapear nomes/ids
        cur.execute("SELECT id, nome FROM dre_mascaras WHERE ativo=true ORDER BY ordem ASC, nome ASC")
        all_masks = cur.fetchall()
        alias_to_id = {re.sub(r"[^0-9a-zA-Z]+", "", (m['nome'] or '').lower()): m['id'] for m in all_masks}
        # Calcula totais das máscaras base
        totals_by_id = {}
        for m in all_masks:
            if m['id'] == mascara_id:
                continue
            cur.execute(
                "SELECT * FROM dre_nos WHERE mascara_id=%s ORDER BY parent_id NULLS FIRST, ordem ASC, id ASC",
                (m['id'],),
            )
            nos_m = [dict(n) for n in cur.fetchall()]
            arv = _montar_arvore_nos(nos_m)
            arv_val, tot = _calcular_totais(cur, arv, base, data_inicio, data_fim)
            if hide_zeros:
                arv_val = _prune_zero_nodes(arv_val)
                _, tot = _calcular_totais(cur, arv_val, base, data_inicio, data_fim)
            totals_by_id[m['id']] = tot

        def _eval_formula(expr, totals_by_id, alias_to_id):
            if not expr:
                return Decimal('0')
            def repl_token(mm):
                tok = mm.group(0)
                if tok.startswith('#') and tok[1:].isdigit():
                    mid = int(tok[1:])
                    return str(totals_by_id.get(mid, Decimal('0')))
                norm = re.sub(r"[^0-9a-zA-Z]+", "", tok.lower())
                mid = alias_to_id.get(norm)
                if mid is not None:
                    return str(totals_by_id.get(mid, Decimal('0')))
                return tok
            expr_num = re.sub(r"#\d+|[A-Za-zÀ-ÿ0-9_]+", repl_token, mascara_sel.get('formula') or '')
            tokens = re.findall(r"\d+(?:\.\d+)?|[()+\-*/]", expr_num)
            prec = {'+':1,'-':1,'*':2,'/':2}
            output, ops = [], []
            for t in tokens:
                if re.match(r"\d", t):
                    output.append(Decimal(t))
                elif t in "+-*/":
                    while ops and ops[-1] in "+-*/" and prec[ops[-1]] >= prec[t]:
                        output.append(ops.pop())
                    ops.append(t)
                elif t == '(':
                    ops.append(t)
                elif t == ')':
                    while ops and ops[-1] != '(':
                        output.append(ops.pop())
                    if ops and ops[-1] == '(':
                        ops.pop()
            while ops:
                output.append(ops.pop())
            stack = []
            for t in output:
                if isinstance(t, str):
                    b = stack.pop() if stack else Decimal('0')
                    a = stack.pop() if stack else Decimal('0')
                    if t == '+': stack.append(a + b)
                    elif t == '-': stack.append(a - b)
                    elif t == '*': stack.append(a * b)
                    elif t == '/': stack.append(b and (a / b) or Decimal('0'))
                else:
                    stack.append(t)
            return stack[-1] if stack else Decimal('0')

        total = _eval_formula(mascara_sel.get('formula') or '', totals_by_id, alias_to_id)
    else:
        cur.execute(
            "SELECT * FROM dre_nos WHERE mascara_id=%s ORDER BY parent_id NULLS FIRST, ordem ASC, id ASC",
            (mascara_id,),
        )
        nos = [dict(n) for n in cur.fetchall()]
        arvore = _montar_arvore_nos(nos)
        arvore_val, total = _calcular_totais(cur, arvore, base, data_inicio, data_fim)
        if hide_zeros:
            arvore_val = _prune_zero_nodes(arvore_val)
            _, total = _calcular_totais(cur, arvore_val, base, data_inicio, data_fim)
    cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
    empresa = cur.fetchone()
    cur.close()
    conn.close()

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width, 6, "Demonstração do Resultado (DRE)", 0, 1, "C")
            self.set_font("Arial", "", 10)
            empresa_nome = empresa["razao_social_nome"] if empresa else ""
            periodo = (
                datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
                + " a "
                + datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
            )
            subtitulo = f"{empresa_nome} | {periodo}"
            self.cell(page_width, 5, subtitulo, 0, 1, "C")
            if mascara_sel:
                self.cell(page_width, 5, f"Máscara: {mascara_sel['nome']}", 0, 1, "C")
            self.ln(2)

    pdf = PDF()
    pdf.set_auto_page_break(True, 15)
    pdf.add_page()

    def render_lines(nos, nivel=0):
        indent = 5 * nivel
        for n in nos:
            # Definições de layout
            left_margin = pdf.l_margin + indent
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            value_col_w = 45
            title_w = page_width - value_col_w - indent

            # Fonte por tipo
            is_group = (n.get("tipo") == "grupo")
            pdf.set_font("Arial", "B" if is_group else "", 10)

            # Título
            # Estilos para grupo (subtotais): sombreamento + borda
            fill = False
            border = 0
            if is_group:
                pdf.set_fill_color(240, 240, 240)
                fill = True
                border = "TB"

            # Título
            pdf.set_x(left_margin)
            pdf.cell(title_w, 6, str(n.get("titulo") or ""), border, 0, "L", fill)

            # Valor
            valor_fmt = format_currency(n.get("valor") or 0)
            pdf.cell(value_col_w, 6, valor_fmt, border, 1, "R", fill)

            # Filhos
            filhos = n.get("filhos") or []
            if filhos:
                render_lines(filhos, nivel + 1)

    render_lines(arvore_val, 0)

    # Total
    pdf.ln(2)
    pdf.set_font("Arial", "B", 11)
    page_width = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(page_width - 45, 7, "Total do Período", "TB", 0, "L", True)
    pdf.cell(45, 7, format_currency(total), "TB", 1, "R", True)

    output = pdf.output(dest="S").encode("latin1", "ignore")
    filename = f"dre_{mascara_sel['nome'] if mascara_sel else 'mascara'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(io.BytesIO(output), mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/relatorios/gerencial/dre/pdf_full", methods=["POST"])
@login_required
def relatorio_dre_pdf_full():
    """Gera PDF completo do DRE (todas as máscaras ativas) para um período."""
    base = (request.form.get("base") or "caixa").lower()
    data_inicio = request.form.get("data_inicio")
    data_fim = request.form.get("data_fim")
    hide_zeros = request.form.get("hide_zeros") in ("1", "on", "true", "True")

    if not (data_inicio and data_fim):
        flash("Parâmetros inválidos para gerar o PDF.", "danger")
        return redirect(url_for("relatorio_dre"))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Máscaras ativas ordenadas
    cur.execute("SELECT id, nome, eh_formula, formula FROM dre_mascaras WHERE ativo=true ORDER BY ordem ASC, nome ASC")
    mascaras = cur.fetchall()

    # Empresa
    cur.execute("SELECT razao_social_nome FROM empresa_licenciada ORDER BY id LIMIT 1")
    empresa = cur.fetchone()

    # Calcula itens por máscara
    resultados = []
    alias_to_id = {re.sub(r"[^0-9a-zA-Z]+", "", (m["nome"] or "").lower()): m["id"] for m in mascaras}
    totals_by_id = {}
    item_by_id = {}

    # Base (não fórmula)
    for mask in mascaras:
        if not mask.get("eh_formula"):
            cur.execute(
                "SELECT * FROM dre_nos WHERE mascara_id=%s ORDER BY parent_id NULLS FIRST, ordem ASC, id ASC",
                (mask["id"],),
            )
            nos = [dict(n) for n in cur.fetchall()]
            arvore = _montar_arvore_nos(nos)
            arvore_val, total = _calcular_totais(cur, arvore, base, data_inicio, data_fim)
            if hide_zeros:
                arvore_val = _prune_zero_nodes(arvore_val)
                _, total = _calcular_totais(cur, arvore_val, base, data_inicio, data_fim)
            totals_by_id[mask["id"]] = total
            item_by_id[mask["id"]] = {"mascara": mask, "arvore": arvore_val, "total": total}

    # Fórmula
    def _eval_formula(expr):
        if not expr:
            return Decimal("0")
        def repl_token(mm):
            tok = mm.group(0)
            if tok.startswith('#') and tok[1:].isdigit():
                mid = int(tok[1:])
                return str(totals_by_id.get(mid, Decimal('0')))
            norm = re.sub(r"[^0-9a-zA-Z]+", "", tok.lower())
            mid = alias_to_id.get(norm)
            if mid is not None:
                return str(totals_by_id.get(mid, Decimal('0')))
            return tok
        expr_num = re.sub(r"#\d+|[A-Za-zÀ-ÿ0-9_]+", repl_token, expr)
        tokens = re.findall(r"\d+(?:\.\d+)?|[()+\-*/]", expr_num)
        prec = {'+':1,'-':1,'*':2,'/':2}
        output, ops = [], []
        for t in tokens:
            if re.match(r"\d", t):
                output.append(Decimal(t))
            elif t in "+-*/":
                while ops and ops[-1] in "+-*/" and prec[ops[-1]] >= prec[t]:
                    output.append(ops.pop())
                ops.append(t)
            elif t == '(':
                ops.append(t)
            elif t == ')':
                while ops and ops[-1] != '(':
                    output.append(ops.pop())
                if ops and ops[-1] == '(':
                    ops.pop()
        while ops:
            output.append(ops.pop())
        stack = []
        for t in output:
            if isinstance(t, str):
                b = stack.pop() if stack else Decimal('0')
                a = stack.pop() if stack else Decimal('0')
                if t == '+': stack.append(a + b)
                elif t == '-': stack.append(a - b)
                elif t == '*': stack.append(a * b)
                elif t == '/': stack.append(b and (a / b) or Decimal('0'))
            else:
                stack.append(t)
        return stack[-1] if stack else Decimal('0')

    for mask in mascaras:
        if mask.get("eh_formula"):
            total = _eval_formula(mask.get("formula") or "")
            item_by_id[mask["id"]] = {"mascara": mask, "arvore": [], "total": total}

    # Ordena por ordem e compõe lista
    for mask in mascaras:
        it = item_by_id.get(mask["id"])
        if it:
            resultados.append(it)

    # Empresa/Período já capturados
    cur.close()
    conn.close()

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            page_width = self.w - self.l_margin - self.r_margin
            self.cell(page_width, 6, "Demonstração do Resultado (DRE)", 0, 1, "C")
            self.set_font("Arial", "", 10)
            empresa_nome = empresa["razao_social_nome"] if empresa else ""
            periodo = (
                datetime.strptime(data_inicio, "%Y-%m-%d").strftime("%d/%m/%Y")
                + " a "
                + datetime.strptime(data_fim, "%Y-%m-%d").strftime("%d/%m/%Y")
            )
            self.cell(page_width, 5, f"{empresa_nome} | {periodo}", 0, 1, "C")
            self.ln(2)

    pdf = PDF()
    pdf.set_auto_page_break(True, 15)
    pdf.add_page()

    def render_lines(nos, nivel=0):
        indent = 5 * nivel
        for n in nos:
            left_margin = pdf.l_margin + indent
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            value_col_w = 45
            title_w = page_width - value_col_w - indent
            is_group = (n.get("tipo") == "grupo")
            pdf.set_font("Arial", "B" if is_group else "", 10)
            fill = False
            border = 0
            pdf.set_x(left_margin)
            pdf.cell(title_w, 6, str(n.get("titulo") or ""), border, 0, "L", fill)
            valor_fmt = format_currency(n.get("valor") or 0)
            pdf.cell(value_col_w, 6, valor_fmt, border, 1, "R", fill)
            filhos = n.get("filhos") or []
            if filhos:
                render_lines(filhos, nivel + 1)

    # Conteúdo por máscara
    page_width = pdf.w - pdf.l_margin - pdf.r_margin
    value_col_w = 45
    for item in resultados:
        mask = item["mascara"]
        total = item["total"]
        arvore_val = item["arvore"]
        # Cabeçalho da máscara com total
        pdf.set_font("Arial", "B", 11)
        pdf.cell(page_width - value_col_w, 7, str(mask["nome"]), 0, 0, "L")
        pdf.cell(value_col_w, 7, format_currency(total), 0, 1, "R")
        pdf.set_font("Arial", "", 10)
        if mask.get("eh_formula"):
            formula = mask.get("formula") or ""
            if formula:
                pdf.cell(page_width, 6, f"Fórmula: {formula}", 0, 1, "L")
        else:
            render_lines(arvore_val, 0)
        # separador
        pdf.ln(2)

    output = pdf.output(dest="S").encode("latin1", "ignore")
    filename = f"dre_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(io.BytesIO(output), mimetype="application/pdf", as_attachment=True, download_name=filename)

# --- Módulo de Administração do Sistema ---
@app.route("/admin/backup", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Incluir") # Permissão para gerar backup
def backup_db():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == "POST":
        backup_filename = request.form.get(
            "backup_filename",
            f"imobiliaria_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
        )
        backup_path = os.path.join(
            app.config["UPLOAD_FOLDER"], "backups", backup_filename
        )
        
        # Garante que o diretório de backups exista
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)

        try:
            # Extrai informações do DATABASE_URL para pg_dump
            # Ex: postgresql://user:password@host:port/dbname
            db_url_parts = DATABASE_URL.split("://")[1].split("@")
            user_pass = db_url_parts[0].split(":")
            host_port_db = db_url_parts[1].split("/")
            
            db_user = user_pass[0]
            db_password = user_pass[1] if len(user_pass) > 1 else ""
            db_host = host_port_db[0].split(":")[0]
            db_port = (
                host_port_db[0].split(":")[1]
                if len(host_port_db[0].split(":")) > 1
                else "5433"
            )
            db_name = host_port_db[1]

            # Comando pg_dump
            # Atenção: O pg_dump precisa estar no PATH do sistema onde o Flask está rodando.
            # Em produção, considere um método mais robusto e seguro para backups.
            command = [
                "pg_dump",
                "-h",
                db_host,
                "-p",
                db_port,
                "-U",
                db_user,
                "-F",
                "p",  # Formato plain-text SQL
                "-f",
                backup_path,
                db_name,
            ]
            
            # Define a variável de ambiente PGPASSWORD para pg_dump
            env = os.environ.copy()
            env["PGPASSWORD"] = db_password

            # Executa o comando
            result = subprocess.run(command, capture_output=True, text=True, env=env)

            if result.returncode == 0:
                status = "Sucesso"
                message = "Backup gerado com sucesso!"
                flash(message, "success")
            else:
                status = "Falha"
                message = f"Erro ao gerar backup: {result.stderr}"
                flash(message, "danger")
            
            # Registra no histórico
            cur.execute(
                "INSERT INTO historico_backups (data_backup, nome_arquivo, caminho_arquivo, status_backup, observacao, usuario_id) VALUES (%s, %s, %s, %s, %s, %s)",
                (
                    datetime.now(),
                    backup_filename,
                    backup_path,
                    status,
                    message,
                    session["user_id"],
                ),
            )
            conn.commit()

        except Exception as e:
            conn.rollback()
            status = "Falha"
            message = f"Erro inesperado ao gerar backup: {e}"
            flash(message, "danger")
            cur.execute(
                "INSERT INTO historico_backups (data_backup, nome_arquivo, caminho_arquivo, status_backup, observacao, usuario_id) VALUES (%s, %s, %s, %s, %s, %s)",
                (
                    datetime.now(),
                    backup_filename,
                    backup_path,
                    status,
                    message,
                    session["user_id"],
                ),
            )
            conn.commit()
        
    # Carrega histórico de backups
    cur.execute("SELECT * FROM historico_backups ORDER BY data_backup DESC")
    historico = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admin/backup.html", historico=historico)


@app.route("/admin/backup/restore", methods=["POST"])
@login_required
@permission_required(
    "Administracao Sistema", "Bloquear"
)  # Ação de restaurar é mais restritiva
def restore_db():
    if request.method == "POST":
        backup_id = request.form.get("backup_id")
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute(
            "SELECT caminho_arquivo FROM historico_backups WHERE id = %s", (backup_id,)
        )
        backup_record = cur.fetchone()
        cur.close()
        conn.close()

        if not backup_record:
            flash("Backup não encontrado.", "danger")
            return redirect(url_for("backup_db"))

        backup_path = backup_record["caminho_arquivo"]

        try:
            # Extrai informações do DATABASE_URL para psql
            db_url_parts = DATABASE_URL.split("://")[1].split("@")
            user_pass = db_url_parts[0].split(":")
            host_port_db = db_url_parts[1].split("/")
            
            db_user = user_pass[0]
            db_password = user_pass[1] if len(user_pass) > 1 else ""
            db_host = host_port_db[0].split(":")[0]
            db_port = (
                host_port_db[0].split(":")[1]
                if len(host_port_db[0].split(":")) > 1
                else "5433"
            )
            db_name = host_port_db[1]

            # Comando psql para restaurar
            # Atenção: Isso irá apagar e recriar o banco de dados. Tenha certeza do que está fazendo.
            # Para restaurar, o banco de dados não pode ter conexões ativas.
            # Em um cenário real, você precisaria de um script mais complexo para desconectar usuários,
            # dropar o banco, recriar e então restaurar.
            # Este é um exemplo simplificado.
            command = [
                "psql",
                "-h",
                db_host,
                "-p",
                db_port,
                "-U",
                db_user,
                "-d",
                db_name,
                "-f",
                backup_path,
            ]

            env = os.environ.copy()
            env["PGPASSWORD"] = db_password

            result = subprocess.run(command, capture_output=True, text=True, env=env)

            if result.returncode == 0:
                flash("Banco de dados restaurado com sucesso!", "success")
            else:
                flash(f"Erro ao restaurar banco de dados: {result.stderr}", "danger")

        except Exception as e:
            flash(f"Erro inesperado ao restaurar banco de dados: {e}", "danger")
    
    return redirect(url_for("backup_db"))


@app.route("/usuarios")
@login_required
@permission_required("Administracao Sistema", "Consultar")
def usuarios_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM usuarios ORDER BY id")
    usuarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admin/usuarios/list.html", usuarios=usuarios)


@app.route("/usuarios/add", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Incluir")
def usuarios_add():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        tipo = request.form.get("tipo_usuario", "Operador")
        status = request.form.get("status", "Ativo")

        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO usuarios (nome_usuario, senha_hash, tipo_usuario, status) VALUES (%s, %s, %s, %s)",
                (username, hashed_password, tipo, status),
            )
            conn.commit()
            flash("Usuário cadastrado com sucesso!", "success")
            return redirect(url_for("usuarios_list"))
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            flash("Nome de usuário já existe.", "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar usuário: {e}", "danger")
        finally:
            cur.close()
            conn.close()
    return render_template("admin/usuarios/add_list.html", usuario={})


@app.route("/usuarios/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def usuarios_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        username = request.form["username"]
        password = request.form.get("password")
        tipo = request.form.get("tipo_usuario")
        status = request.form.get("status")
        try:
            if password:
                hashed_password = generate_password_hash(password)
                cur.execute(
                    "UPDATE usuarios SET nome_usuario = %s, senha_hash = %s, tipo_usuario = %s, status = %s WHERE id = %s",
                    (username, hashed_password, tipo, status, id),
                )
            else:
                cur.execute(
                    "UPDATE usuarios SET nome_usuario = %s, tipo_usuario = %s, status = %s WHERE id = %s",
                    (username, tipo, status, id),
                )
            conn.commit()
            flash("Usuário atualizado com sucesso!", "success")
            return redirect(url_for("usuarios_list"))
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            flash("Nome de usuário já existe.", "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar usuário: {e}", "danger")
    cur.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
    usuario = cur.fetchone()
    cur.close()
    conn.close()
    if usuario is None:
        flash("Usuário não encontrado.", "danger")
        return redirect(url_for("usuarios_list"))
    return render_template("admin/usuarios/add_list.html", usuario=usuario)


@app.route("/usuarios/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Excluir")
def usuarios_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM usuarios WHERE id = %s", (id,))
        conn.commit()
        flash("Usuário excluído com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir usuário: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("usuarios_list"))


@app.route("/usuarios/permissoes/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def usuarios_permissoes(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        cur.execute("DELETE FROM permissoes WHERE usuario_id = %s", (id,))
        for module in MODULES:
            for action in ACTIONS:
                field = f"{module}:{action}"
                if request.form.get(field):
                    cur.execute(
                        "INSERT INTO permissoes (usuario_id, modulo, acao) VALUES (%s, %s, %s)",
                        (id, module, action),
                    )
        conn.commit()
        flash("Permissões atualizadas com sucesso!", "success")

    cur.execute("SELECT modulo, acao FROM permissoes WHERE usuario_id = %s", (id,))
    existing = {(row["modulo"], row["acao"]) for row in cur.fetchall()}
    cur.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
    usuario = cur.fetchone()
    cur.close()
    conn.close()
    if usuario is None:
        flash("Usuário não encontrado.", "danger")
        return redirect(url_for("usuarios_list"))
    return render_template(
        "admin/usuarios/permissoes.html",
        usuario=usuario,
        existing=existing,
        modules=MODULES,
        actions=ACTIONS,
    )


@app.route("/empresa")
@login_required
@permission_required("Administracao Sistema", "Consultar")
def empresa_list():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    search_query = request.args.get("search", "")
    if search_query:
        cur.execute(
            """
            SELECT * FROM empresa_licenciada
            WHERE documento ILIKE %s OR razao_social_nome ILIKE %s OR nome_fantasia ILIKE %s
            ORDER BY data_cadastro DESC
        """,
            (f"%{search_query}%", f"%{search_query}%", f"%{search_query}%"),
        )
    else:
        cur.execute("SELECT * FROM empresa_licenciada ORDER BY data_cadastro DESC")
    empresas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template(
        "admin/empresa/list.html", empresas=empresas, search_query=search_query
    )


@app.route("/empresa/add", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Incluir")
def empresa_add():
    if request.method == "POST":
        try:
            documento = (
                request.form["documento"].replace(".", "").replace("/", "").replace("-", "")
            )
            razao_social_nome = request.form["razao_social_nome"]
            nome_fantasia = request.form.get("nome_fantasia")
            endereco = request.form.get("endereco")
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep", "").replace("-", "")
            telefone = request.form.get("telefone")
            observacao = request.form.get("observacao")
            status = request.form["status"]

            if len(documento) == 11 and not documento.isdigit():
                flash("CPF inválido. Deve conter apenas números.", "danger")
                return render_template("admin/empresa/add_list.html", empresa=request.form)
            elif len(documento) == 14 and not documento.isdigit():
                flash("CNPJ inválido. Deve conter apenas números.", "danger")
                return render_template("admin/empresa/add_list.html", empresa=request.form)
            elif len(documento) != 11 and len(documento) != 14:
                flash(
                    "Documento deve ser um CPF (11 dígitos) ou CNPJ (14 dígitos).",
                    "danger",
                )
                return render_template("admin/empresa/add_list.html", empresa=request.form)

            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO empresa_licenciada (documento, razao_social_nome, nome_fantasia, endereco, bairro, cidade, estado, cep, telefone, observacao, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    documento,
                    razao_social_nome,
                    nome_fantasia,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    telefone,
                    observacao,
                    status,
                ),
            )
            conn.commit()
            cur.close()
            conn.close()
            flash("Empresa cadastrada com sucesso!", "success")
            return redirect(url_for("empresa_list"))
        except psycopg2.errors.UniqueViolation:
            flash("Erro: Documento (CNPJ/CPF) já cadastrado.", "danger")
            conn.rollback()
            return render_template("admin/empresa/add_list.html", empresa=request.form)
        except Exception as e:
            flash(f"Erro ao cadastrar empresa: {e}", "danger")
            return render_template("admin/empresa/add_list.html", empresa=request.form)
    return render_template("admin/empresa/add_list.html", empresa={})


@app.route("/empresa/edit/<int:id>", methods=["GET", "POST"])
@login_required
@permission_required("Administracao Sistema", "Editar")
def empresa_edit(id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == "POST":
        try:
            documento = (
                request.form["documento"].replace(".", "").replace("/", "").replace("-", "")
            )
            razao_social_nome = request.form["razao_social_nome"]
            nome_fantasia = request.form.get("nome_fantasia")
            endereco = request.form.get("endereco")
            bairro = request.form.get("bairro")
            cidade = request.form.get("cidade")
            estado = request.form.get("estado")
            cep = request.form.get("cep", "").replace("-", "")
            telefone = request.form.get("telefone")
            observacao = request.form.get("observacao")
            status = request.form["status"]

            if len(documento) == 11 and not documento.isdigit():
                flash("CPF inválido. Deve conter apenas números.", "danger")
                cur.execute("SELECT * FROM empresa_licenciada WHERE id = %s", (id,))
                empresa = cur.fetchone()
                return render_template("admin/empresa/add_list.html", empresa=empresa)
            elif len(documento) == 14 and not documento.isdigit():
                flash("CNPJ inválido. Deve conter apenas números.", "danger")
                cur.execute("SELECT * FROM empresa_licenciada WHERE id = %s", (id,))
                empresa = cur.fetchone()
                return render_template("admin/empresa/add_list.html", empresa=empresa)
            elif len(documento) != 11 and len(documento) != 14:
                flash(
                    "Documento deve ser um CPF (11 dígitos) ou CNPJ (14 dígitos).",
                    "danger",
                )
                cur.execute("SELECT * FROM empresa_licenciada WHERE id = %s", (id,))
                empresa = cur.fetchone()
                return render_template("admin/empresa/add_list.html", empresa=empresa)

            cur.execute(
                """
                UPDATE empresa_licenciada
                SET documento = %s, razao_social_nome = %s, nome_fantasia = %s, endereco = %s, bairro = %s, cidade = %s, estado = %s, cep = %s, telefone = %s, observacao = %s, status = %s
                WHERE id = %s
                """,
                (
                    documento,
                    razao_social_nome,
                    nome_fantasia,
                    endereco,
                    bairro,
                    cidade,
                    estado,
                    cep,
                    telefone,
                    observacao,
                    status,
                    id,
                ),
            )
            conn.commit()
            flash("Empresa atualizada com sucesso!", "success")
            return redirect(url_for("empresa_list"))
        except psycopg2.errors.UniqueViolation:
            flash("Erro: Documento (CNPJ/CPF) já cadastrado para outra empresa.", "danger")
            conn.rollback()
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar empresa: {e}", "danger")
        cur.execute("SELECT * FROM empresa_licenciada WHERE id = %s", (id,))
        empresa = cur.fetchone()
        cur.close()
        conn.close()
        if empresa is None:
            flash("Empresa não encontrada.", "danger")
            return redirect(url_for("empresa_list"))
        return render_template("admin/empresa/add_list.html", empresa=empresa)

    cur.execute("SELECT * FROM empresa_licenciada WHERE id = %s", (id,))
    empresa = cur.fetchone()
    cur.close()
    conn.close()
    if empresa is None:
        flash("Empresa não encontrada.", "danger")
        return redirect(url_for("empresa_list"))
    return render_template("admin/empresa/add_list.html", empresa=empresa)


@app.route("/empresa/delete/<int:id>", methods=["POST"])
@login_required
@permission_required("Administracao Sistema", "Excluir")
def empresa_delete(id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM empresa_licenciada WHERE id = %s", (id,))
        conn.commit()
        flash("Empresa excluída com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir empresa: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("empresa_list"))

# --- Execução da Aplicação ---

if __name__ == "__main__":
    # Ao definir host="0.0.0.0" o Flask escuta em todas as interfaces de
    # rede, permitindo acesso por outras máquinas da rede local.
    # Altere debug para False em produção.
    app.run(host="0.0.0.0", port=5000, debug=True)

